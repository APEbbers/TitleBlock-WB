# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/

import FreeCAD as App
import Standard_Functions_TitleBlock as Standard_Functions
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


# Define the translation
translate = App.Qt.translate


def createExcel():
    import Settings
    from Settings import preferences
    from Settings import IMPORT_SETTINGS_XL
    from Settings import ENABLE_DEBUG

    # Create a workbook and activate the first sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "TitleBlockData"
    preferences.SetString("SheetName", "TitleBlockData")
    preferences.SetString("StartCell", "A1")

    # Get the startcell and the next cells
    StartCell = str("A1")
    TopRow = int(StartCell[1:])
    PropCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    IncreaseCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    MultiplierCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    RemarkCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        message = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(message)

    # Set the headers
    ws[StartCell].value = "Property Name"
    ws[PropCell].value = "Property Value"
    ws[IncreaseCell].value = "Increase value"
    ws[MultiplierCell].value = "Factor"
    ws[RemarkCell].value = "Remarks"

    # region Format the settings with the values as a Table
    #
    # Define the the last cell
    EndCell = "E10"

    # Define the table
    NumberOfSheets = str(len(wb.sheetnames))
    tab = Table(
        displayName="TitleBlockTable_" + NumberOfSheets,
        ref=f"{StartCell}:{EndCell}",
    )

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    # add the style to the table
    tab.tableStyleInfo = style

    # add the table to the worksheet
    ws.add_table(tab)
    # endregion

    # Make the columns to autofit the date
    for col in ws.columns:
        SetLen = 0
        column = col[0].column_letter  # Get the column name

        for cell in col:
            if len(str(cell.value)) > SetLen:
                SetLen = len(str(cell.value))

        set_col_width = SetLen + 5
        # Setting the column width
        ws.column_dimensions[column].width = set_col_width

    # Save the excel file in a folder of your choosing
    Filter = [
        ("Excel", "*.xlsx"),
    ]
    FileName = Standard_Functions.SaveDialog(Filter)
    if FileName is not None:
        # Save the workbook
        wb.save(str(FileName))
        # Close the workbook
        wb.close()
        # Update the preferences
        preferences.SetString("ExternalFile", FileName)

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        Settings.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {ws.title}",
    )
    Standard_Functions.Mbox(text=message, title="TitleBlock Workbench", style=0)
