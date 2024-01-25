# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/

import FreeCAD as App
import Standard_Functions_TitleBlock as Standard_Functions
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# Get the settings
import SettingsTB
from SettingsTB import preferences
from SettingsTB import IMPORT_SETTINGS_XL
from SettingsTB import ENABLE_DEBUG
from SettingsTB import SPREADSHEET_COLUMNFONTSTYLE_UNDERLINE
from SettingsTB import SPREADSHEET_COLUMNFONTSTYLE_ITALIC
from SettingsTB import SPREADSHEET_COLUMNFONTSTYLE_BOLD
from SettingsTB import SPREADSHEET_TABLEFONTSTYLE_UNDERLINE
from SettingsTB import SPREADSHEET_TABLEFONTSTYLE_ITALIC
from SettingsTB import SPREADSHEET_TABLEFONTSTYLE_BOLD
from SettingsTB import SPREADSHEET_TABLEFOREGROUND
from SettingsTB import SPREADSHEET_TABLEBACKGROUND_2
from SettingsTB import SPREADSHEET_TABLEBACKGROUND_1
from SettingsTB import SPREADSHEET_HEADERFONTSTYLE_UNDERLINE
from SettingsTB import SPREADSHEET_HEADERFONTSTYLE_ITALIC
from SettingsTB import SPREADSHEET_HEADERFONTSTYLE_BOLD
from SettingsTB import SPREADSHEET_HEADERFOREGROUND
from SettingsTB import SPREADSHEET_HEADERBACKGROUND
from SettingsTB import AUTOFIT_FACTOR

# Define the translation
translate = App.Qt.translate


def CcreateExcel():
    # Create a workbook and activate the first sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "TitleBlockData"
    preferences.SetString("SheetName", "TitleBlockData")
    preferences.SetString("StartCell", "A1")

    # Get the startcell and the next cells
    StartCell = str("A1")
    TopRow = int(StartCell[1:])
    PropCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    IncreaseCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    MultiplierCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    RemarkCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        message = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(message)

    # Set the headers
    ws[StartCell].value = "Property Name"
    ws[PropCell].value = "Property Value"
    ws[IncreaseCell].value = "Increase value"
    ws[MultiplierCell].value = "Factor"
    ws[RemarkCell].value = "Remarks"

    # region Format the settings with the values as a Table
    #
    # Define the the last cell
    EndCell = "E10"

    # Define the table
    NumberOfSheets = str(len(wb.sheetnames))
    tab = Table(
        displayName="TitleBlockTable_" + NumberOfSheets,
        ref=f"{StartCell}:{EndCell}",
    )

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    # add the style to the table
    tab.tableStyleInfo = style

    # add the table to the worksheet
    ws.add_table(tab)
    # endregion

    # Make the columns to autofit the date
    for col in ws.columns:
        SetLen = 0
        column = col[0].column_letter  # Get the column name

        for cell in col:
            if len(str(cell.value)) > SetLen:
                SetLen = len(str(cell.value))

        set_col_width = SetLen + 5
        # Setting the column width
        ws.column_dimensions[column].width = set_col_width

    # Save the excel file in a folder of your choosing
    Filter = [
        ("Excel", "*.xlsx"),
    ]
    FileName = Standard_Functions.GetFileDialog(Filter)
    if FileName != "":
        # Save the workbook
        wb.save(str(FileName))
        # Close the workbook
        wb.close()
        # Update the preferences
        preferences.SetString("ExternalFile", rf"{FileName}")
    if FileName == "":
        return

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        SettingsTB.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {ws.title}",
    )
    Standard_Functions.Mbox(text=message, title="TitleBlock Workbench", style=0)

    return


def CreateFreeCAD():
    # Get the active document
    doc = App.ActiveDocument
    # Save the name of the active document to reactivate it at the end of this function.
    LastActiveDoc = doc.Name

    # Create a placeholder for the new document
    ff = ""
    # Save the FreeCAD file in a folder of your choosing
    Filter = [
        ("FreeCAD", "*.FCStd"),
    ]
    FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=True)
    if FileName != "":
        # Create a new FreeCAD file
        ff = App.newDocument()
        # Save the workbook
        ff.saveAs(FileName)
        # Close the document before reopening
        App.closeDocument(ff.Name)
    if FileName == "":
        return

    # Open the document hidden, recompute and save it
    ff = App.openDocument(FileName, True)
    ff.recompute(None, True, True)
    ff.save

    # If there already are objects, clean the document
    Objects = ff.Objects
    for i in range(len(Objects)):
        ff.removeObject(Objects[i].Name)

    # Create a spreadsheet for the titleblock data.
    TitleBlockData = ff.addObject("Spreadsheet::Sheet", "TitleBlockData")
    preferences.SetString("SheetName", "TitleBlockData")

    # Get the startcell and the next cells
    StartCell = "A1"
    TopRow = int(StartCell[1:])
    PropCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    IncreaseCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    MultiplierCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    RemarkCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(Text)

    # Set the headers
    TitleBlockData.set(StartCell, "Property Name")
    TitleBlockData.set(PropCell, "Property Value")
    TitleBlockData.set(IncreaseCell, "Increase value")
    TitleBlockData.set(MultiplierCell, "Factor")
    TitleBlockData.set(RemarkCell, "Remarks")

    # region Format the settings with the values as a Table
    #
    # Define the the last cell
    EndRow = 10

    FormatTable(sheet=TitleBlockData, Endrow=EndRow)
    # endregion

    # recompute the document
    ff.recompute(None, True, True)
    # Save the workbook
    ff.save()

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        SettingsTB.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {TitleBlockData}",
    )
    Standard_Functions.Mbox(text=message, title="TitleBlock Workbench", style=0)

    # Close the FreeCAD file
    App.closeDocument(ff.Name)
    # Activate the document which was active when this command started.
    try:
        App.setActiveDocument(LastActiveDoc)
    except Exception:
        pass

    return


# region - supporting functions


# Function the return the correct string to use in the FormatTable function
def FontStyle(Bold: bool, Italic: bool, UnderLine: bool) -> str:
    result = ""
    if Bold is True:
        if result == "":
            result = "bold"
        if result != "":
            result = result + "|bold"
    if Italic is True:
        if result == "":
            result = "italic"
        if result != "":
            result = result + "|italic"
    if UnderLine is True:
        if result == "":
            result = "underline"
        if result != "":
            result = result + "|underline"
    return result


# Format the table based on the preferences
def FormatTable(sheet, Endrow):
    # HeaderRange
    RangeAlign1 = "A1:A" + str(Endrow)
    RangeStyle1 = "A1:E1"

    # TableRange
    RangeAlign2 = "B1:E" + str(Endrow)
    RangeStyle2 = "B2:E" + str(Endrow)
    # First column
    RangeStyle3 = "A2:A" + str(Endrow)

    # Font style for the top row
    sheet.setStyle(
        RangeStyle1,
        FontStyle(
            SPREADSHEET_HEADERFONTSTYLE_BOLD,
            SPREADSHEET_HEADERFONTSTYLE_ITALIC,
            SPREADSHEET_HEADERFONTSTYLE_UNDERLINE,
        ),
    )

    # Font style for the first column
    sheet.setStyle(
        RangeStyle3,
        FontStyle(
            SPREADSHEET_COLUMNFONTSTYLE_BOLD,
            SPREADSHEET_COLUMNFONTSTYLE_ITALIC,
            SPREADSHEET_COLUMNFONTSTYLE_UNDERLINE,
        ),
    )

    # Font style for the rest of the table
    sheet.setStyle(
        RangeStyle2,
        FontStyle(
            SPREADSHEET_TABLEFONTSTYLE_BOLD,
            SPREADSHEET_TABLEFONTSTYLE_ITALIC,
            SPREADSHEET_TABLEFONTSTYLE_UNDERLINE,
        ),
    )

    sheet.setBackground(RangeStyle1, SPREADSHEET_HEADERBACKGROUND)
    sheet.setForeground(RangeStyle1, SPREADSHEET_HEADERFOREGROUND)

    # Style the rest of the table
    for i in range(2, int(Endrow) + 1, 2):
        RangeStyle3 = f"A{i}:E{i}"
        RangeStyle4 = f"A{i+1}:E{i+1}"
        sheet.setBackground(RangeStyle3, SPREADSHEET_TABLEBACKGROUND_1)
        sheet.setBackground(RangeStyle4, SPREADSHEET_TABLEBACKGROUND_2)
        sheet.setForeground(RangeStyle3, SPREADSHEET_TABLEFOREGROUND)
        sheet.setForeground(RangeStyle4, SPREADSHEET_TABLEFOREGROUND)

    # align the columns
    sheet.setAlignment(RangeAlign1, "left|vcenter")
    sheet.setAlignment(RangeAlign2, "center|vcenter")

    # Set the column width
    for i in range(1, Endrow):
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"A{i}", cellValue=sheet.getContents(f"A{i}"), factor=AUTOFIT_FACTOR)
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"B{i}", cellValue=sheet.getContents(f"B{i}"), factor=AUTOFIT_FACTOR)
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"C{i}", cellValue=sheet.getContents(f"C{i}"), factor=AUTOFIT_FACTOR)
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"D{i}", cellValue=sheet.getContents(f"D{i}"), factor=AUTOFIT_FACTOR)
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"E{i}", cellValue=sheet.getContents(f"E{i}"), factor=AUTOFIT_FACTOR)
    return

# endregion
