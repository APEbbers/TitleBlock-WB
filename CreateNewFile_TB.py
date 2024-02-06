# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/

import FreeCAD as App
import Standard_Functions_TB as Standard_Functions
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Alignment
import TableFormat_Functions_TB

# Get the settings
import Settings_TB
from Settings_TB import preferences
from Settings_TB import IMPORT_SETTINGS_XL
from Settings_TB import ENABLE_DEBUG
from Settings_TB import SPREADSHEET_COLUMNFONTSTYLE_UNDERLINE
from Settings_TB import SPREADSHEET_COLUMNFONTSTYLE_ITALIC
from Settings_TB import SPREADSHEET_COLUMNFONTSTYLE_BOLD
from Settings_TB import SPREADSHEET_TABLEFONTSTYLE_UNDERLINE
from Settings_TB import SPREADSHEET_TABLEFONTSTYLE_ITALIC
from Settings_TB import SPREADSHEET_TABLEFONTSTYLE_BOLD
from Settings_TB import SPREADSHEET_TABLEFOREGROUND
from Settings_TB import SPREADSHEET_TABLEBACKGROUND_2
from Settings_TB import SPREADSHEET_TABLEBACKGROUND_1
from Settings_TB import SPREADSHEET_HEADERFONTSTYLE_UNDERLINE
from Settings_TB import SPREADSHEET_HEADERFONTSTYLE_ITALIC
from Settings_TB import SPREADSHEET_HEADERFONTSTYLE_BOLD
from Settings_TB import SPREADSHEET_HEADERFOREGROUND
from Settings_TB import SPREADSHEET_HEADERBACKGROUND
from Settings_TB import AUTOFIT_FACTOR

# Define the translation
translate = App.Qt.translate


# region - TitleBlockData
def CreateTitleBlockData_Excel():
    # Create a workbook and activate the first sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "TitleBlockData"
    preferences.SetString("SheetName", "TitleBlockData")
    preferences.SetString("StartCell", "A1")

    # Get the startcell and the next cells
    StartCell = str("A1")
    TopRow = int(StartCell[1:])
    PropCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    IncreaseCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    MultiplierCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    RemarkCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        message = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(message)

    # Set the headers
    ws[StartCell].value = "Property Name"
    ws[PropCell].value = "Property Value"
    ws[IncreaseCell].value = "Increase value"
    ws[MultiplierCell].value = "Factor"
    ws[RemarkCell].value = "Remarks"

    # region Format the settings with the values as a Table
    #
    # Define the the last cell
    EndCell = "E10"

    # Define the table
    NumberOfSheets = str(len(wb.sheetnames))
    tab = Table(
        displayName="TitleBlockTable_" + NumberOfSheets,
        ref=f"{StartCell}:{EndCell}",
    )

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    # add the style to the table
    tab.tableStyleInfo = style

    # add the table to the worksheet
    ws.add_table(tab)
    # endregion

    # Make the columns to autofit the date
    for col in ws.columns:
        SetLen = 0
        column = col[0].column_letter  # Get the column name

        for cell in col:
            if len(str(cell.value)) > SetLen:
                SetLen = len(str(cell.value))

        set_col_width = SetLen + 5
        # Setting the column width
        ws.column_dimensions[column].width = set_col_width

    # Save the excel file in a folder of your choosing
    Filter = [
        ("Excel", "*.xlsx"),
    ]
    FileName = Standard_Functions.GetFileDialog(Filter)
    if FileName != "":
        # Save the workbook
        wb.save(str(FileName))
        # Close the workbook
        wb.close()
        # Update the preferences
        preferences.SetString("ExternalFile", rf"{FileName}")
    if FileName == "":
        return

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        Settings_TB.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {ws.title}",
    )
    Standard_Functions.Mbox(text=message, title="TitleBlock Workbench", style=0)

    return


def CreateTitleBlockData_FreeCAD():
    # Get the active document
    doc = App.ActiveDocument
    # Save the name of the active document to reactivate it at the end of this function.
    LastActiveDoc = doc.Name

    # Create a placeholder for the new document
    ff = ""
    # Save the FreeCAD file in a folder of your choosing
    Filter = [
        ("FreeCAD", "*.FCStd"),
    ]
    FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=True)
    if FileName != "":
        # Create a new FreeCAD file
        ff = App.newDocument()
        # Save the workbook
        ff.saveAs(FileName)
        # Close the document before reopening
        App.closeDocument(ff.Name)
    if FileName == "":
        return

    # Open the document hidden, recompute and save it
    ff = App.openDocument(FileName, True)
    ff.recompute(None, True, True)
    ff.save

    # If there already are objects, clean the document
    Objects = ff.Objects
    for i in range(len(Objects)):
        ff.removeObject(Objects[i].Name)

    # Create a spreadsheet for the titleblock data.
    TitleBlockData = ff.addObject("Spreadsheet::Sheet", "TitleBlockData")
    preferences.SetString("SheetName", "TitleBlockData")

    # Get the startcell and the next cells
    StartCell = "A1"
    TopRow = int(StartCell[1:])
    PropCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    IncreaseCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    MultiplierCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    RemarkCell = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(Text)

    # Set the headers
    TitleBlockData.set(StartCell, "Property Name")
    TitleBlockData.set(PropCell, "Property Value")
    TitleBlockData.set(IncreaseCell, "Increase value")
    TitleBlockData.set(MultiplierCell, "Factor")
    TitleBlockData.set(RemarkCell, "Remarks")

    # Add data, otherwise formatting doesn't work.
    for i in range(9):
        TitleBlockData.set(f"A{i+2}", "-")
        TitleBlockData.set(f"B{i+2}", "-")
        TitleBlockData.set(f"C{i+2}", "-")
        TitleBlockData.set(f"D{i+2}", "-")
        TitleBlockData.set(f"E{i+2}", "-")

    # region Format the settings with the values as a Table
    #
    TableFormat_Functions_TB.FormatTable(
        sheet=TitleBlockData,
        HeaderRange="A1:E1",
        TableRange="A2:E10",
        FirstColumnRange="A2:A10",
    )
    # endregion

    # recompute the document
    ff.recompute(None, True, True)
    # Save the workbook
    ff.save()

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        Settings_TB.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {TitleBlockData}",
    )
    Standard_Functions.Mbox(text=message, title="TitleBlock Workbench", style=0)

    # Close the FreeCAD file
    App.closeDocument(ff.Name)
    # Activate the document which was active when this command started.
    try:
        App.setActiveDocument(LastActiveDoc)
    except Exception:
        pass

    return


# endregion


# region - SimpleDrawingList
def CreateSimpleDrawingList_Excel():
    # Create a workbook and activate the first sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "DrawingList"
    preferences.SetString("SheetName_SimpleList", "DrawingList")
    preferences.SetString("StartCell_SimpleList", "A1")

    # Get the startcell and the next cells
    StartCell = str("A1")
    TopRow = int(StartCell[1:])
    EditableText_1 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    EditableText_2 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    EditableText_3 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    EditableText_4 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        message = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(message)

    # Set the headers
    ws[StartCell].value = "Property Value"
    ws[EditableText_1].value = "<Editable text - Name>(1)"
    ws[EditableText_2].value = "<Editable text - Name>(2)"
    ws[EditableText_3].value = "<Editable text - Name>(3)"
    ws[EditableText_4].value = "<Editable text - Name>(4)"

    # region Format the settings with the values as a Table
    #
    # Define the the last cell
    EndCell = "E10"

    # Define the table
    NumberOfSheets = str(len(wb.sheetnames))
    tab = Table(
        displayName="DrawingList_" + NumberOfSheets,
        ref=f"{StartCell}:{EndCell}",
    )

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    # add the style to the table
    tab.tableStyleInfo = style

    # add the table to the worksheet
    ws.add_table(tab)
    # endregion

    # Make the columns to autofit the date
    for col in ws.columns:
        SetLen = 0
        column = col[0].column_letter  # Get the column name

        for cell in col:
            if len(str(cell.value)) > SetLen:
                SetLen = len(str(cell.value))

        set_col_width = SetLen + 5
        # Setting the column width
        ws.column_dimensions[column].width = set_col_width

    # Save the excel file in a folder of your choosing
    Filter = [
        ("Excel", "*.xlsx"),
    ]
    FileName = Standard_Functions.GetFileDialog(Filter)
    if FileName != "":
        # Save the workbook
        wb.save(str(FileName))
        # Close the workbook
        wb.close()
        # Update the preferences
        preferences.SetString("ExternalFile_SimpleList", rf"{FileName}")
    if FileName == "":
        return

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        Settings_TB.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {ws.title}",
    )
    Standard_Functions.Mbox(text=message, title="TitleBlock Workbench", style=0)

    return


def CreateSimpleDrawingList_FreeCAD():
    # Get the active document
    doc = App.ActiveDocument
    # Save the name of the active document to reactivate it at the end of this function.
    LastActiveDoc = doc.Name

    # Create a placeholder for the new document
    ff = ""
    # Save the FreeCAD file in a folder of your choosing
    Filter = [
        ("FreeCAD", "*.FCStd"),
    ]
    FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=True)
    if FileName != "":
        # Create a new FreeCAD file
        ff = App.newDocument()
        # Save the workbook
        ff.saveAs(FileName)
        # Close the document before reopening
        App.closeDocument(ff.Name)
    if FileName == "":
        return

    # Open the document hidden, recompute and save it
    ff = App.openDocument(FileName, True)
    ff.recompute(None, True, True)
    ff.save

    # If there already are objects, clean the document
    Objects = ff.Objects
    for i in range(len(Objects)):
        ff.removeObject(Objects[i].Name)

    # Create a spreadsheet for the titleblock data.
    DrawingList = ff.addObject("Spreadsheet::Sheet", "DrawingList")
    preferences.SetString("SheetName_SimpleList", "DrawingList")

    # Get the startcell and the next cells
    StartCell = "A1"
    preferences.SetString("StartCell_SimpleList", StartCell)
    TopRow = int(StartCell[1:])
    EditableText_1 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    EditableText_2 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    EditableText_3 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    EditableText_4 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        Standard_Functions.Print(Text, "Log")

    # Set the headers
    DrawingList.set(StartCell, "Property Value")
    DrawingList.set(EditableText_1, "<Editable text - Name>(1)")
    DrawingList.set(EditableText_2, "<Editable text - Name>(2)")
    DrawingList.set(EditableText_3, "<Editable text - Name>(3)")
    DrawingList.set(EditableText_4, "<Editable text - Name>(4)")

    # Add data, otherwise formatting doesn't work.
    i = 0
    for i in range(9):
        DrawingList.set(f"A{i+2}", "Value")
        DrawingList.set(f"B{i+2}", "<Editable text - Value>(1)")
        DrawingList.set(f"C{i+2}", "<Editable text - Value>(2)")
        DrawingList.set(f"D{i+2}", "<Editable text - Value>(3)")
        DrawingList.set(f"E{i+2}", "<Editable text - Value>(4)")
    # Set instruction text
    DrawingList.set(f"A{i+3}", "Value to look up")
    DrawingList.mergeCells(f"B{i+3}:E{i+3}")
    DrawingList.set(f"B{i+3}", "Values to fill in titleblock")

    # region Format the settings with the values as a Table
    #
    DrawingList = TableFormat_Functions_TB.FormatTable(
        sheet=DrawingList,
        HeaderRange="A1:E1",
        TableRange="A2:E10",
        FirstColumnRange="A2:A10",
    )

    DrawingList.setBackground("A11", Standard_Functions.ColorConvertor([255, 230, 153]))
    DrawingList.setAlignment("A11", "left|vcenter")
    DrawingList.mergeCells("B11:E11")
    DrawingList.setBackground(
        "B11:E11", Standard_Functions.ColorConvertor([198, 224, 180])
    )
    DrawingList.setAlignment("B11:E11", "center|vcenter")
    # endregion

    # recompute the document
    ff.recompute(None, True, True)
    # Save the workbook
    ff.save()

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        Settings_TB.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {DrawingList}",
    )
    DrawingList = Standard_Functions.Mbox(
        text=message, title="TitleBlock Workbench", style=0
    )

    # Close the FreeCAD file
    App.closeDocument(ff.Name)
    # Activate the document which was active when this command started.
    try:
        App.setActiveDocument(LastActiveDoc)
    except Exception:
        pass

    return


def CreateSimpleDrawingList_Internal():
    from Settings_TB import SHEETNAME_SIMPLE_LIST
    import Standard_Functions_TB

    # Get the active document
    doc = App.ActiveDocument

    DrawingList = doc.getObject(SHEETNAME_SIMPLE_LIST)
    if DrawingList is not None:
        Text = translate(
            "TitleBlock Workbench",
            "Do you want to replace the drawing list with a new one?",
        )
        reply = Standard_Functions_TB.Mbox(
            text=Text, title="TitleBlock Workbench", style=1, IconType="Question"
        )

        if reply == "yes":
            doc.removeObject(SHEETNAME_SIMPLE_LIST)

            Text = translate(
                "TitleBlock Workbench",
                "Please enter the name of the drawing list.\n",
            )
            DrawingListName = Standard_Functions.Mbox(
                text=Text, title="TitleBlock Workbench", style=20, default="DrawingList"
            )

            if DrawingListName != "":
                DrawingList = doc.addObject("Spreadsheet::Sheet", DrawingListName)

                # Set SHEETNAME_STARTCELL to the chosen sheetname
                preferences.SetString("SheetName_SimpleList", DrawingListName)
                SHEETNAME_SIMPLE_LIST = DrawingList

        if reply == "no":
            return

    if DrawingList is None:
        DrawingList = doc.addObject("Spreadsheet::Sheet", "DrawingList")
        preferences.SetString("SheetName_SimpleList", "DrawingList")

    # Get the startcell and the next cells
    StartCell = "A1"
    TopRow = int(StartCell[1:])
    EditableText_1 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    EditableText_2 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    EditableText_3 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    EditableText_4 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(Text)

    # Set the headers
    DrawingList.set(StartCell, "Property")
    DrawingList.set(EditableText_1, "<Editable text - Name>(1)")
    DrawingList.set(EditableText_2, "<Editable text - Name>(2)")
    DrawingList.set(EditableText_3, "<Editable text - Name>(3)")
    DrawingList.set(EditableText_4, "<Editable text - Name>(4)")

    # Add data, otherwise formatting doesn't work.
    i = 0
    for i in range(9):
        DrawingList.set(f"A{i+2}", "Value")
        DrawingList.set(f"B{i+2}", "<Editable text - Value>(1)")
        DrawingList.set(f"C{i+2}", "<Editable text - Value>(2)")
        DrawingList.set(f"D{i+2}", "<Editable text - Value>(3)")
        DrawingList.set(f"E{i+2}", "<Editable text - Value>(4)")
    # Set instruction text
    DrawingList.set(f"A{i+3}", "Value to look up")
    DrawingList.mergeCells(f"B{i+3}:E{i+3}")
    DrawingList.set(f"B{i+3}", "Values to fill in titleblock")

    # region Format the settings with the values as a Table
    #
    DrawingList = TableFormat_Functions_TB.FormatTable(
        sheet=DrawingList,
        HeaderRange="A1:E1",
        TableRange="A2:E10",
        FirstColumnRange="A2:A10",
    )

    DrawingList.setBackground("A11", Standard_Functions.ColorConvertor([255, 230, 153]))
    DrawingList.setAlignment("A11", "left|vcenter")
    DrawingList.mergeCells("B11:E11")
    DrawingList.setBackground(
        "B11:E11", Standard_Functions.ColorConvertor([198, 224, 180])
    )
    DrawingList.setAlignment("B11:E11", "center|vcenter")
    # endregion

    # recompute the document
    doc.recompute(None, True, True)
    # Save the workbook
    doc.save()

    return


# endregion


# region - AdvancedDrawingList
def CreateAdvancedDrawingList_Excel():
    # Create a workbook and activate the first sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "DrawingList"
    preferences.SetString("SheetName_AdvancedList", "DrawingList")
    preferences.SetString("StartCell_AdvancedList", "A1")

    # Get the startcell and the next cells
    StartCell = str("A1")
    TopRow = int(StartCell[1:])
    EditableText_1 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    EditableText_2 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    EditableText_3 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    EditableText_4 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        message = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(message)

    # Set the headers
    ws[StartCell].value = "Property Value"
    ws[EditableText_1].value = "<Editable text - Name>(1)"
    ws[EditableText_2].value = "<Editable text - Name>(2)"
    ws[EditableText_3].value = "<Editable text - Name>(3)"
    ws[EditableText_4].value = "<Editable text - Name>(4)"

    # region Format the settings with the values as a Table
    #
    # Define the the last cell
    EndCell_G1 = "E9"

    # Define the table
    NumberOfSheets = str(len(wb.sheetnames))
    tab = Table(
        displayName="DrawingList_" + NumberOfSheets,
        ref=f"{StartCell}:{EndCell_G1}",
    )

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    # add the style to the table
    tab.tableStyleInfo = style

    # add the table to the worksheet
    ws.add_table(tab)

    # Create group 1 and 2. Aplly a fill to Group 1 and 2.
    for i in range(1, 6):
        Cell1 = ws[f"{Standard_Functions.GetLetterFromNumber(i)}2"]
        Cell1.fill = PatternFill("solid", fgColor="CC99FF")

        Cell2 = ws[f"{Standard_Functions.GetLetterFromNumber(i)}6"]
        Cell2.fill = PatternFill("solid", fgColor="CC99FF")
    # endregion

    # Fill the values
    # Groups
    ws["A2"].value = "Group 1"
    ws["A6"].value = "Group 2"
    # Content
    for i in range(3, 6):
        ws[f"A{i}"].value = "Value"
        ws[f"B{i}"].value = "<Editable text - Value>(1)"
        ws[f"C{i}"].value = "<Editable text - Value>(2)"
        ws[f"D{i}"].value = "<Editable text - Value>(3)"
        ws[f"E{i}"].value = "<Editable text - Value>(4)"
    for i in range(7, 10):
        ws[f"A{i}"].value = "Value"
        ws[f"B{i}"].value = "<Editable text - Value>(1)"
        ws[f"C{i}"].value = "<Editable text - Value>(2)"
        ws[f"D{i}"].value = "<Editable text - Value>(3)"
        ws[f"E{i}"].value = "<Editable text - Value>(4)"
    ws[f"A{i+1}"].value = "Value to look up"
    ws[f"A{i+1}"].fill = PatternFill("solid", fgColor="FFE699")
    ws.merge_cells(f"B{i+1}:E{i+1}")
    ws[f"B{i+1}"].fill = PatternFill("solid", fgColor="C6E0B4")
    ws[f"B{i+1}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"B{i+1}"].value = "Values to fill in titleblock"

    # Make the columns to autofit the date
    for col in ws.columns:
        SetLen = 0
        column = col[0].column_letter  # Get the column name

        for cell in col:
            if len(str(cell.value)) > SetLen:
                SetLen = len(str(cell.value))

        set_col_width = SetLen + 5
        # Setting the column width
        ws.column_dimensions[column].width = set_col_width

    # Save the excel file in a folder of your choosing
    Filter = [
        ("Excel", "*.xlsx"),
    ]
    FileName = Standard_Functions.GetFileDialog(Filter)
    if FileName != "":
        # Save the workbook
        wb.save(str(FileName))
        # Close the workbook
        wb.close()
        # Update the preferences
        preferences.SetString("ExternalFile_AdvancedList", rf"{FileName}")
    if FileName == "":
        return

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        Settings_TB.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {ws.title}",
    )
    Standard_Functions.Mbox(text=message, title="TitleBlock Workbench", style=0)

    return


def CreateAdvancedDrawingList_FreeCAD():
    # Get the active document
    doc = App.ActiveDocument
    # Save the name of the active document to reactivate it at the end of this function.
    LastActiveDoc = doc.Name

    # Create a placeholder for the new document
    ff = ""
    # Save the FreeCAD file in a folder of your choosing
    Filter = [
        ("FreeCAD", "*.FCStd"),
    ]
    FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=True)
    if FileName != "":
        # Create a new FreeCAD file
        ff = App.newDocument()
        # Save the workbook
        ff.saveAs(FileName)
        # Close the document before reopening
        App.closeDocument(ff.Name)
        # Update the settings
        preferences.SetString("ExternalFile_AdvancedList", rf"{FileName}")
    if FileName == "":
        return

    # Open the document hidden, recompute and save it
    ff = App.openDocument(FileName, True)
    ff.recompute(None, True, True)
    ff.save

    # If there already are objects, clean the document
    Objects = ff.Objects
    for i in range(len(Objects)):
        ff.removeObject(Objects[i].Name)

    # Create a spreadsheet for the titleblock data.
    DrawingList = ff.addObject("Spreadsheet::Sheet", "DrawingList")
    preferences.SetString("SheetName_AdvancedList", "DrawingList")

    # Get the startcell and the next cells
    StartCell = "A1"
    preferences.SetString("StartCell_AdvancedList", StartCell)
    TopRow = int(StartCell[1:])
    EditableText_1 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    EditableText_2 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    EditableText_3 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    EditableText_4 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        Standard_Functions.Print(Text, "Log")

    # Set the headers
    DrawingList.set(StartCell, "Property Value")
    DrawingList.set(EditableText_1, "<Editable text - Name>(1)")
    DrawingList.set(EditableText_2, "<Editable text - Name>(2)")
    DrawingList.set(EditableText_3, "<Editable text - Name>(3)")
    DrawingList.set(EditableText_4, "<Editable text - Name>(4)")

    # Add data, otherwise formatting doesn't work.
    i = 0
    for i in range(8):
        DrawingList.set(f"A{i+2}", "Value")
        DrawingList.set(f"B{i+2}", "<Editable text - Value>(1)")
        DrawingList.set(f"C{i+2}", "<Editable text - Value>(2)")
        DrawingList.set(f"D{i+2}", "<Editable text - Value>(3)")
        DrawingList.set(f"E{i+2}", "<Editable text - Value>(4)")
    # Set instruction text
    DrawingList.set(f"A{i+3}", "Value to look up")
    DrawingList.mergeCells(f"B{i+3}:E{i+3}")
    DrawingList.set(f"B{i+3}", "Values to fill in titleblock")

    # region Format the settings with the values as a Table
    #
    DrawingList = TableFormat_Functions_TB.FormatTable(
        sheet=DrawingList,
        HeaderRange="A1:E1",
        TableRange="A2:E10",
        FirstColumnRange="A2:A10",
    )

    DrawingList.setBackground("A10", Standard_Functions.ColorConvertor([255, 230, 153]))
    DrawingList.setAlignment("A10", "left|vcenter")
    DrawingList.mergeCells("B10:E10")
    DrawingList.setBackground(
        "B10:E10", Standard_Functions.ColorConvertor([198, 224, 180])
    )
    DrawingList.setAlignment("B10:E10", "center|vcenter")
    # endregion

    # Set the groups
    DrawingList.mergeCells("A2:E2")
    DrawingList.set("A2", "Group 1")
    DrawingList.setBackground("A2", Standard_Functions.ColorConvertor([204, 153, 255]))
    DrawingList.mergeCells("A6:E6")
    DrawingList.set("A6", "Group 2")
    DrawingList.setBackground("A6", Standard_Functions.ColorConvertor([204, 153, 255]))

    # recompute the document
    ff.recompute(None, True, True)
    # Save the workbook
    ff.save()

    # If import settings from excel is enabled, export settings to the new excel file.
    if IMPORT_SETTINGS_XL is True:
        Settings_TB.ExportSettings_XL(Silent=True)

    # print a message if you succeded.
    message = translate(
        "TitleBlock Workbench",
        f"The titleblock data is exported to the workbook {FileName} in the worksheet {DrawingList}",
    )
    DrawingList = Standard_Functions.Mbox(
        text=message, title="TitleBlock Workbench", style=0
    )

    # Close the FreeCAD file
    ff.recompute(None, True, True)
    ff.save
    App.closeDocument(ff.Name)

    # Activate the document which was active when this command started.
    try:
        App.setActiveDocument(LastActiveDoc)
    except Exception:
        pass

    return


def CreateAdvancedDrawingList_Internal():
    from Settings_TB import SHEETNAME_ADVANCED_LIST
    import Standard_Functions_TB

    # Get the active document
    doc = App.ActiveDocument

    DrawingList = doc.getObject(SHEETNAME_ADVANCED_LIST)
    if DrawingList is not None:
        Text = translate(
            "TitleBlock Workbench",
            "Do you want to replace the drawing list with a new one?",
        )
        reply = Standard_Functions_TB.Mbox(
            text=Text, title="TitleBlock Workbench", style=1, IconType="Question"
        )

        if reply == "yes":
            doc.removeObject(SHEETNAME_ADVANCED_LIST)

            Text = translate(
                "TitleBlock Workbench",
                "Please enter the name of the drawing list.\n",
            )
            DrawingListName = Standard_Functions.Mbox(
                text=Text, title="TitleBlock Workbench", style=20, default="DrawingList"
            )

            if DrawingListName != "":
                DrawingList = doc.addObject("Spreadsheet::Sheet", DrawingListName)

                # Set SHEETNAME_STARTCELL to the chosen sheetname
                preferences.SetString("SheetName_AdvancedList", DrawingListName)
        if reply == "no":
            return

    if DrawingList is None:
        DrawingList = doc.addObject("Spreadsheet::Sheet", "DrawingList")
        preferences.SetString("SheetName_AdvancedList", "DrawingList")

    # Get the startcell and the next cells
    StartCell = "A1"
    TopRow = int(StartCell[1:])
    EditableText_1 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
        )
    ) + str(TopRow)
    EditableText_2 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
        )
    ) + str(TopRow)
    EditableText_3 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
        )
    ) + str(TopRow)
    EditableText_4 = str(
        Standard_Functions.GetLetterFromNumber(
            Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
        )
    ) + str(TopRow)

    if ENABLE_DEBUG is True:
        Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
        print(Text)

    # Set the headers
    DrawingList.set(StartCell, "Property")
    DrawingList.set(EditableText_1, "<Editable text - Name>(1)")
    DrawingList.set(EditableText_2, "<Editable text - Name>(2)")
    DrawingList.set(EditableText_3, "<Editable text - Name>(3)")
    DrawingList.set(EditableText_4, "<Editable text - Name>(4)")

    # Add data, otherwise formatting doesn't work.
    i = 0
    for i in range(8):
        DrawingList.set(f"A{i+2}", "Value")
        DrawingList.set(f"B{i+2}", "<Editable text - Value>(1)")
        DrawingList.set(f"C{i+2}", "<Editable text - Value>(2)")
        DrawingList.set(f"D{i+2}", "<Editable text - Value>(3)")
        DrawingList.set(f"E{i+2}", "<Editable text - Value>(4)")
    # Set instruction text
    DrawingList.set(f"A{i+3}", "Value to look up")
    DrawingList.mergeCells(f"B{i+3}:E{i+3}")
    DrawingList.set(f"B{i+3}", "Values to fill in titleblock")

    # region Format the settings with the values as a Table
    #
    DrawingList = TableFormat_Functions_TB.FormatTable(
        sheet=DrawingList,
        HeaderRange="A1:E1",
        TableRange="A2:E10",
        FirstColumnRange="A2:A10",
    )

    DrawingList.setBackground("A10", Standard_Functions.ColorConvertor([255, 230, 153]))
    DrawingList.setAlignment("A10", "left|vcenter")
    DrawingList.mergeCells("B10:E10")
    DrawingList.setBackground(
        "B10:E10", Standard_Functions.ColorConvertor([198, 224, 180])
    )
    DrawingList.setAlignment("B10:E10", "center|vcenter")
    # endregion

    # Set the groups
    DrawingList.mergeCells("A2:E2")
    DrawingList.set("A2", "Group 1")
    DrawingList.setBackground("A2", Standard_Functions.ColorConvertor([204, 153, 255]))
    DrawingList.mergeCells("A6:E6")
    DrawingList.set("A6", "Group 2")
    DrawingList.setBackground("A6", Standard_Functions.ColorConvertor([204, 153, 255]))

    # recompute the document
    doc.recompute(None, True, True)
    # Save the workbook
    doc.save()

    return


# endregion
