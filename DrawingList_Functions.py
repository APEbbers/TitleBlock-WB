# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/

# Additional functions to map properties based on values in a drawing list

# region imports
import FreeCAD as App
import Standard_Functions_TitleBlock as Standard_Functions

# Get the settings
from SettingsTB import ENABLE_DEBUG
from SettingsTB import USE_SIMPLE_LIST
from SettingsTB import USE_EXTERNAL_SOURCE_SIMPLE_LIST
from SettingsTB import EXTERNAL_FILE_SIMPLE_LIST
from SettingsTB import SHEETNAME_SIMPLE_LIST
from SettingsTB import STARTCELL_SIMPLE_LIST
from SettingsTB import PROPERTY_NAME_SIMPLE_LIST
from SettingsTB import USE_PAGE_NAMES_SIMPLE_LIST
from SettingsTB import USE_ADVANCED_LIST
from SettingsTB import EXTERNAL_FILE_ADVANCED_LIST
from SettingsTB import SHEETNAME_ADVANCED_LIST
from SettingsTB import STARTCELL_ADVANCED_LIST
from SettingsTB import PROPERTY_NAME_ADVANCED_LIST
from SettingsTB import SORTING_PREFIX_ADVANCED_LIST
from SettingsTB import USE_EXTERNAL_SOURCE_ADVANCED_LIST
from SettingsTB import USE_PAGE_NAMES_ADVANCED_LIST
from SettingsTB import preferences

# Define the translation
translate = App.Qt.translate

# endregion


def MapSimpleDrawingList(sheet):
    # Check if it is allowed to use an external source and if so, continue
    if USE_SIMPLE_LIST is True and USE_EXTERNAL_SOURCE_SIMPLE_LIST is False:
        try:
            # Get the active document
            doc = App.ActiveDocument
            # Save the name of the active document to reactivate it at the end of this function.
            LastActiveDoc = doc.Name
            # Get the Drawing list.
            DrawingList = doc.getObject(SHEETNAME_SIMPLE_LIST)
            if DrawingList is None:
                SheetName = Standard_Functions.Mbox(text="DrawingList", title="Title block workbench", style=20)
                if SheetName != "":
                    doc.addObject("Spreadsheet::Sheet", SheetName)
                    DrawingList = doc.getObject(SheetName)
                if SheetName == "":
                    return

            # Get the startcolumn and the other three columns from there
            StartCell = STARTCELL_SIMPLE_LIST

            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)
            StartColumn = Standard_Functions.RemoveNumbersFromString(StartCell)
            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column for the drawing list is: " + str(StartColumn),
                    ),
                    "Log",
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number for the drawing list is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumn)),
                    ),
                    "Log",
                )

            # Get the start row
            StartRow = Standard_Functions.RemoveLettersFromString(
                StartCell
            )
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row for the drawing list is: " + str(StartRow)
                )
                Standard_Functions.Print(Text, "Log")

            # Define placeholder for the property value in the drawing list.
            PropertyValue_DrawingList = ""
            # Create a list with return values
            ReturnNames_DrawingList = []
            for i in range(1000):
                # Get the colum letter. For example: i = 0, StartColumn = A->1 + 1 results in column B
                Column = Standard_Functions.GetLetterFromNumber(
                    i + Standard_Functions.GetNumberFromLetter(StartColumn) + 1)

                # If the cell is not empty, add the contents to the list
                if str(DrawingList.getContents(f"{Column}{StartRow}")) != "":
                    ReturnName = str(DrawingList.getContents(f"{Column}{StartRow}"))
                    if ReturnName[:1] == "'":
                        ReturnName = ReturnName[1:]
                    ReturnNames_DrawingList.append(ReturnName)

                    for j in range(1, 1000):
                        # check if you reached the end of the data.
                        TitleBlockPropName = sheet.getContents(f"A{str(j)}")
                        if TitleBlockPropName[:1] == "'":
                            TitleBlockPropName = TitleBlockPropName[1:]
                        if TitleBlockPropName == "" or TitleBlockPropName is None:
                            break

                        if TitleBlockPropName == ReturnName:
                            sheet.set(f"B{j}", "-")
                            sheet.set(f"E{j}", "Property value retrieved from drawing list")

                # If the cell is empty, you are on the end. Break the loop.
                if str(DrawingList.getContents(f"{Column}{StartRow}")) == "":
                    break

            # Get the pages in the document
            pages = doc.findObjects("TechDraw::DrawPage")

            # Go through the drawing list and collect the property value based on the property name searched for.
            for i in range(1000):
                # Define the start row. This is the Header row + 1 + i as counter
                RowNumber = int(StartRow) + i + 1

                # If the property name in the drawing list is empty, it is the end of the list.
                # Exit the function
                if str(DrawingList.getContents(f"{StartColumn}{RowNumber}")) == "":
                    break

                # Get the property name in the drawing list. If it starts with "'", remove it
                PropertyValue_DrawingList = str(DrawingList.getContents(f"{StartColumn}{RowNumber}"))
                if PropertyValue_DrawingList[:1] == "'":
                    PropertyValue_DrawingList = PropertyValue_DrawingList[1:]

                # Go through the columns starting from the column right from the column with the property value
                for j in range(len(ReturnNames_DrawingList)):
                    # Get the column letter
                    Column = Standard_Functions.GetLetterFromNumber(
                        j + Standard_Functions.GetNumberFromLetter(StartColumn) + 1)

                    # If the cell is not empty and j is lower then NoColumns, continue.
                    if str(DrawingList.getContents(f"{Column}{RowNumber}")) != "":
                        # Get the property value in the excel list. If it starts with "'", remove it
                        ReturnValue_DrawingList = str(DrawingList.getContents(f"{Column}{RowNumber}"))
                        if ReturnValue_DrawingList[:1] == "'":
                            ReturnValue_DrawingList = ReturnValue_DrawingList[1:]

                        # If page names are not to be mapped, go here
                        if USE_PAGE_NAMES_SIMPLE_LIST is False:
                            # Go through all the pages
                            for page in pages:
                                # Get the editable texts
                                texts = page.Template.EditableTexts
                                # If the value editable text in the titleblock to search for
                                # matches the property value in the drawing list,
                                # fill in the editable text that needs to be updated.
                                if texts[PROPERTY_NAME_SIMPLE_LIST] == PropertyValue_DrawingList:
                                    # Get the property name in the titleblock spreadsheet and fill it with the property value from the drawing list
                                    texts[ReturnNames_DrawingList[j]] = ReturnValue_DrawingList

                                    # Write all the updated text to the page.
                                    page.Template.EditableTexts = texts
                                    page.recompute()

                        # If page names are to be mapped, go here
                        if USE_PAGE_NAMES_SIMPLE_LIST is True:
                            # Go through the pages
                            for page in pages:
                                # If the Property name in the drawing list matches the page label:
                                # Fill in the desired editable text with the property value from the drawing list
                                if PropertyValue_DrawingList == page.Label:
                                    # Get the editable texts
                                    texts = page.Template.EditableTexts
                                    texts[ReturnNames_DrawingList[j]] = ReturnValue_DrawingList

                                    # Write all the updated text to the page.
                                    page.Template.EditableTexts = texts
                                    page.recompute()

            # recompute the document
            doc.recompute(None, True, True)
            # Save the workbook
            doc.save()
            # Activate the document which was active when this command started.
            try:
                App.setActiveDocument(LastActiveDoc)
            except Exception:
                pass

            # Recomute the document
            App.ActiveDocument.recompute(None, True, True)

        except Exception as e:
            Text = translate(
                "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!\n"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: an error occurred!!\n"
                    + "See the report view for details",
                )
                raise e
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        return


def MapSimpleDrawingList_Excel(sheet):
    from openpyxl import load_workbook

    # Check if it is allowed to use an external source and if so, continue
    if USE_SIMPLE_LIST is True and USE_EXTERNAL_SOURCE_SIMPLE_LIST is True:
        # if debug mode is enabled, show the external file including path.
        if ENABLE_DEBUG is True:
            Text = translate("TitleBlock Workbench", f"The drawing list is: {EXTERNAL_FILE_SIMPLE_LIST}")
            Standard_Functions.Print(Text, "Log")

        # try to open the source. if not show an messagebox and if debug mode is enabled, show the exeption as well
        try:
            # Define a placeholder for the workbook
            wb = ""
            # If the drawinglist is an FreeCAD document instead of an Excel workbook,
            # Let the user select the correct workbook. Otherwise just load the workbook.
            if EXTERNAL_FILE_SIMPLE_LIST.lower().endswith(".fcstd"):
                Filter = [
                    ("Excel", "*.xlsx"),
                ]
                FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=False)
                if FileName != "":
                    wb = load_workbook(FileName, read_only=True, data_only=True)
                if FileName == "":
                    return
            else:
                wb = load_workbook(
                    str(EXTERNAL_FILE_SIMPLE_LIST), read_only=True, data_only=True
                )
            # If the sheetname not set, let the user set the correct name.
            if SHEETNAME_SIMPLE_LIST == "":
                # Set the sheetname with a inputbox
                Worksheets_List = [i for i in wb.sheetnames if i != "Settings"]
                Text = translate(
                    "TitleBlock Workbench", "Please enter the name of the worksheet"
                )
                Input_SheetName = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=3,
                        default="TitleBlockData",
                        stringList=Worksheets_List,
                    )
                )
                # if the user canceled, exit this function.
                if not Input_SheetName.strip():
                    return
                ws = wb[str(Input_SheetName)]
            if SHEETNAME_SIMPLE_LIST != "":
                # Get the list of worksheets
                Worksheets_List = [i for i in wb.sheetnames if i != "Settings"]
                # Assume that the worksheet doesn't exits
                WorksheetExits = False
                # Go through the names of the worksheets
                for WorksheetName in Worksheets_List:
                    # If a name matches the sheetname in preferences, set WorksheetsExits to True
                    if WorksheetName == SHEETNAME_SIMPLE_LIST:
                        WorksheetExits is True
                # If WorksheetsExits is true, define the worksheet
                if WorksheetExits is True:
                    ws = wb[str(SHEETNAME_SIMPLE_LIST)]
                # If WorksheetsExits is false, ask the user to enter the correct name.
                if WorksheetExits is False:
                    # Set the sheetname with a inputbox
                    Text = translate(
                        "TitleBlock Workbench", "The worksheet doesn't exits!\nPlease enter the name of the worksheet"
                    )
                    Input_SheetName = str(
                        Standard_Functions.Mbox(
                            text=Text,
                            title="TitleBlock Workbench",
                            style=21,
                            default="TitleBlockData",
                            stringList=Worksheets_List,
                        )
                    )
                    # if the user canceled, exit this function.
                    if not Input_SheetName.strip():
                        return
                    # Define the worksheets
                    ws = wb[str(Input_SheetName)]
                    # Save the sheetname to the preferences
                    preferences.SetString("SheetName_SimpleList", Input_SheetName)
        except IOError:
            Standard_Functions.Mbox(
                "Permission error!!\nDo you have the file open?",
                "Titleblock Workbench",
                0,
                IconType="Critical",
            )
            return
        except Exception as e:
            if ENABLE_DEBUG is True:
                raise (e)
            Text = translate(
                "TitleBlock Workbench",
                "an problem occured while openening the excel file!\nDo you have it open in an another application",
            )
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
            return

        try:
            # Get the spreadsheet.
            if sheet is None:
                sheet = App.ActiveDocument.getObject("TitleBlock")

            # Get the startcolumn and the other three columns from there
            StartCell = STARTCELL_SIMPLE_LIST
            if SHEETNAME_SIMPLE_LIST == "":
                # Save drawinglist and sheetname to the preferences
                preferences.SetString("ExternalFile", FileName)
                preferences.SetString("SheetName_SimpleList", Input_SheetName)
                ws = wb[Input_SheetName]

                # Set the startcell with an inputbox
                Text = translate(
                    "TitleBlock Workbench",
                    "Please enter the name of the cell.\n"
                    + "Enter a single cell like 'A1', 'B2', etc. Other notations will be ignored!",
                )
                StartCell = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=20,
                        default="A1",
                    )
                )
                if not StartCell.strip():
                    StartCell = "A1"

                # Set SHEETNAME_STARTCELL to the chosen sheetname
                preferences.SetString("StartCell_SimpleList", StartCell)

            # If the StartCell is R1C1 style, change it
            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)

            # Get the start column
            StartColumn = Standard_Functions.RemoveNumbersFromString(StartCell)
            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column for the drawing list is: " + str(StartColumn),
                        "Log",
                    )
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number for the drawing list is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumn)),
                    ),
                    "Log",
                )

            # Get the start row
            StartRow = Standard_Functions.RemoveLettersFromString(StartCell)
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row for the drawing list is: " + str(StartRow)
                )
                Standard_Functions.Print(Text, "Log")

            # Define placeholder for the property value in the excel list.
            PropertyValueExcel = ""
            # Define a placeholder for the qty of columns with return values
            NoColumns = 0
            # Create a list with return values
            ReturnNamesExcel = []
            for i in range(1000):
                # Get the colum letter. For example: i = 0, StartColumn = A->1 + 1 results in column B
                Column = Standard_Functions.GetLetterFromNumber(
                    i + Standard_Functions.GetNumberFromLetter(StartColumn) + 1)
                NoColumns = NoColumns + i

                # If the cell is not empty, add the contents to the list
                if ws[f"{Column}{StartRow}"].value is not None:
                    ReturnNamesExcel.append(str(ws[f"{Column}{StartRow}"].value))

                    for j in range(1, 1000):
                        # check if you reached the end of the data.
                        test = sheet.getContents(f"A{str(j)}")
                        if test == "" or test is None:
                            break

                        if sheet.getContents(f"A{str(j)}") == str(ws[f"{Column}{StartRow}"].value):
                            sheet.set(f"B{j}", "-")
                            sheet.set(f"E{j}", "Property value retrieved from drawing list")

                # If the cell is empty, you are on the end. Break the loop.
                if ws[f"{Column}{StartRow}"].value is None:
                    break

            # Get the pages in the document
            pages = App.ActiveDocument.findObjects("TechDraw::DrawPage")

            # Go through the excel list and collect the property value based on the property name searched for.
            for i in range(1000):
                # Define the start row. This is the Header row + 1 + i as counter
                RowNumber = int(StartRow) + i + 1

                # If the property name in the drawing list is empty, it is the end of the list.
                # Exit the function
                if ws[f"{StartColumn}{RowNumber}"].value is None:
                    break

                # Get the property name in the excel list. If it starts with "'", remove it
                PropertyValueExcel = str(ws[f"{StartColumn}{RowNumber}"].value)
                if PropertyValueExcel[:1] == "'":
                    PropertyValueExcel = PropertyValueExcel[1:]

                # Go through the columns starting from the column right from the column with the property value
                for j in range(len(ReturnNamesExcel)):
                    # Get the column letter
                    Column = Standard_Functions.GetLetterFromNumber(
                        j + Standard_Functions.GetNumberFromLetter(StartColumn) + 1)

                    # If the cell is not empty and j is lower then NoColumns, continue.
                    if ws[f"{Column}{RowNumber}"].value is not None:
                        # Get the property value in the excel list. If it starts with "'", remove it
                        ReturnValueExcel = str(ws[f"{Column}{RowNumber}"].value)
                        if ReturnValueExcel[:1] == "'":
                            ReturnValueExcel = ReturnValueExcel[1:]

                        # If page names are not to be mapped, go here
                        if USE_PAGE_NAMES_SIMPLE_LIST is False:
                            # Go through all the pages
                            for page in pages:
                                # Get the editable texts
                                texts = page.Template.EditableTexts
                                # If the value editable text in the titleblock to search for
                                # matches the property value in the drawing list,
                                # fill in the editable text that needs to be updated.
                                if texts[PROPERTY_NAME_SIMPLE_LIST] == PropertyValueExcel:
                                    # Get the property name in the titleblock spreadsheet and fill it with the property value from the drawing list
                                    texts[ReturnNamesExcel[j]] = ReturnValueExcel

                                    # Write all the updated text to the page.
                                    page.Template.EditableTexts = texts
                                    page.recompute()

                        # If page names are to be mapped, go here
                        if USE_PAGE_NAMES_SIMPLE_LIST is True:
                            # Go through the pages
                            for page in pages:
                                # If the Property name in the drawing list matches the page label:
                                # Fill in the desired editable text with the property value from the drawing list
                                if PropertyValueExcel == page.Label:
                                    # Get the editable texts
                                    texts = page.Template.EditableTexts
                                    texts[ReturnNamesExcel[j]] = ReturnValueExcel

                                    # Write all the updated text to the page.
                                    page.Template.EditableTexts = texts
                                    page.recompute()

            # Recomute the document
            App.ActiveDocument.recompute(None, True, True)

        except Exception as e:
            Text = translate(
                "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!\n"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: an error occurred!!\n"
                    + "See the report view for details",
                )
                raise e
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        finally:
            # Close the excel workbook
            wb.close()
    else:
        Text = translate("TitleBlock Workbench", "Use of a simple drawing list is not enabled!")
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    return


def MapSimpleDrawingList_FreeCAD(sheet):
    # Check if it is allowed to use an external source and if so, continue
    if USE_SIMPLE_LIST is True and USE_EXTERNAL_SOURCE_SIMPLE_LIST is True:
        # if debug mode is enabled, show the external file including path.
        if ENABLE_DEBUG is True:
            Text = translate("TitleBlock Workbench", f"The drawing list is: {EXTERNAL_FILE_SIMPLE_LIST}")
            Standard_Functions.Print(Text, "Log")
        # Get the active document
        doc = App.ActiveDocument
        # Save the name of the active document to reactivate it at the end of this function.
        LastActiveDoc = doc.Name
        # Get the name of the external source
        Input_SheetName = SHEETNAME_SIMPLE_LIST
        # Define the External sheet and document
        DrawingList = None
        # ff = None

        # try to open the source. if not show an messagebox and if debug mode is enabled, show the exeption as well
        try:
            if EXTERNAL_FILE_SIMPLE_LIST.lower().endswith(".xlsx"):
                Filter = [
                    ("FreeCAD", "*.FCStd"),
                ]
                FileName = Standard_Functions.GetFileDialog(
                    files=Filter, SaveAs=False
                )
                if FileName != "":
                    ff = App.openDocument(FileName, True)
                if FileName == "":
                    return
            else:
                ff = App.openDocument(EXTERNAL_FILE_SIMPLE_LIST, True)
            if EXTERNAL_FILE_SIMPLE_LIST == "":
                # Set the sheetname with a inputbox
                Spreadsheet_List = ff.findObjects("Spreadsheet::Sheet")
                Text = translate(
                    "TitleBlock Workbench",
                    "Please enter the name of the spreadsheet",
                )
                Input_SheetName = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=21,
                        default="Settings",
                        stringList=Spreadsheet_List,
                    )
                )
                # if the user canceled, exit this function.
                if not Input_SheetName.strip():
                    return

                DrawingList = ff.getObject(Input_SheetName)
            if SHEETNAME_SIMPLE_LIST != "":
                # Get the list of worksheets
                Spreadsheet_List = ff.findObjects("Spreadsheet::Sheet")
                # Assume that the worksheet doesn't exits
                SpreadsheetExits = False
                # Go through the names of the worksheets
                for Spreadsheet in Spreadsheet_List:
                    # If a name matches the sheetname in preferences, set WorksheetsExits to True
                    if Spreadsheet.Label == SHEETNAME_SIMPLE_LIST:
                        SpreadsheetExits is True
                # If WorksheetsExits is true, define the worksheet
                if SpreadsheetExits is True:
                    DrawingList = ff.getObject(Input_SheetName)
                # If WorksheetsExits is false, ask the user to enter the correct name.
                if SpreadsheetExits is False:
                    # Set the sheetname with a inputbox
                    Text = translate(
                        "TitleBlock Workbench",
                        "The spreadsheet doesn't exits!\nPlease enter the name of the spreadsheet")
                    Input_SheetName = str(
                        Standard_Functions.Mbox(
                            text=Text,
                            title="TitleBlock Workbench",
                            style=21,
                            default="TitleBlockData",
                            stringList=Spreadsheet_List,
                        )
                    )
                    # if the user canceled, exit this function.
                    if not Input_SheetName.strip():
                        return
                    # Define the worksheets
                    DrawingList = ff.getObject(Input_SheetName)
                    # Save the sheetname to the preferences
                    preferences.SetString("SheetName_SimpleList", Input_SheetName)
        except Exception as e:
            if ENABLE_DEBUG is True:
                raise (e)
            Text = translate(
                "TitleBlock Workbench",
                "an problem occured while openening the FreeCAD file!\nDo you have it open in an another application",
            )
            Standard_Functions.Mbox(
                text=Text, title="TitleBlock Workbench", style=0
            )
            return

        try:
            # Get the startcolumn and the other three columns from there
            StartCell = STARTCELL_SIMPLE_LIST
            if STARTCELL_SIMPLE_LIST == "":
                # Set EXTERNAL_SOURCE_SHEET_NAME to the chosen sheetname
                preferences.SetString("SheetName", Input_SheetName)
                DrawingList = ff.getObject(Input_SheetName)
                # Set the startcell with an inputbox
                Text = translate(
                    "TitleBlock Workbench",
                    "Please enter the name of the cell.\n"
                    + "Enter a single cell like 'A1', 'B2', etc. Other notations will be ignored!",
                )
                StartCell = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=20,
                        default="A1",
                    )
                )
                if not StartCell.strip():
                    StartCell = "A1"

                # Set SHEETNAME_STARTCELL to the chosen sheetname
                preferences.SetString("StartCell", StartCell)

            # If the StartCell is R1C1 style, change it
            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)

            # Get the start column
            StartColumn_DrawingList = Standard_Functions.RemoveNumbersFromString(StartCell)

            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column for the drawing list is: " + str(StartColumn_DrawingList),
                    ),
                    "Log",
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number for the drawing list is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList)),
                    ),
                    "Log",
                )

            # Get the start row
            StartRow_DrawingList = Standard_Functions.RemoveLettersFromString(
                StartCell
            )
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row for the drawing list is: " + str(StartRow_DrawingList)
                )
                Standard_Functions.Print(Text, "Log")

            # Define placeholder for the property value in the drawing list.
            PropertyValue_DrawingList = ""
            # Define a placeholder for the qty of columns with return values
            NoColumns = 0
            # Create a list with return values
            ReturnNames_DrawingList = []
            for i in range(1000):
                # Get the colum letter. For example: i = 0, StartColumn = A->1 + 1 results in column B
                Column = Standard_Functions.GetLetterFromNumber(
                    i + Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList) + 1)
                NoColumns = NoColumns + i

                # If the cell is not empty, add the contents to the list
                if str(DrawingList.getContents(f"{Column}{StartRow_DrawingList}")) != "":
                    ReturnName = str(DrawingList.getContents(f"{Column}{StartRow_DrawingList}"))
                    if ReturnName[:1] == "'":
                        ReturnName = ReturnName[1:]
                    ReturnNames_DrawingList.append(ReturnName)

                    for j in range(1, 1000):
                        # check if you reached the end of the data.
                        TitleBlockPropName = sheet.getContents(f"A{str(j)}")
                        if TitleBlockPropName[:1] == "'":
                            TitleBlockPropName = TitleBlockPropName[1:]
                        if TitleBlockPropName == "" or TitleBlockPropName is None:
                            break

                        if TitleBlockPropName == ReturnName:
                            sheet.set(f"B{j}", "-")
                            sheet.set(f"E{j}", "Property value retrieved from drawing list")

                # If the cell is empty, you are on the end. Break the loop.
                if str(DrawingList.getContents(f"{Column}{StartRow_DrawingList}")) == "":
                    break

            # Get the pages in the document
            pages = doc.findObjects("TechDraw::DrawPage")

            # Go through the drawing list and collect the property value based on the property name searched for.
            for i in range(1000):
                # Define the start row. This is the Header row + 1 + i as counter
                RowNumber = int(StartRow_DrawingList) + i + 1

                # If the property name in the drawing list is empty, it is the end of the list.
                # Exit the function
                if str(DrawingList.getContents(f"{StartColumn_DrawingList}{RowNumber}")) == "":
                    break

                # Get the property name in the excel list. If it starts with "'", remove it
                PropertyValue_DrawingList = str(DrawingList.getContents(f"{StartColumn_DrawingList}{RowNumber}"))
                if PropertyValue_DrawingList[:1] == "'":
                    PropertyValue_DrawingList = PropertyValue_DrawingList[1:]

                # Go through the columns starting from the column right from the column with the property value
                for j in range(len(ReturnNames_DrawingList)):
                    # Get the column letter
                    Column = Standard_Functions.GetLetterFromNumber(
                        j + Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList) + 1)

                    # If the cell is not empty and j is lower then NoColumns, continue.
                    if str(DrawingList.getContents(f"{Column}{RowNumber}")) != "":
                        # Get the property value in the excel list. If it starts with "'", remove it
                        ReturnValue_DrawingList = str(DrawingList.getContents(f"{Column}{RowNumber}"))
                        if ReturnValue_DrawingList[:1] == "'":
                            ReturnValue_DrawingList = ReturnValue_DrawingList[1:]

                        # If page names are not to be mapped, go here
                        if USE_PAGE_NAMES_SIMPLE_LIST is False:
                            # Go through all the pages
                            for page in pages:
                                # Get the editable texts
                                texts = page.Template.EditableTexts
                                # If the value editable text in the titleblock to search for
                                # matches the property value in the drawing list,
                                # fill in the editable text that needs to be updated.
                                if texts[PROPERTY_NAME_SIMPLE_LIST] == PropertyValue_DrawingList:
                                    # Get the property name in the titleblock spreadsheet and fill it with the property value from the drawing list
                                    texts[ReturnNames_DrawingList[j]] = ReturnValue_DrawingList

                                    # Write all the updated text to the page.
                                    page.Template.EditableTexts = texts
                                    page.recompute()

                        # If page names are to be mapped, go here
                        if USE_PAGE_NAMES_SIMPLE_LIST is True:
                            # Go through the pages
                            for page in pages:
                                # If the Property name in the drawing list matches the page label:
                                # Fill in the desired editable text with the property value from the drawing list
                                if PropertyValue_DrawingList == page.Label:
                                    # Get the editable texts
                                    texts = page.Template.EditableTexts
                                    # Get the property name in the titleblock spreadsheet and fill it with the property value from the drawing list
                                    texts[ReturnNames_DrawingList[j]] = ReturnValue_DrawingList

                                    # Write all the updated text to the page.
                                    page.Template.EditableTexts = texts
                                    page.recompute()

            # ----------------------------------------------------------------------------------------------
            # recompute the document
            doc.recompute(None, True, True)
            # Save the workbook
            doc.save()
            # Close the FreeCAD file
            App.closeDocument(ff.Name)
            # Activate the document which was active when this command started.
            try:
                App.setActiveDocument(LastActiveDoc)
            except Exception:
                pass

        except Exception as e:
            Text = translate(
                "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!\n"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: an error occurred!!\n"
                    + "See the report view for details",
                )
                raise e
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    else:
        Text = translate("TitleBlock Workbench", "Use of a simple drawing list is not enabled!")
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    return


def MapAdvancedDrawingList(doc, sheet):
    # Check if it is allowed to use an external source and if so, continue
    if USE_ADVANCED_LIST is True and USE_EXTERNAL_SOURCE_ADVANCED_LIST is False:
        try:
            # Get the active document
            doc = App.ActiveDocument
            # Save the name of the active document to reactivate it at the end of this function.
            LastActiveDoc = doc.Name
            # Get the Drawing list.
            DrawingList = doc.getObject(SHEETNAME_ADVANCED_LIST)
            if DrawingList is None:
                SheetName = Standard_Functions.Mbox(text="DrawingList", title="Title block workbench", style=20)
                if SheetName != "":
                    doc.addObject("Spreadsheet::Sheet", SheetName)
                    DrawingList = doc.getObject(SheetName)
                if SheetName == "":
                    return

            # Get folders with spreadsheets and add them to a list
            # Define a list for the groups
            GroupList = []
            # Get all the objects in the documents
            docObjects = doc.Objects
            # Go through the objects
            for i in range(len(docObjects)):
                # Create a bool for detecting if a folder has a page.
                HasPage = False
                # If the object is a group, continue
                if docObjects[i].TypeId == "App::DocumentObjectGroup":
                    # Go through the objects in the group
                    for j in range(len(docObjects[i].Group)):
                        # If the object is a page, set HasPage to True
                        if docObjects[i].Group[j].TypeId == "TechDraw::DrawPage":
                            HasPage = True
                # If HasPage is true, this is a group with page(s). Add it to the group list.
                if HasPage is True:
                    GroupList.append(docObjects[i])

            # Get the startcolumn and the other three columns from there
            StartCell = STARTCELL_ADVANCED_LIST

            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)
            StartColumn = Standard_Functions.RemoveNumbersFromString(StartCell)
            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column for the drawing list is: " + str(StartColumn),
                    ),
                    "Log",
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number for the drawing list is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumn)),
                    ),
                    "Log",
                )

            # Get the start row
            StartRow = Standard_Functions.RemoveLettersFromString(
                StartCell
            )
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row for the drawing list is: " + str(StartRow)
                )
                Standard_Functions.Print(Text, "Log")

            # Get the start column
            StartColumn_DrawingList = Standard_Functions.RemoveNumbersFromString(StartCell)

            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column for the drawing list is: " + str(StartColumn_DrawingList),
                    ),
                    "Log",
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number for the drawing list is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList)),
                    ),
                    "Log",
                )

            # Get the start row
            StartRow_DrawingList = int(Standard_Functions.RemoveLettersFromString(StartCell))
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row for the drawing list is: " + str(StartRow_DrawingList)
                )
                Standard_Functions.Print(Text, "Log")

            # Get the cell adresses for the cell with the group names
            # Define a dict for the adresses
            ExtSourceGroupAdress = []
            # Go through the group list
            i = 0
            j = 0
            for i in range(len(GroupList)):
                # Go through the first column in the drawing list.
                for j in range(1, 1000):
                    # Get the cell value
                    CellValue = str(
                        DrawingList.getContents(f"{StartColumn_DrawingList}{StartRow_DrawingList + j}")
                    )
                    if CellValue[:1] == "'":
                        CellValue = CellValue[1:]

                    # If the cell value is equal to the group label, this is the cell with a group name.
                    # Add it to the adress dict.
                    if GroupList[i].Label == CellValue:
                        ExtSourceGroupAdress.append(
                            [GroupList[i].Label, f"{StartColumn_DrawingList}{StartRow_DrawingList+j}"])

                    # If the cell value is empty, you are at the end of the drawing list.
                    if DrawingList.getContents(f"{StartColumn_DrawingList}{StartRow_DrawingList + j}") != "":
                        break
            # Add an endrow, so you can determine the last range in the next function
            ExtSourceGroupAdress.append(["EndRow", f"{StartColumn_DrawingList}{StartRow_DrawingList+j}"])

            # Create ranges for the different groups, where the function must search.
            NewList = []
            for i in range(len(ExtSourceGroupAdress) - 1):
                # Get the start cell. This is one row below the group cell
                StartCell_Range = ExtSourceGroupAdress[i][1]
                StartCell_Range = f"{Standard_Functions.RemoveNumbersFromString(StartCell_Range)}" + \
                    f"{str(int(Standard_Functions.RemoveLettersFromString(StartCell_Range)) + 1)}"
                # Get the end cell.
                EndCell_Range = ExtSourceGroupAdress[i + 1][1]

                # Add the end cell to the corresponding list item.
                NewList.append([ExtSourceGroupAdress[i][0], StartCell_Range, EndCell_Range])
            ExtSourceGroupAdress = NewList

            # Go through the grouplist
            for Group in GroupList:
                # Go through the list with adresses fro each group
                for j in range(len(ExtSourceGroupAdress)):
                    if Group.Label == ExtSourceGroupAdress[j][0]:
                        # Define the start, end and column
                        StartRow_Range = int(Standard_Functions.RemoveLettersFromString(ExtSourceGroupAdress[j][1]))
                        EndRow_Range = int(Standard_Functions.RemoveLettersFromString(ExtSourceGroupAdress[j][2]))
                        StartRow_DrawingList = int(Standard_Functions.RemoveLettersFromString(StartCell))
                        StartColumn_DrawingList = Standard_Functions.RemoveNumbersFromString(StartCell)

                        # Get the pages in the group
                        pages = []
                        for k in range(len(Group.Group)):
                            if Group.Group[k].TypeId == "TechDraw::DrawPage":
                                pages.append(Group.Group[k])

                        # Go through the range of the group
                        for k in range(StartRow_Range, EndRow_Range):
                            # Define the rownumber. This is the Header row + 1 + i as counter
                            RowNumber = k
                            # Define placeholder for the property value in the drawing list.
                            PropertyValueExt = ""
                            # Define a placeholder for the qty of columns with return values
                            NoColumns = 0
                            # Create a list with return values
                            ReturnNamesExt = []
                            for i in range(1000):
                                # Get the colum letter. For example: i = 0, StartColumn = A->1 + 1 results in column B
                                Column = Standard_Functions.GetLetterFromNumber(
                                    i + Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList) + 1)
                                NoColumns = NoColumns + i

                                # If the cell is not empty, add the contents to the list
                                if DrawingList.getContents(f"{Column}{StartRow_DrawingList}") != "":
                                    CellValue = str(DrawingList.getContents(f"{Column}{StartRow_DrawingList}"))
                                    if CellValue[:1] == "'":
                                        CellValue = CellValue[1:]
                                    ReturnNamesExt.append(CellValue)

                                    for j in range(1, 1000):
                                        # check if you reached the end of the data.
                                        test = sheet.getContents(f"A{str(j)}")
                                        if test == "" or test is None:
                                            break

                                        if sheet.getContents(f"A{str(j)}") == str(
                                                DrawingList.getContents(f"{Column}{StartRow_DrawingList}")):
                                            sheet.set(f"B{j}", "-")
                                            sheet.set(f"E{j}", "Property value retrieved from drawing list")

                                # If the cell is empty, you are on the end. Break the loop.
                                if DrawingList.getContents(f"{Column}{StartRow_DrawingList}") != "":
                                    break

                            # Get the property name in the drawing list. If it starts with "'", remove it
                            PropertyValueExt = str(DrawingList.getContents(f"{StartColumn_DrawingList}{RowNumber}"))
                            if PropertyValueExt[:1] == "'":
                                PropertyValueExt = PropertyValueExt[1:]
                            # If a prefix is used for sorting the groups in the tree, remove it from the Property
                            if SORTING_PREFIX_ADVANCED_LIST != "" or SORTING_PREFIX_ADVANCED_LIST is not None:
                                lengthPrefix = len(SORTING_PREFIX_ADVANCED_LIST)
                                PropertyValueExt = PropertyValueExt[lengthPrefix:]

                            # Go through the columns starting from the column right from the column with the property value
                            for j in range(len(ReturnNamesExt)):
                                # Get the column letter
                                Column = Standard_Functions.GetLetterFromNumber(
                                    j + Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList) + 1)

                                # If the cell is not empty and j is lower then NoColumns, continue.
                                if DrawingList.getContents(f"{Column}{RowNumber}") is not None:
                                    # Get the property value in the excel list. If it starts with "'", remove it
                                    ReturnValueExcel = str(DrawingList.getContents(f"{Column}{RowNumber}"))
                                    if ReturnValueExcel[:1] == "'":
                                        ReturnValueExcel = ReturnValueExcel[1:]

                                    # If page names are not to be mapped, go here
                                    if USE_PAGE_NAMES_ADVANCED_LIST is False:
                                        # Go through all the pages
                                        for page in pages:
                                            # Get the editable texts
                                            texts = page.Template.EditableTexts
                                            # If the value editable text in the titleblock to search for
                                            # matches the property value in the drawing list,
                                            # fill in the editable text that needs to be updated.
                                            if texts[PROPERTY_NAME_ADVANCED_LIST] == PropertyValueExt:
                                                # Get the property name in the titleblock spreadsheet
                                                # and fill it with the property value from the drawing list
                                                texts[ReturnNamesExt[j]] = ReturnValueExcel

                                                # Write all the updated text to the page.
                                                page.Template.EditableTexts = texts
                                                page.recompute()

                                    # If page names are to be mapped, go here
                                    if USE_PAGE_NAMES_ADVANCED_LIST is True:
                                        # Go through the pages
                                        for page in pages:
                                            # If the Property name in the drawing list matches the page label:
                                            # Fill in the desired editable text with the property value from the drawing list
                                            if PropertyValueExt == page.Label:
                                                # Get the editable texts
                                                texts = page.Template.EditableTexts
                                                texts[ReturnNamesExt[j]] = ReturnValueExcel

                                                # Write all the updated text to the page.
                                                page.Template.EditableTexts = texts
                                                page.recompute()

            # ----------------------------------------------------------------------------------------------
            # recompute the document
            doc.recompute(None, True, True)
            # Save the workbook
            doc.save()
            # Activate the document which was active when this command started.
            try:
                App.setActiveDocument(LastActiveDoc)
            except Exception:
                pass

        except Exception as e:
            Text = translate(
                "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!\n"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: an error occurred!!\n"
                    + "See the report view for details",
                )
                raise e
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    else:
        Text = translate("TitleBlock Workbench", "Use of an advanced drawing list is not enabled!")
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    return


def MapAdvancedDrawingList_Excel(doc, sheet):
    from openpyxl import load_workbook

    # Check if it is allowed to use an external source and if so, continue
    if USE_ADVANCED_LIST is True and USE_EXTERNAL_SOURCE_ADVANCED_LIST is True:
        # if debug mode is enabled, show the external file including path.
        if ENABLE_DEBUG is True:
            Text = translate("TitleBlock Workbench", f"The drawing list is: {EXTERNAL_FILE_ADVANCED_LIST}")
            Standard_Functions.Print(Text, "Log")

        # try to open the source. if not show an messagebox and if debug mode is enabled, show the exeption as well
        try:
            # Define a placeholder for the workbook
            wb = ""
            # If the drawinglist is an FreeCAD document instead of an Excel workbook,
            # Let the user select the correct workbook. Otherwise just load the workbook.
            if EXTERNAL_FILE_ADVANCED_LIST.lower().endswith(".fcstd"):
                Filter = [
                    ("Excel", "*.xlsx"),
                ]
                FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=False)
                if FileName != "":
                    wb = load_workbook(FileName, read_only=True, data_only=True)
                if FileName == "":
                    return
            else:
                wb = load_workbook(
                    str(EXTERNAL_FILE_ADVANCED_LIST), read_only=True, data_only=True
                )
            # If the sheetname not set, let the user set the correct name.
            if SHEETNAME_ADVANCED_LIST == "":
                # Set the sheetname with a inputbox
                Worksheets_List = [i for i in wb.sheetnames if i != "Settings"]
                Text = translate(
                    "TitleBlock Workbench", "Please enter the name of the worksheet"
                )
                Input_SheetName = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=3,
                        default="TitleBlockData",
                        stringList=Worksheets_List,
                    )
                )
                # if the user canceled, exit this function.
                if not Input_SheetName.strip():
                    return
                ws = wb[str(Input_SheetName)]
            if SHEETNAME_ADVANCED_LIST != "":
                # Get the list of worksheets
                Worksheets_List = [i for i in wb.sheetnames if i != "Settings"]
                # Assume that the worksheet doesn't exits
                WorksheetExits = False
                # Go through the names of the worksheets
                for WorksheetName in Worksheets_List:
                    # If a name matches the sheetname in preferences, set WorksheetsExits to True
                    if WorksheetName == SHEETNAME_ADVANCED_LIST:
                        WorksheetExits is True
                # If WorksheetsExits is true, define the worksheet
                if WorksheetExits is True:
                    ws = wb[str(SHEETNAME_ADVANCED_LIST)]
                # If WorksheetsExits is false, ask the user to enter the correct name.
                if WorksheetExits is False:
                    # Set the sheetname with a inputbox
                    Text = translate(
                        "TitleBlock Workbench", "The worksheet doesn't exits!\nPlease enter the name of the worksheet"
                    )
                    Input_SheetName = str(
                        Standard_Functions.Mbox(
                            text=Text,
                            title="TitleBlock Workbench",
                            style=21,
                            default="TitleBlockData",
                            stringList=Worksheets_List,
                        )
                    )
                    # if the user canceled, exit this function.
                    if not Input_SheetName.strip():
                        return
                    # Define the worksheets
                    ws = wb[str(Input_SheetName)]
                    # Save the sheetname to preferences
                    preferences.SetString("SheetName_AdvancedList", Input_SheetName)
        except IOError:
            Standard_Functions.Mbox(
                "Permission error!!\nDo you have the file open?",
                "Titleblock Workbench",
                0,
                IconType="Critical",
            )
            return
        except Exception as e:
            if ENABLE_DEBUG is True:
                raise (e)
            Text = translate(
                "TitleBlock Workbench",
                "an problem occured while openening the excel file!\nDo you have it open in an another application",
            )
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
            return

        try:
            # Get the spreadsheet.
            if sheet is None:
                sheet = doc.getObject("TitleBlock")

            # Get folders with spreadsheets and add them to a list
            # Define a list for the groups
            GroupList = []
            # Get all the objects in the documents
            docObjects = doc.Objects
            # Go through the objects
            for i in range(len(docObjects)):
                # Create a bool for detecting if a folder has a page.
                HasPage = False
                # If the object is a group, continue
                if docObjects[i].TypeId == "App::DocumentObjectGroup":
                    # Go through the objects in the group
                    for j in range(len(docObjects[i].Group)):
                        # If the object is a page, set HasPage to True
                        if docObjects[i].Group[j].TypeId == "TechDraw::DrawPage":
                            HasPage = True
                # If HasPage is true, this is a group with page(s). Add it to the group list.
                if HasPage is True:
                    GroupList.append(docObjects[i])

            # Get the startcolumn and the other three columns from there
            StartCell = STARTCELL_ADVANCED_LIST
            if SHEETNAME_ADVANCED_LIST == "":
                # Save drawinglist and sheetname to the preferences
                preferences.SetString("ExternalFile", FileName)
                preferences.SetString("SheetName_AdvancedList", Input_SheetName)
                ws = wb[Input_SheetName]

                # Set the startcell with an inputbox
                Text = translate(
                    "TitleBlock Workbench",
                    "Please enter the name of the cell.\n"
                    + "Enter a single cell like 'A1', 'B2', etc. Other notations will be ignored!",
                )
                StartCell = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=20,
                        default="A1",
                    )
                )
                if not StartCell.strip():
                    StartCell = "A1"

                # Set SHEETNAME_STARTCELL to the chosen sheetname
                preferences.SetString("StartCell_AdvancedList", StartCell)

            # If the StartCell is R1C1 style, change it
            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)

            # Get the start column
            StartColumn_DrawingList = Standard_Functions.RemoveNumbersFromString(StartCell)
            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column for the drawing list is: " + str(StartColumn_DrawingList),
                    ),
                    "Log",
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number for the drawing list is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList)),
                    ),
                    "Log",
                )

            # Get the start row
            StartRow_DrawingList = int(Standard_Functions.RemoveLettersFromString(StartCell))
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row for the drawing list is: " + str(StartRow_DrawingList)
                )
                Standard_Functions.Print(Text, "Log")

            # Get the cell adresses for the cell with the group names
            # Define a dict for the adresses
            ExtSourceGroupAdress = []
            # Go through the group list
            i = 0
            j = 0
            for i in range(len(GroupList)):
                # Go through the first column in the drawing list.
                for j in range(1, 1000):
                    # Get the cell value
                    CellValue = str(ws[f"{StartColumn_DrawingList}{StartRow_DrawingList + j}"].value)

                    # If the cell value is equal to the group label, this is the cell with a group name.
                    # Add it to the adress dict.
                    if GroupList[i].Label == CellValue:
                        ExtSourceGroupAdress.append(
                            [GroupList[i].Label, f"{StartColumn_DrawingList}{StartRow_DrawingList+j}"])

                    # If the cell value is empty, you are at the end of the drawing list.
                    if ws[f"{StartColumn_DrawingList}{StartRow_DrawingList + j}"].value is None:
                        break
            # Add an endrow, so you can determine the last range in the next function
            ExtSourceGroupAdress.append(["EndRow", f"{StartColumn_DrawingList}{StartRow_DrawingList+j}"])

            # Create ranges for the different groups, where the function must search.
            NewList = []
            for i in range(len(ExtSourceGroupAdress) - 1):
                # Get the start cell. This is one row below the group cell
                StartCell_Range = ExtSourceGroupAdress[i][1]
                StartCell_Range = f"{Standard_Functions.RemoveNumbersFromString(StartCell_Range)}" + \
                    f"{str(int(Standard_Functions.RemoveLettersFromString(StartCell_Range)) + 1)}"
                # Get the end cell.
                EndCell_Range = ExtSourceGroupAdress[i + 1][1]

                # Add the end cell to the corresponding list item.
                NewList.append([ExtSourceGroupAdress[i][0], StartCell_Range, EndCell_Range])
            ExtSourceGroupAdress = NewList

            # Go through the grouplist
            for Group in GroupList:
                # Go through the list with adresses fro each group
                for j in range(len(ExtSourceGroupAdress)):
                    if Group.Label == ExtSourceGroupAdress[j][0]:
                        # Define the start, end and column
                        StartRow_Range = int(Standard_Functions.RemoveLettersFromString(ExtSourceGroupAdress[j][1]))
                        EndRow_Range = int(Standard_Functions.RemoveLettersFromString(ExtSourceGroupAdress[j][2]))
                        StartRow_DrawingList = int(Standard_Functions.RemoveLettersFromString(StartCell))
                        StartColumn_DrawingList = Standard_Functions.RemoveNumbersFromString(StartCell)

                        # Get the pages in the group
                        pages = []
                        for k in range(len(Group.Group)):
                            if Group.Group[k].TypeId == "TechDraw::DrawPage":
                                pages.append(Group.Group[k])

                        # Go through the range of the group
                        for k in range(StartRow_Range, EndRow_Range):
                            # Define the rownumber. This is the Header row + 1 + i as counter
                            RowNumber = k
                            # Define placeholder for the property value in the drawing list.
                            PropertyValueExcel = ""
                            # Define a placeholder for the qty of columns with return values
                            NoColumns = 0
                            # Create a list with return values
                            ReturnNamesExcel = []
                            for i in range(1000):
                                # Get the colum letter. For example: i = 0, StartColumn = A->1 + 1 results in column B
                                Column = Standard_Functions.GetLetterFromNumber(
                                    i + Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList) + 1)
                                NoColumns = NoColumns + i

                                # If the cell is not empty, add the contents to the list
                                if ws[f"{Column}{StartRow_DrawingList}"].value is not None:
                                    ReturnNamesExcel.append(str(ws[f"{Column}{StartRow_DrawingList}"].value))

                                    for j in range(1, 1000):
                                        # check if you reached the end of the data.
                                        test = sheet.getContents(f"A{str(j)}")
                                        if test == "" or test is None:
                                            break

                                        if sheet.getContents(f"A{str(j)}") == str(
                                                ws[f"{Column}{StartRow_DrawingList}"].value):
                                            sheet.set(f"B{j}", "-")
                                            sheet.set(f"E{j}", "Property value retrieved from drawing list")

                                # If the cell is empty, you are on the end. Break the loop.
                                if ws[f"{Column}{StartRow_DrawingList}"].value is None:
                                    break

                            # Get the property name in the drawing list. If it starts with "'", remove it
                            PropertyValueExcel = str(ws[f"{StartColumn_DrawingList}{RowNumber}"].value)
                            if PropertyValueExcel[:1] == "'":
                                PropertyValueExcel = PropertyValueExcel[1:]
                            # If a prefix is used for sorting the groups in the tree, remove it from the Property
                            if SORTING_PREFIX_ADVANCED_LIST != "" or SORTING_PREFIX_ADVANCED_LIST is not None:
                                lengthPrefix = len(SORTING_PREFIX_ADVANCED_LIST)
                                PropertyValueExcel = PropertyValueExcel[lengthPrefix:]

                            # Go through the columns starting from the column right from the column with the property value
                            for j in range(len(ReturnNamesExcel)):
                                # Get the column letter
                                Column = Standard_Functions.GetLetterFromNumber(
                                    j + Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList) + 1)

                                # If the cell is not empty and j is lower then NoColumns, continue.
                                if ws[f"{Column}{RowNumber}"].value is not None:
                                    # Get the property value in the excel list. If it starts with "'", remove it
                                    ReturnValueExcel = str(ws[f"{Column}{RowNumber}"].value)
                                    if ReturnValueExcel[:1] == "'":
                                        ReturnValueExcel = ReturnValueExcel[1:]

                                    # If page names are not to be mapped, go here
                                    if USE_PAGE_NAMES_ADVANCED_LIST is False:
                                        # Go through all the pages
                                        for page in pages:
                                            # Get the editable texts
                                            texts = page.Template.EditableTexts
                                            # If the value editable text in the titleblock to search for
                                            # matches the property value in the drawing list,
                                            # fill in the editable text that needs to be updated.
                                            if texts[PROPERTY_NAME_ADVANCED_LIST] == PropertyValueExcel:
                                                # Get the property name in the titleblock spreadsheet
                                                # and fill it with the property value from the drawing list
                                                texts[ReturnNamesExcel[j]] = ReturnValueExcel

                                                # Write all the updated text to the page.
                                                page.Template.EditableTexts = texts
                                                page.recompute()

                                    # If page names are to be mapped, go here
                                    if USE_PAGE_NAMES_ADVANCED_LIST is True:
                                        # Go through the pages
                                        for page in pages:
                                            # If the Property name in the drawing list matches the page label:
                                            # Fill in the desired editable text with the property value from the drawing list
                                            if PropertyValueExcel == page.Label:
                                                # Get the editable texts
                                                texts = page.Template.EditableTexts
                                                texts[ReturnNamesExcel[j]] = ReturnValueExcel

                                                # Write all the updated text to the page.
                                                page.Template.EditableTexts = texts
                                                page.recompute()

            # ----------------------------------------------------------------------------------------------
            # Recomute the document
            App.ActiveDocument.recompute(None, True, True)

        except Exception as e:
            Text = translate(
                "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!\n"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: an error occurred!!\n"
                    + "See the report view for details",
                )
                raise e
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        finally:
            # Close the excel workbook
            wb.close()
    else:
        Text = translate("TitleBlock Workbench", "Use of a advanced drawing list is not enabled!")
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    return


def MapAdvancedDrawingList_FreeCAD(doc, sheet):
    # Check if it is allowed to use an external source and if so, continue
    if USE_ADVANCED_LIST is True and USE_EXTERNAL_SOURCE_ADVANCED_LIST is True:
        # if debug mode is enabled, show the external file including path.
        if ENABLE_DEBUG is True:
            Text = translate("TitleBlock Workbench", f"The drawing list is: {EXTERNAL_FILE_ADVANCED_LIST}")
            Standard_Functions.Print(Text, "Log")
        # Get the active document
        doc = App.ActiveDocument
        # Save the name of the active document to reactivate it at the end of this function.
        LastActiveDoc = doc.Name
        # Get the name of the external source
        Input_SheetName = SHEETNAME_ADVANCED_LIST
        # Define the External sheet and document
        DrawingList = None
        # ff = None

        # try to open the source. if not show an messagebox and if debug mode is enabled, show the exeption as well
        try:
            if EXTERNAL_FILE_ADVANCED_LIST.lower().endswith(".xlsx"):
                Filter = [
                    ("FreeCAD", "*.FCStd"),
                ]
                FileName = Standard_Functions.GetFileDialog(
                    files=Filter, SaveAs=False
                )
                if FileName != "":
                    ff = App.openDocument(FileName, True)
                if FileName == "":
                    return
            else:
                ff = App.openDocument(EXTERNAL_FILE_ADVANCED_LIST, True)
            if EXTERNAL_FILE_ADVANCED_LIST == "":
                # Set the sheetname with a inputbox
                Spreadsheet_List = ff.findObjects("Spreadsheet::Sheet")
                Text = translate(
                    "TitleBlock Workbench",
                    "Please enter the name of the spreadsheet",
                )
                Input_SheetName = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=21,
                        default="Settings",
                        stringList=Spreadsheet_List,
                    )
                )
                # if the user canceled, exit this function.
                if not Input_SheetName.strip():
                    return

                DrawingList = ff.getObject(Input_SheetName)
            if SHEETNAME_ADVANCED_LIST != "":
                # Get the list of worksheets
                Spreadsheet_List = ff.findObjects("Spreadsheet::Sheet")
                # Assume that the worksheet doesn't exits
                SpreadsheetExits = False
                # Go through the names of the worksheets
                for Spreadsheet in Spreadsheet_List:
                    # If a name matches the sheetname in preferences, set WorksheetsExits to True
                    if Spreadsheet.Label == SHEETNAME_ADVANCED_LIST:
                        SpreadsheetExits is True
                # If WorksheetsExits is true, define the worksheet
                if SpreadsheetExits is True:
                    DrawingList = ff.getObject(Input_SheetName)
                # If WorksheetsExits is false, ask the user to enter the correct name.
                if SpreadsheetExits is False:
                    # Set the sheetname with a inputbox
                    Text = translate(
                        "TitleBlock Workbench",
                        "The spreadsheet doesn't exits!\nPlease enter the name of the spreadsheet")
                    Input_SheetName = str(
                        Standard_Functions.Mbox(
                            text=Text,
                            title="TitleBlock Workbench",
                            style=21,
                            default="TitleBlockData",
                            stringList=Spreadsheet_List,
                        )
                    )
                    # if the user canceled, exit this function.
                    if not Input_SheetName.strip():
                        return
                    # Define the worksheets
                    DrawingList = ff.getObject(Input_SheetName)
                    # Save the sheetname to preferences
                    preferences.SetString("SheetName_AdvancedList", Input_SheetName)
        except Exception as e:
            if ENABLE_DEBUG is True:
                raise (e)
            Text = translate(
                "TitleBlock Workbench",
                "an problem occured while openening the FreeCAD file!\nDo you have it open in an another application",
            )
            Standard_Functions.Mbox(
                text=Text, title="TitleBlock Workbench", style=0
            )
            return

        try:
            # Get the spreadsheet.
            if sheet is None:
                sheet = doc.getObject("TitleBlock")

            # Get folders with spreadsheets and add them to a list
            # Define a list for the groups
            GroupList = []
            # Get all the objects in the documents
            docObjects = doc.Objects
            # Go through the objects
            for i in range(len(docObjects)):
                # Create a bool for detecting if a folder has a page.
                HasPage = False
                # If the object is a group, continue
                if docObjects[i].TypeId == "App::DocumentObjectGroup":
                    # Go through the objects in the group
                    for j in range(len(docObjects[i].Group)):
                        # If the object is a page, set HasPage to True
                        if docObjects[i].Group[j].TypeId == "TechDraw::DrawPage":
                            HasPage = True
                # If HasPage is true, this is a group with page(s). Add it to the group list.
                if HasPage is True:
                    GroupList.append(docObjects[i])

            # Get the startcolumn and the other three columns from there
            StartCell = STARTCELL_ADVANCED_LIST
            if SHEETNAME_ADVANCED_LIST == "":
                # Set EXTERNAL_SOURCE_SHEET_NAME to the chosen sheetname
                preferences.SetString("SheetName", Input_SheetName)
                DrawingList = ff.getObject(Input_SheetName)
                # Set the startcell with an inputbox
                Text = translate(
                    "TitleBlock Workbench",
                    "Please enter the name of the cell.\n"
                    + "Enter a single cell like 'A1', 'B2', etc. Other notations will be ignored!",
                )
                StartCell = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=20,
                        default="A1",
                    )
                )
                if not StartCell.strip():
                    StartCell = "A1"

                # Set SHEETNAME_STARTCELL to the chosen sheetname
                preferences.SetString("StartCell", StartCell)

            # If the StartCell is R1C1 style, change it
            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)

            # Get the start column
            StartColumn_DrawingList = Standard_Functions.RemoveNumbersFromString(StartCell)

            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column for the drawing list is: " + str(StartColumn_DrawingList),
                    ),
                    "Log",
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number for the drawing list is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList)),
                    ),
                    "Log",
                )

            # Get the start row
            StartRow_DrawingList = int(Standard_Functions.RemoveLettersFromString(StartCell))
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row for the drawing list is: " + str(StartRow_DrawingList)
                )
                Standard_Functions.Print(Text, "Log")

            # Get the cell adresses for the cell with the group names
            # Define a dict for the adresses
            ExtSourceGroupAdress = []
            # Go through the group list
            i = 0
            j = 0
            for i in range(len(GroupList)):
                # Go through the first column in the drawing list.
                for j in range(1, 1000):
                    # Get the cell value
                    CellValue = str(
                        DrawingList.getContents(f"{StartColumn_DrawingList}{StartRow_DrawingList + j}")
                    )
                    if CellValue[:1] == "'":
                        CellValue = CellValue[1:]

                    # If the cell value is equal to the group label, this is the cell with a group name.
                    # Add it to the adress dict.
                    if GroupList[i].Label == CellValue:
                        ExtSourceGroupAdress.append(
                            [GroupList[i].Label, f"{StartColumn_DrawingList}{StartRow_DrawingList+j}"])

                    # If the cell value is empty, you are at the end of the drawing list.
                    if DrawingList.getContents(f"{StartColumn_DrawingList}{StartRow_DrawingList + j}") == "":
                        break
            # Add an endrow, so you can determine the last range in the next function
            ExtSourceGroupAdress.append(["EndRow", f"{StartColumn_DrawingList}{StartRow_DrawingList+j}"])

            # Create ranges for the different groups, where the function must search.
            NewList = []
            for i in range(len(ExtSourceGroupAdress) - 1):
                # Get the start cell. This is one row below the group cell
                StartCell_Range = ExtSourceGroupAdress[i][1]
                StartCell_Range = f"{Standard_Functions.RemoveNumbersFromString(StartCell_Range)}" + \
                    f"{str(int(Standard_Functions.RemoveLettersFromString(StartCell_Range)) + 1)}"
                # Get the end cell.
                EndCell_Range = ExtSourceGroupAdress[i + 1][1]

                # Add the end cell to the corresponding list item.
                NewList.append([ExtSourceGroupAdress[i][0], StartCell_Range, EndCell_Range])
            ExtSourceGroupAdress = NewList

            # Go through the grouplist
            for Group in GroupList:
                # Go through the list with adresses fro each group
                for j in range(len(ExtSourceGroupAdress)):
                    if Group.Label == ExtSourceGroupAdress[j][0]:
                        # Define the start, end and column
                        StartRow_Range = int(Standard_Functions.RemoveLettersFromString(ExtSourceGroupAdress[j][1]))
                        EndRow_Range = int(Standard_Functions.RemoveLettersFromString(ExtSourceGroupAdress[j][2]))
                        StartRow_DrawingList = int(Standard_Functions.RemoveLettersFromString(StartCell))
                        StartColumn_DrawingList = Standard_Functions.RemoveNumbersFromString(StartCell)

                        # Get the pages in the group
                        pages = []
                        for k in range(len(Group.Group)):
                            if Group.Group[k].TypeId == "TechDraw::DrawPage":
                                pages.append(Group.Group[k])

                        # Go through the range of the group
                        for k in range(StartRow_Range, EndRow_Range):
                            # Define the rownumber. This is the Header row + 1 + i as counter
                            RowNumber = k
                            # Define placeholder for the property value in the drawing list.
                            PropertyValueExt = ""
                            # Define a placeholder for the qty of columns with return values
                            NoColumns = 0
                            # Create a list with return values
                            ReturnNamesExt = []
                            for i in range(1000):
                                # Get the colum letter. For example: i = 0, StartColumn = A->1 + 1 results in column B
                                Column = Standard_Functions.GetLetterFromNumber(
                                    i + Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList) + 1)
                                NoColumns = NoColumns + i

                                # If the cell is not empty, add the contents to the list
                                if DrawingList.getContents(f"{Column}{StartRow_DrawingList}") != "":
                                    CellValue = str(DrawingList.getContents(f"{Column}{StartRow_DrawingList}"))
                                    if CellValue[:1] == "'":
                                        CellValue = CellValue[1:]
                                    ReturnNamesExt.append(CellValue)

                                    for j in range(1, 1000):
                                        # check if you reached the end of the data.
                                        test = sheet.getContents(f"A{str(j)}")
                                        if test == "" or test is None:
                                            break

                                        if sheet.getContents(f"A{str(j)}") == str(
                                                DrawingList.getContents(f"{Column}{StartRow_DrawingList}")):
                                            sheet.set(f"B{j}", "-")
                                            sheet.set(f"E{j}", "Property value retrieved from drawing list")

                                # If the cell is empty, you are on the end. Break the loop.
                                if DrawingList.getContents(f"{Column}{StartRow_DrawingList}") == "":
                                    break

                            # Get the property name in the drawing list. If it starts with "'", remove it
                            PropertyValueExt = str(DrawingList.getContents(f"{StartColumn_DrawingList}{RowNumber}"))
                            if PropertyValueExt[:1] == "'":
                                PropertyValueExt = PropertyValueExt[1:]
                            # If a prefix is used for sorting the groups in the tree, remove it from the Property
                            if SORTING_PREFIX_ADVANCED_LIST != "" or SORTING_PREFIX_ADVANCED_LIST is not None:
                                lengthPrefix = len(SORTING_PREFIX_ADVANCED_LIST)
                                PropertyValueExt = PropertyValueExt[lengthPrefix:]

                            # Go through the columns starting from the column right from the column with the property value
                            for j in range(len(ReturnNamesExt)):
                                # Get the column letter
                                Column = Standard_Functions.GetLetterFromNumber(
                                    j + Standard_Functions.GetNumberFromLetter(StartColumn_DrawingList) + 1)

                                # If the cell is not empty and j is lower then NoColumns, continue.
                                if DrawingList.getContents(f"{Column}{RowNumber}") != "":
                                    # Get the property value in the excel list. If it starts with "'", remove it
                                    ReturnValueExt = str(DrawingList.getContents(f"{Column}{RowNumber}"))
                                    if ReturnValueExt[:1] == "'":
                                        ReturnValueExt = ReturnValueExt[1:]

                                    # If page names are not to be mapped, go here
                                    if USE_PAGE_NAMES_ADVANCED_LIST is False:
                                        # Go through all the pages
                                        for page in pages:
                                            # Get the editable texts
                                            texts = page.Template.EditableTexts
                                            # If the value editable text in the titleblock to search for
                                            # matches the property value in the drawing list,
                                            # fill in the editable text that needs to be updated.
                                            if texts[PROPERTY_NAME_ADVANCED_LIST] == PropertyValueExt:
                                                # Get the property name in the titleblock spreadsheet
                                                # and fill it with the property value from the drawing list
                                                texts[ReturnNamesExt[j]] = ReturnValueExt

                                                # Write all the updated text to the page.
                                                page.Template.EditableTexts = texts
                                                page.recompute()

                                    # If page names are to be mapped, go here
                                    if USE_PAGE_NAMES_ADVANCED_LIST is True:
                                        # Go through the pages
                                        for page in pages:
                                            # If the Property name in the drawing list matches the page label:
                                            # Fill in the desired editable text with the property value from the drawing list
                                            if PropertyValueExt == page.Label:
                                                # Get the editable texts
                                                texts = page.Template.EditableTexts
                                                texts[ReturnNamesExt[j]] = ReturnValueExt

                                                # Write all the updated text to the page.
                                                page.Template.EditableTexts = texts
                                                page.recompute()

            # ----------------------------------------------------------------------------------------------
            # recompute the document
            doc.recompute(None, True, True)
            # Save the workbook
            doc.save()
            # Close the FreeCAD file
            App.closeDocument(ff.Name)
            # Activate the document which was active when this command started.
            try:
                App.setActiveDocument(LastActiveDoc)
            except Exception:
                pass

        except Exception as e:
            Text = translate(
                "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!\n"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: an error occurred!!\n"
                    + "See the report view for details",
                )
                raise e
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    else:
        Text = translate("TitleBlock Workbench", "Use of an advanced drawing list is not enabled!")
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    return


def SortGroups(Reverse: bool = False):
    # Get the active document
    doc = App.ActiveDocument

    # Get folders with spreadsheets and add them to a list
    # Define a list for the groups
    GroupList = []
    # Get all the objects in the documents
    docObjects = doc.Objects
    # Go through the objects
    for i in range(len(docObjects)):
        # Create a bool for detecting if a folder has a page.
        HasPage = False
        # If the object is a group, continue
        if docObjects[i].TypeId == "App::DocumentObjectGroup":
            # Go through the objects in the group
            for j in range(len(docObjects[i].Group)):
                # If the object is a page, set HasPage to True
                if docObjects[i].Group[j].TypeId == "TechDraw::DrawPage":
                    HasPage = True
        # If HasPage is true, this is a group with page(s). Add it to the group list.
        if HasPage is True:
            GroupList.append(docObjects[i])

    # Sort each group in the group list
    for Group in GroupList:
        Standard_Functions.sortGroup(group=Group, reverse=Reverse)

    # Recompute the document
    doc.recompute(None, True, True)
