# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/


import FreeCAD as App
import Standard_Functions_TitleBlock as Standard_Functions
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# Get the settings
import Settings
from Settings import IMPORT_SETTINGS_XL
from Settings import SPREADSHEET_COLUMNFONTSTYLE_UNDERLINE
from Settings import SPREADSHEET_COLUMNFONTSTYLE_ITALIC
from Settings import SPREADSHEET_COLUMNFONTSTYLE_BOLD
from Settings import SPREADSHEET_TABLEFONTSTYLE_UNDERLINE
from Settings import SPREADSHEET_TABLEFONTSTYLE_ITALIC
from Settings import SPREADSHEET_TABLEFONTSTYLE_BOLD
from Settings import SPREADSHEET_TABLEFOREGROUND
from Settings import SPREADSHEET_TABLEBACKGROUND_2
from Settings import SPREADSHEET_TABLEBACKGROUND_1
from Settings import SPREADSHEET_HEADERFONTSTYLE_UNDERLINE
from Settings import SPREADSHEET_HEADERFONTSTYLE_ITALIC
from Settings import SPREADSHEET_HEADERFONTSTYLE_BOLD
from Settings import SPREADSHEET_HEADERFOREGROUND
from Settings import SPREADSHEET_HEADERBACKGROUND
from Settings import AUTOFIT_FACTOR

# Define the translation
translate = App.Qt.translate

# region - supporting functions


# Function the return the correct string to use in the FormatTable function
def FontStyle(Bold: bool, Italic: bool, UnderLine: bool) -> str:
    result = ""
    if Bold is True:
        if result == "":
            result = "bold"
        if result != "":
            result = result + "|bold"
    if Italic is True:
        if result == "":
            result = "italic"
        if result != "":
            result = result + "|italic"
    if UnderLine is True:
        if result == "":
            result = "underline"
        if result != "":
            result = result + "|underline"
    return result


# Format the table based on the preferences
def FormatTable(sheet, Endrow):
    # HeaderRange
    RangeAlign1 = "A1:A" + str(Endrow)
    RangeStyle1 = "A1:E1"

    # TableRange
    RangeAlign2 = "B1:E" + str(Endrow)
    RangeStyle2 = "B2:E" + str(Endrow)
    # First column
    RangeStyle3 = "A2:A" + str(Endrow)

    # Font style for the top row
    sheet.setStyle(
        RangeStyle1,
        FontStyle(
            SPREADSHEET_HEADERFONTSTYLE_BOLD,
            SPREADSHEET_HEADERFONTSTYLE_ITALIC,
            SPREADSHEET_HEADERFONTSTYLE_UNDERLINE,
        ),
    )

    # Font style for the first column
    sheet.setStyle(
        RangeStyle3,
        FontStyle(
            SPREADSHEET_COLUMNFONTSTYLE_BOLD,
            SPREADSHEET_COLUMNFONTSTYLE_ITALIC,
            SPREADSHEET_COLUMNFONTSTYLE_UNDERLINE,
        ),
    )

    # Font style for the rest of the table
    sheet.setStyle(
        RangeStyle2,
        FontStyle(
            SPREADSHEET_TABLEFONTSTYLE_BOLD,
            SPREADSHEET_TABLEFONTSTYLE_ITALIC,
            SPREADSHEET_TABLEFONTSTYLE_UNDERLINE,
        ),
    )

    sheet.setBackground(RangeStyle1, SPREADSHEET_HEADERBACKGROUND)
    sheet.setForeground(RangeStyle1, SPREADSHEET_HEADERFOREGROUND)

    # Style the rest of the table
    for i in range(2, int(Endrow) + 1, 2):
        RangeStyle3 = f"A{i}:E{i}"
        RangeStyle4 = f"A{i+1}:E{i+1}"
        sheet.setBackground(RangeStyle3, SPREADSHEET_TABLEBACKGROUND_1)
        sheet.setBackground(RangeStyle4, SPREADSHEET_TABLEBACKGROUND_2)
        sheet.setForeground(RangeStyle3, SPREADSHEET_TABLEFOREGROUND)
        sheet.setForeground(RangeStyle4, SPREADSHEET_TABLEFOREGROUND)

    # align the columns
    sheet.setAlignment(RangeAlign1, "left|vcenter")
    sheet.setAlignment(RangeAlign2, "center|vcenter")

    # Set the column width
    for i in range(1, Endrow):
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"A{i}", cellValue=sheet.getContents(f"A{i}"), factor=AUTOFIT_FACTOR)
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"B{i}", cellValue=sheet.getContents(f"B{i}"), factor=AUTOFIT_FACTOR)
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"C{i}", cellValue=sheet.getContents(f"C{i}"), factor=AUTOFIT_FACTOR)
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"D{i}", cellValue=sheet.getContents(f"D{i}"), factor=AUTOFIT_FACTOR)
        Standard_Functions.SetColumnWidth_SpreadSheet(
            sheet=sheet, column=f"E{i}", cellValue=sheet.getContents(f"E{i}"), factor=AUTOFIT_FACTOR)
    return

# endregion


def ExportSpreadSheet_Excel():
    import Settings
    from Settings import preferences
    from Settings import IMPORT_SETTINGS_XL
    from Settings import EXTERNAL_SOURCE_STARTCELL
    from Settings import ENABLE_DEBUG

    try:
        # get the spreadsheet "TitleBlock"
        sheet = sheet = App.ActiveDocument.getObject("TitleBlock")
        if sheet is None:
            Text = translate(
                "TitleBlock Workbench", "No spreadsheet named 'TitleBlock'!!!"
            )
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
            return
        sheet.recompute()

        # Create a workbook and activate the first sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "TitleBlockData"
        preferences.SetString("SheetName", "TitleBlockData")

        # Get the startcell and the next cells
        StartCell = str(EXTERNAL_SOURCE_STARTCELL)
        TopRow = int(StartCell[1:])
        PropCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
            )
        ) + str(TopRow)
        IncreaseCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
            )
        ) + str(TopRow)
        MultiplierCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
            )
        ) + str(TopRow)
        RemarkCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
            )
        ) + str(TopRow)

        if ENABLE_DEBUG is True:
            Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
            print(Text)

        # Set the headers
        ws[StartCell].value = str(sheet.getContents("A1"))
        ws[PropCell].value = str(sheet.getContents("B1"))
        ws[IncreaseCell].value = str(sheet.getContents("C1"))
        ws[MultiplierCell].value = str(sheet.getContents("D1"))
        ws[RemarkCell].value = str(sheet.getContents("E1"))

        # Go through the spreadsheet.
        for RowNumber in range(1000):
            # Start with x+1 first, to make sure that x is at least 1.
            RowNumber = RowNumber + 1 + TopRow

            try:
                ws[StartCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("A" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[StartCell[:1] + str(RowNumber)].value = ""

            try:
                ws[PropCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("B" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[PropCell[:1] + str(RowNumber)].value = ""

            try:
                ws[IncreaseCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("C" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[IncreaseCell[:1] + str(RowNumber)].value = ""

            try:
                ws[MultiplierCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("D" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[MultiplierCell[:1] + str(RowNumber)].value = ""

            try:
                ws[RemarkCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("E" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[RemarkCell[:1] + str(RowNumber)].value = ""

            # Check if the next row exits. If not this is the end of all the available values.
            try:
                sheet.getContents("A" + str(RowNumber - TopRow + 1))
            except Exception:
                break

        # region Format the settings with the values as a Table
        #
        # Define the the last cell
        EndCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
            )
        ) + str(RowNumber - 1)

        # Define the table
        NumberOfSheets = str(len(wb.sheetnames))
        tab = Table(
            displayName="TitleBlockTable_" + NumberOfSheets,
            ref=f"{StartCell}:{EndCell}",
        )

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )

        # add the style to the table
        tab.tableStyleInfo = style

        # add the table to the worksheet
        ws.add_table(tab)

        # Align the columns
        for row in ws[1: ws.max_row]:
            Column_1 = row[Standard_Functions.GetNumberFromLetter(StartCell[:1]) - 1]
            Column_2 = row[Standard_Functions.GetNumberFromLetter(PropCell[:1]) - 1]
            Column_3 = row[Standard_Functions.GetNumberFromLetter(IncreaseCell[:1]) - 1]
            Column_4 = row[
                Standard_Functions.GetNumberFromLetter(MultiplierCell[:1]) - 1
            ]
            Column_5 = row[Standard_Functions.GetNumberFromLetter(RemarkCell[:1]) - 1]
            Column_1.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_2.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_3.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_4.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_5.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
        # endregion

        # Make the columns to autofit the date
        for col in ws.columns:
            SetLen = 0
            column = col[0].column_letter  # Get the column name

            for cell in col:
                if len(str(cell.value)) > SetLen:
                    SetLen = len(str(cell.value))

            set_col_width = SetLen + 5
            # Setting the column width
            ws.column_dimensions[column].width = set_col_width

        # Save the excel file in a folder of your choosing
        Filter = [
            ("Excel", "*.xlsx"),
        ]
        FileName = Standard_Functions.SaveDialog(Filter)
        if FileName is not None:
            # Save the workbook
            wb.save(str(FileName))
            # Close the workbook
            wb.close()

        # If import settings from excel is enabled, export settings to the new excel file.
        if IMPORT_SETTINGS_XL is True:
            Settings.ExportSettings_XL(Silent=True)

        # print a message if you succeded.
        Text = translate(
            "TitleBlock Workbench",
            f"The titleblock data is exported to the workbook {FileName} in the worksheet {ws.title}",
        )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            raise (e)


def ExportSpreadSheet_FreeCAD():
    # import Settings
    from Settings import preferences
    # from Settings import IMPORT_SETTINGS_XL
    # from Settings import EXTERNAL_SOURCE_STARTCELL
    from Settings import ENABLE_DEBUG

    try:
        # get the spreadsheet "TitleBlock"
        doc = App.ActiveDocument
        sheet = doc.getObject("TitleBlock")
        LastActiveDoc = doc.Name

        if sheet is None:
            Text = translate(
                "TitleBlock Workbench", "No spreadsheet named 'TitleBlock'!!!"
            )
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
            return
        sheet.recompute()

        # Create a new FreeCAD file
        ff = App.newDocument()
        # Save the FreeCAD file in a folder of your choosing
        Filter = [
            ("FreeCAD", "*.FCStd"),
        ]
        FileName = Standard_Functions.SaveDialog(Filter)
        if FileName is not None:
            # Save the workbook
            ff.saveAs(FileName)
            App.closeDocument(ff.Name)
        if FileName is None:
            return

        ff = App.openDocument(FileName, True)
        ff.recompute(None, True, True)
        ff.save

        # Create a spreadsheet in it.
        TitleBlockData = ff.addObject("Spreadsheet::Sheet", "TitleBlockData")
        preferences.SetString("SheetName", "TitleBlockData")

        # Get the startcell and the next cells
        StartCell = "A1"
        TopRow = int(StartCell[1:])
        PropCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
            )
        ) + str(TopRow)
        IncreaseCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
            )
        ) + str(TopRow)
        MultiplierCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
            )
        ) + str(TopRow)
        RemarkCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
            )
        ) + str(TopRow)

        if ENABLE_DEBUG is True:
            Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
            print(Text)

        # Set the headers
        TitleBlockData.set(StartCell, str(sheet.getContents("A1")))
        TitleBlockData.set(PropCell, str(sheet.getContents("B1")))
        TitleBlockData.set(IncreaseCell, str(sheet.getContents("C1")))
        TitleBlockData.set(MultiplierCell, str(sheet.getContents("D1")))
        TitleBlockData.set(RemarkCell, str(sheet.getContents("E1")))

        # Go through the spreadsheet.
        for RowNumber in range(1000):
            # Start with x+1 first, to make sure that x is at least 1.
            RowNumber = RowNumber + 1 + TopRow

            try:
                TitleBlockData.set(
                    StartCell[: 1] + str(RowNumber),
                    str(sheet.getContents("A" + str(RowNumber - TopRow + 1)))
                )
            except Exception:
                TitleBlockData.set(StartCell[: 1] + str(RowNumber), "")

            try:
                TitleBlockData.set(
                    PropCell[: 1] + str(RowNumber),
                    str(sheet.getContents("B" + str(RowNumber - TopRow + 1)))
                )
            except Exception:
                TitleBlockData.set(PropCell[: 1] + str(RowNumber), "")

            try:
                TitleBlockData.set(
                    IncreaseCell[: 1] + str(RowNumber),
                    str(sheet.getContents("C" + str(RowNumber - TopRow + 1)))
                )
            except Exception:
                TitleBlockData.set(PropCell[: 1] + str(RowNumber), "")

            try:
                TitleBlockData.set(
                    MultiplierCell[: 1] + str(RowNumber),
                    str(sheet.getContents("D" + str(RowNumber - TopRow + 1)))
                )
            except Exception:
                TitleBlockData.set(PropCell[: 1] + str(RowNumber), "")

            try:
                TitleBlockData.set(
                    RemarkCell[: 1] + str(RowNumber),
                    str(sheet.getContents("E" + str(RowNumber - TopRow + 1)))
                )
            except Exception:
                TitleBlockData.set(PropCell[: 1] + str(RowNumber), "")

            # Check if the next row exits. If not this is the end of all the available values.
            try:
                test = sheet.getContents("A" + str(RowNumber - TopRow + 1))
                if test == "" or test is None:
                    break
            except Exception:
                break

        # region Format the settings with the values as a Table
        #
        FormatTable(sheet=TitleBlockData, Endrow=RowNumber - 2)
        # endregion

        # recompute the document
        ff.recompute(None, True, True)
        # Save the workbook
        ff.save()
        # Close the FreeCAD file
        App.closeDocument(ff.Name)
        # Activate the document which was active when this command started.
        App.setActiveDocument(LastActiveDoc)

        # If import settings from external source is enabled, export settings to the new excel file.
        if IMPORT_SETTINGS_XL is True:
            Settings.ExportSettings_FreeCAD(Silent=True)

        # print a message if you succeded.
        Text = translate(
            "TitleBlock Workbench",
            f"The titleblock data is exported to the FreeCAD file {FileName} in the spreadsheet TitleBlockData",
        )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            raise (e)
