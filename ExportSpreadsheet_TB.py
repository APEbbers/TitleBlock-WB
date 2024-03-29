# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/


import FreeCAD as App
import Standard_Functions_TB as Standard_Functions
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import TableFormat_Functions_TB
import os

# Get the settings
import Settings_TB
from Settings_TB import preferences
from Settings_TB import EXTERNAL_SOURCE_STARTCELL

# from Settings import EXTERNAL_SOURCE_SHEET_NAME
# from Settings import EXTERNAL_SOURCE_PATH
from Settings_TB import IMPORT_SETTINGS_XL
from Settings_TB import ENABLE_DEBUG

# Define the translation
translate = App.Qt.translate


def ExportSpreadSheet_Excel() -> bool:
    result = False
    try:
        # get the spreadsheet "TitleBlock"
        sheet = App.ActiveDocument.getObject("TitleBlock")
        if sheet is None:
            Text = translate(
                "TitleBlock Workbench", "No spreadsheet named 'TitleBlock'!!!"
            )
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
            return
        sheet.recompute()

        # Create a workbook and activate the first sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "TitleBlockData"
        preferences.SetString("SheetName", "TitleBlockData")

        # Get the startcell and the next cells
        StartCell = str(EXTERNAL_SOURCE_STARTCELL)
        TopRow = int(StartCell[1:])
        PropCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
            )
        ) + str(TopRow)
        IncreaseCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
            )
        ) + str(TopRow)
        MultiplierCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
            )
        ) + str(TopRow)
        RemarkCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
            )
        ) + str(TopRow)

        if ENABLE_DEBUG is True:
            Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
            Standard_Functions.Print(Input=Text, Type="Log")

        # Set the headers
        ws[StartCell].value = str(sheet.getContents("A1"))
        ws[PropCell].value = str(sheet.getContents("B1"))
        ws[IncreaseCell].value = str(sheet.getContents("C1"))
        ws[MultiplierCell].value = str(sheet.getContents("D1"))
        ws[RemarkCell].value = str(sheet.getContents("E1"))

        # Go through the spreadsheet.
        for RowNumber in range(1000):
            # Start with x+1 first, to make sure that x is at least 1.
            RowNumber = RowNumber + 1 + TopRow

            try:
                ws[StartCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("A" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[StartCell[:1] + str(RowNumber)].value = ""

            try:
                ws[PropCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("B" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[PropCell[:1] + str(RowNumber)].value = ""

            try:
                ws[IncreaseCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("C" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[IncreaseCell[:1] + str(RowNumber)].value = ""

            try:
                ws[MultiplierCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("D" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[MultiplierCell[:1] + str(RowNumber)].value = ""

            try:
                ws[RemarkCell[:1] + str(RowNumber)].value = str(
                    sheet.getContents("E" + str(RowNumber - TopRow + 1))
                )
            except Exception:
                ws[RemarkCell[:1] + str(RowNumber)].value = ""

            # Check if the next row exits. If not this is the end of all the available values.
            try:
                sheet.getContents("A" + str(RowNumber - TopRow + 1))
            except Exception:
                break

        # region Format the settings with the values as a Table
        #
        # Define the the last cell
        EndCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
            )
        ) + str(RowNumber - 1)

        # Define the table
        NumberOfSheets = str(len(wb.sheetnames))
        tab = Table(
            displayName="TitleBlockTable_" + NumberOfSheets,
            ref=f"{StartCell}:{EndCell}",
        )

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )

        # add the style to the table
        tab.tableStyleInfo = style

        # add the table to the worksheet
        ws.add_table(tab)

        # Align the columns
        for row in ws[1 : ws.max_row]:
            Column_1 = row[Standard_Functions.GetNumberFromLetter(StartCell[:1]) - 1]
            Column_2 = row[Standard_Functions.GetNumberFromLetter(PropCell[:1]) - 1]
            Column_3 = row[Standard_Functions.GetNumberFromLetter(IncreaseCell[:1]) - 1]
            Column_4 = row[
                Standard_Functions.GetNumberFromLetter(MultiplierCell[:1]) - 1
            ]
            Column_5 = row[Standard_Functions.GetNumberFromLetter(RemarkCell[:1]) - 1]
            Column_1.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_2.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_3.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_4.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_5.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
        # endregion

        # Make the columns to autofit the date
        for col in ws.columns:
            SetLen = 0
            column = col[0].column_letter  # Get the column name

            for cell in col:
                if len(str(cell.value)) > SetLen:
                    SetLen = len(str(cell.value))

            set_col_width = SetLen + 5
            # Setting the column width
            ws.column_dimensions[column].width = set_col_width

        # Save the excel file in a folder of your choosing
        Filter = [
            ("Excel", "*.xlsx"),
        ]
        FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=True)
        if FileName != "":
            # Save the workbook
            wb.save(str(FileName))
            # Close the workbook
            wb.close()
        if FileName == "":
            return

        # If import settings from excel is enabled, export settings to the new excel file.
        if IMPORT_SETTINGS_XL is True:
            Settings_TB.ExportSettings_XL(Silent=True)

        # print a message if you succeded.
        Text = translate(
            "TitleBlock Workbench",
            f"The titleblock data is exported to the workbook {FileName} in the worksheet {ws.title}",
        )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            raise (e)
    finally:
        # Close the excel workbook
        wb.close()
        result = True
    return result


def ExportSpreadSheet_FreeCAD():
    try:
        # Get the active document
        doc = App.ActiveDocument
        # get the spreadsheet "TitleBlock"
        sheet = doc.getObject("TitleBlock")
        # Save the name of the active document to reactivate it at the end of this function.
        LastActiveDoc = doc.Name

        # If there is no spreadsheet named TitleBlock, show a message and exit this function.
        if sheet is None:
            Text = translate(
                "TitleBlock Workbench", "No spreadsheet named 'TitleBlock'!!!"
            )
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
            return

        # Create a placeholder for the new document
        ff = ""
        # Save the FreeCAD file in a folder of your choosing
        Filter = [
            ("FreeCAD", "*.FCStd"),
        ]
        FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=True)
        if FileName != "":
            FileNameOnly = os.path.basename(FileName)
            if Standard_Functions.CheckIfDocumentIsOpen(FileNameOnly) is False:
                # Create a new FreeCAD file
                ff = App.newDocument()
                # Save the workbook
                ff.saveAs(FileName)
                # Close the document before reopening
                App.closeDocument(ff.Name)

            if Standard_Functions.CheckIfDocumentIsOpen(FileNameOnly) is True:
                App.closeDocument(FileNameOnly)
        if FileName == "":
            return

        # Open the document hidden, recompute and save it
        ff = App.openDocument(FileName, True)
        ff.recompute(None, True, True)
        ff.save

        # Create a spreadsheet for the titleblock data.
        TitleBlockData = ff.addObject("Spreadsheet::Sheet", "TitleBlockData")
        preferences.SetString("SheetName", "TitleBlockData")

        # Get the startcell and the next cells
        StartCell = EXTERNAL_SOURCE_STARTCELL
        if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
            StartCell = Standard_Functions.GetA1fromR1C1(StartCell)
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    f"TitleBlock Workbench: the startcell converted from {EXTERNAL_SOURCE_STARTCELL} to {StartCell}",
                )
                Standard_Functions.Print(Text, "Log")

        TopRow = int(Standard_Functions.RemoveLettersFromString(StartCell))
        PropCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
            )
        ) + str(TopRow)
        IncreaseCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 2
            )
        ) + str(TopRow)
        MultiplierCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 3
            )
        ) + str(TopRow)
        RemarkCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 4
            )
        ) + str(TopRow)

        if ENABLE_DEBUG is True:
            Text = translate("TitleBlock Workbench", f"the startcell is: {StartCell}")
            Standard_Functions.Print(Input=Text, Type="Log")

        # Set the headers
        TitleBlockData.set(StartCell, str(sheet.getContents("A1")))
        TitleBlockData.set(PropCell, str(sheet.getContents("B1")))
        TitleBlockData.set(IncreaseCell, str(sheet.getContents("C1")))
        TitleBlockData.set(MultiplierCell, str(sheet.getContents("D1")))
        TitleBlockData.set(RemarkCell, str(sheet.getContents("E1")))

        # Go through the external spreadsheet and fill in the data.
        for RowNumber in range(1000):
            # Start with x+1 first, to make sure that x is at least 1.
            RowNumber = RowNumber + 1 + TopRow

            try:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(StartCell)
                    + str(RowNumber),
                    str(sheet.getContents("A" + str(RowNumber - TopRow + 1))),
                )
            except Exception:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(StartCell)
                    + str(RowNumber),
                    "",
                )

            try:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(PropCell)
                    + str(RowNumber),
                    str(sheet.getContents("B" + str(RowNumber - TopRow + 1))),
                )
            except Exception:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(PropCell), ""
                )

            try:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(IncreaseCell)
                    + str(RowNumber),
                    str(sheet.getContents("C" + str(RowNumber - TopRow + 1))),
                )
            except Exception:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(IncreaseCell)
                    + str(RowNumber),
                    "",
                )

            try:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(MultiplierCell)
                    + str(RowNumber),
                    str(sheet.getContents("D" + str(RowNumber - TopRow + 1))),
                )
            except Exception:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(MultiplierCell)
                    + str(RowNumber),
                    "",
                )

            try:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(RemarkCell)
                    + str(RowNumber),
                    str(sheet.getContents("E" + str(RowNumber - TopRow + 1))),
                )
            except Exception:
                TitleBlockData.set(
                    Standard_Functions.RemoveNumbersFromString(RemarkCell)
                    + str(RowNumber),
                    "",
                )

            # Check if the next row of the spreadsheet has data. If not this is the end of all the available values.
            try:
                test = sheet.getContents("A" + str(RowNumber - TopRow + 1))
                if test == "" or test is None:
                    break
            except Exception:
                break

        # region Format the data with the values as a Table
        #
        # Define the header range
        HeaderRange = str(f"{StartCell}:{RemarkCell}")

        # Get the first row below the header
        FirstTableRow = ""
        for i in range(len(StartCell)):
            if StartCell[i].isdigit():
                FirstTableRow = FirstTableRow + str(StartCell[i])
        FirstTableRow = int(FirstTableRow) + 1

        # Get the first column
        FirstColumn = Standard_Functions.RemoveNumbersFromString(StartCell)

        # Get the last column
        LastColumn = Standard_Functions.RemoveNumbersFromString(RemarkCell)

        # Define the table range
        TableRange = str(f"{FirstColumn}{FirstTableRow}:{LastColumn}{RowNumber - 1}")

        # Define the First column range
        FirstColumnRange = str(
            f"{FirstColumn}{FirstTableRow}:{FirstColumn}{RowNumber - 1}"
        )

        # Format the table
        TableFormat_Functions_TB.FormatTable(
            sheet=TitleBlockData,
            HeaderRange=HeaderRange,
            TableRange=TableRange,
            FirstColumnRange=FirstColumnRange,
        )

        # endregion

        # Save the name of the external file before Export settings closes
        ExternalFileName = ff.Name

        # If import settings from external source is enabled, export settings to the new excel file.
        if IMPORT_SETTINGS_XL is True:
            Settings_TB.ExportSettings_FreeCAD(Silent=True)

        # recompute the document
        if Standard_Functions.CheckIfDocumentIsOpen(ExternalFileName):
            ff.recompute(None, True, True)
            # Save the workbook
            ff.save()
            # Close the FreeCAD file
            App.closeDocument(ff.Name)
        # Activate the document which was active when this command started.
        try:
            doc = App.setActiveDocument(LastActiveDoc)
            doc.recompute(None, True, True)
        except Exception:
            pass

        # print a message if you succeded.
        Text = translate(
            "TitleBlock Workbench",
            f"The titleblock data is exported to the FreeCAD file {FileName} in the spreadsheet TitleBlockData",
        )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            raise (e)
