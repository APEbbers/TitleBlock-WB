# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/


# This macro creates an Spreadsheet named "TitleBlock", reads editable texts from the first page
# and fills the spreadsheet.
# The data in spreadsheet can be changed by the user as long the data in column A remains unchanged.
# There will be three colums created in the spreadsheet:
#  - Property Name           -> the name of the editable text item
#  - Property Value          -> the value of the editable text item
#  - Increase, Yes or No?    -> do you want to increase the  value by 1 for the next sheet.
# there a multiple senarios for using these macro's:
#  - When changing the template for an different size, you can quickly refill the titleblock.
#    Of course, the editable text in the new template must be equal.
#  - Link the date to model properties like parameters. Basicly all what is achievable with the Spreadsheet workbench
#  - Import data from excel or libre office
#  - Etc.

# region imports

import os
import FreeCAD as App
import Standard_Functions_TitleBlock as Standard_Functions
import TableFormat_Functions

# Get the settings
from Settings import DRAW_NO_FIELD
from Settings import USE_FILENAME_DRAW_NO
from Settings import DRAW_NO_FIELD_PAGE
from Settings import USE_PAGENAME_DRAW_NO
from Settings import ENABLE_DEBUG
from Settings import DOCINFO_COMMENT
from Settings import DOCINFO_LICENSEURL
from Settings import DOCINFO_LICENSE
from Settings import DOCINFO_COMPANY
from Settings import DOCINFO_LASTMODIFIEDDATE
from Settings import DOCINFO_LASTMODIFIEDBY
from Settings import DOCINFO_CREATEDDATE
from Settings import DOCINFO_CREATEDBY
from Settings import DOCINFO_NAME
from Settings import MAP_NOSHEETS
from Settings import MAP_MASS
from Settings import MAP_ANGLE
from Settings import MAP_LENGTH
from Settings import EXTERNAL_SOURCE_STARTCELL
from Settings import EXTERNAL_SOURCE_SHEET_NAME
from Settings import EXTERNAL_SOURCE_PATH
from Settings import USE_EXTERNAL_SOURCE
from Settings import INCLUDE_NO_SHEETS
from Settings import INCLUDE_MASS
from Settings import INCLUDE_ANGLE
from Settings import INCLUDE_LENGTH
from Settings import preferences
# endregion

# Define the translation
translate = App.Qt.translate

# region - Setting corrections

# If no start cell is defined. the start cell will be "A1"
if len(EXTERNAL_SOURCE_STARTCELL) == 0:
    EXTERNAL_SOURCE_STARTCELL = "A1"

# When you copy/paste from a cell in the spreadsheet workbench, the char ' is automaticly added
# at the begining of the text.
# To avoid annociance, if this char is detected, it is removed from the setting.
if str(MAP_LENGTH).startswith("'"):
    MAP_LENGTH = str(MAP_LENGTH)[1:]
if str(MAP_ANGLE).startswith("'"):
    MAP_ANGLE = str(MAP_ANGLE)[1:]
if str(MAP_MASS).startswith("'"):
    MAP_MASS = str(MAP_MASS)[1:]
if str(MAP_NOSHEETS).startswith("'"):
    MAP_NOSHEETS = str(MAP_NOSHEETS)[1:]
if str(DOCINFO_NAME).startswith("'"):
    DOCINFO_NAME = str(DOCINFO_NAME)[1:]
if str(DOCINFO_CREATEDBY).startswith("'"):
    DOCINFO_CREATEDBY = str(DOCINFO_CREATEDBY)[1:]
if str(DOCINFO_CREATEDDATE).startswith("'"):
    DOCINFO_CREATEDDATE = str(DOCINFO_CREATEDDATE)[1:]
if str(DOCINFO_LASTMODIFIEDBY).startswith("'"):
    DOCINFO_LASTMODIFIEDBY = str(DOCINFO_LASTMODIFIEDBY)[1:]
if str(DOCINFO_LASTMODIFIEDDATE).startswith("'"):
    DOCINFO_LASTMODIFIEDDATE = str(DOCINFO_LASTMODIFIEDDATE)[1:]
if str(DOCINFO_COMPANY).startswith("'"):
    DOCINFO_COMPANY = str(DOCINFO_COMPANY)[1:]
if str(DOCINFO_LICENSE).startswith("'"):
    DOCINFO_LICENSE = str(DOCINFO_LICENSE)[1:]
if str(DOCINFO_LICENSEURL).startswith("'"):
    DOCINFO_LICENSEURL = str(DOCINFO_LICENSEURL)[1:]
if str(DOCINFO_COMMENT).startswith("'"):
    DOCINFO_COMMENT = str(DOCINFO_COMMENT)[1:]
if str(DRAW_NO_FIELD).startswith("'"):
    DRAW_NO_FIELD = str(DRAW_NO_FIELD)[1:]
# endregion

# region - suporting functions

# The following system values can be added: length unit, angle unit, mass unti and number of sheets.
# You can use these for you specific templates.
# Use the "bind function" of the spreadsheet workbench to create the correct entry for your template.:
#   1. add the correct Property name
#   2. bind the cell to the value of your specific template property.
#   3. Mark the cell in column C if this value needs to be increased per page


def AddExtraData(sheet, StartRow, doc=None):
    if doc is None:
        doc = App.ActiveDocument

    # If the debug mode is active, show which property is includex.difference(y)
    if ENABLE_DEBUG is True:
        Text = ""
        if INCLUDE_LENGTH is True:
            Text = translate(
                "TitleBlock Workbench",
                "Length unit is included: " + str(INCLUDE_LENGTH),
            )
        if INCLUDE_ANGLE is True:
            Text = translate(
                "TitleBlock Workbench", "Angle unit is included: " + str(INCLUDE_ANGLE)
            )
        if INCLUDE_MASS is True:
            Text = translate(
                "TitleBlock Workbench", "Mass unit is included: " + str(INCLUDE_MASS)
            )
        if INCLUDE_NO_SHEETS is True:
            Text = translate(
                "TitleBlock Workbench",
                "Number of pages is included: " + str(INCLUDE_NO_SHEETS),
            )
        Standard_Functions.Print(Text, "Log")
    # get units scheme
    SchemeNumber = App.Units.getSchema()

    # Add the length units of your FreeCAD application
    if INCLUDE_LENGTH is True:
        sheet.set("A" + str(StartRow + 1), "Length_Units")
        sheet.set(
            "B" + str(StartRow + 1),
            str(
                App.Units.schemaTranslate(
                    App.Units.Quantity(50, App.Units.Length), SchemeNumber
                )
            )
            .split()[1]
            .replace("'", "")
            .replace(",", ""),
        )
        StartRow = StartRow + 1

    # Add the angular units of your FreeCAD application
    if INCLUDE_ANGLE is True:
        sheet.set("A" + str(StartRow + 1), "Angle_Units")
        sheet.set(
            "B" + str(StartRow + 1),
            str(
                App.Units.schemaTranslate(
                    App.Units.Quantity(50, App.Units.Angle), SchemeNumber
                )
            )
            .split()[1]
            .replace("'", "")
            .replace(",", ""),
        )
        StartRow = StartRow + 1

    # Add the mass units of your FreeCAD application
    if INCLUDE_MASS is True:
        sheet.set("A" + str(StartRow + 1), "Mass_Units")
        sheet.set(
            "B" + str(StartRow + 1),
            str(
                App.Units.schemaTranslate(
                    App.Units.Quantity(50, App.Units.Mass), SchemeNumber
                )
            )
            .split()[1]
            .replace("'", "")
            .replace(",", ""),
        )
        StartRow = StartRow + 1

    # Add the total number of sheets. You can use this for your title block
    if INCLUDE_NO_SHEETS is True:
        # Get all the pages
        pages = App.ActiveDocument.findObjects("TechDraw::DrawPage")
        # Create the Property and add the value
        sheet.set("A" + str(StartRow + 1), "Number of sheets")
        sheet.set("B" + str(StartRow + 1), str(len(pages)))
    return


# Map data from the system and/or document to the spreadsheet
def MapData(sheet, doc=None):
    if doc is None:
        doc = App.ActiveDocument

    # Get the filename
    filename = os.path.basename(doc.FileName).split(".")[0]
    pages = doc.findObjects("TechDraw::DrawPage")
    # get the fist page.
    pagename = ""
    pages = App.ActiveDocument.findObjects("TechDraw::DrawPage")
    if len(pages) > 0:
        pagename = pages[0].Label

    # if the debug mode is on, show what is mapped to which property
    if ENABLE_DEBUG is True:
        Text = ""
        if str(MAP_LENGTH).strip():
            Text = translate(
                "TitleBlock Workbench", "Length unit is mapped to: " + str(MAP_LENGTH)
            )
        if str(MAP_ANGLE).strip():
            Text = translate(
                "TitleBlock Workbench", "Angle unit is mapped to: " + str(MAP_ANGLE)
            )
        if str(MAP_MASS).strip():
            Text = translate(
                "TitleBlock Workbench", "Mass unit is mapped to: " + str(MAP_MASS)
            )
        if str(MAP_NOSHEETS).strip():
            Text = translate(
                "TitleBlock Workbench",
                "the number of pages is mapped to: " + str(MAP_NOSHEETS),
            )
        if USE_FILENAME_DRAW_NO is True:
            Text = translate(
                "TitleBlock Workbench",
                "The filename ("
                + str(filename)
                + ") is mapped to: "
                + str(DRAW_NO_FIELD),
            )
        if USE_PAGENAME_DRAW_NO is True:
            Text = translate(
                "TitleBlock Workbench",
                "The pagenames will be mapped to: "
                + str(DRAW_NO_FIELD_PAGE),
            )
        Standard_Functions.Print(Text, "Log")

    # get units scheme
    SchemeNumber = App.Units.getSchema()

    # Go through the column A in the spreadsheet and find the properties.
    for RowNum in range(1000):
        # Start with x+1 first, to make sure that x is at least 1.
        RowNum = RowNum + 2

        # Get the property name. If it starts with ' remove it.
        PropertyName = str(sheet.getContents("A" + str(RowNum)))
        if PropertyName.startswith("'"):
            PropertyName = PropertyName[1:]

        # Map length units of your FreeCAD application
        # Map only as requested
        if str(MAP_LENGTH).strip():
            # If the cell in column A is equal to MAP_LENGTH, add the value in column B
            if PropertyName == MAP_LENGTH:
                sheet.set(
                    "B" + str(RowNum),
                    str(
                        App.Units.schemaTranslate(
                            App.Units.Quantity(50, App.Units.Length), SchemeNumber
                        )
                    )
                    .split()[1]
                    .replace("'", "")
                    .replace(",", ""),
                )

        # Map angle units of your FreeCAD application
        # Map only as requested
        if str(MAP_ANGLE).strip():
            # If the cell in column A is equal to MAP_ANGLE, add the value in column B
            if PropertyName == MAP_ANGLE:
                sheet.set(
                    "B" + str(RowNum),
                    str(
                        App.Units.schemaTranslate(
                            App.Units.Quantity(50, App.Units.Angle), SchemeNumber
                        )
                    )
                    .split()[1]
                    .replace("'", "")
                    .replace(",", ""),
                )

        # Map mass units of your FreeCAD application
        # Map only as requested
        if str(MAP_MASS).strip():
            # If the cell in column A is equal to MAP_MASS, add the value in column B
            if PropertyName == MAP_MASS:
                sheet.set(
                    "B" + str(RowNum),
                    str(
                        App.Units.schemaTranslate(
                            App.Units.Quantity(50, App.Units.Mass), SchemeNumber
                        )
                    )
                    .split()[1]
                    .replace("'", "")
                    .replace(",", ""),
                )

        # Map the number of pages
        # Map only as requested
        if str(MAP_NOSHEETS).strip():
            # Get all the pages
            pages = App.ActiveDocument.findObjects("TechDraw::DrawPage")
            # If the cell in column A is equal to MAP_NOSHEETS, add the value in column B
            if PropertyName == MAP_NOSHEETS:
                sheet.set("B" + str(RowNum), str(len(pages)))

        # Map the filename
        # Map only as requested
        if USE_FILENAME_DRAW_NO is True:
            if str(DRAW_NO_FIELD).strip():
                # If the cell in column A is equal to DRAW_NO_FIELD, add the value in column B
                if PropertyName == DRAW_NO_FIELD:
                    sheet.set("B" + str(RowNum), filename)

        # Map the pagename
        # Map only as requested
        if USE_PAGENAME_DRAW_NO is True:
            if str(DRAW_NO_FIELD_PAGE).strip() and pagename != "":
                # If the cell in column A is equal to DRAW_NO_FIELD, add the value in column B
                if PropertyName == DRAW_NO_FIELD_PAGE:
                    sheet.set("B" + str(RowNum), pagename)
                    sheet.set("E" + str(RowNum), "The name of each page will be mapped to its titleblock")

        # Check if the next row exits. If not this is the end of all the available values.
        try:
            sheet.getContents("A" + str(RowNum + 1))
        except Exception:
            return


# Map document information
def MapDocInfo(sheet, doc=None):
    if doc is None:
        doc = App.ActiveDocument

    # if the debug mode is on, show what is mapped to which property
    if ENABLE_DEBUG is True:
        Text = ""
        if str(DOCINFO_NAME).strip():
            Text = translate(
                "TitleBlock Workbench",
                "Document name is mapped to:  " + str(DOCINFO_NAME),
            )
        if str(DOCINFO_CREATEDBY).strip():
            Text = translate(
                "TitleBlock Workbench",
                "Created by value is mapped to:  " + str(DOCINFO_CREATEDBY),
            )
        if str(DOCINFO_CREATEDDATE).strip():
            Text = translate(
                "TitleBlock Workbench",
                "Created date is mapped to:  " + str(DOCINFO_CREATEDDATE),
            )
        if str(DOCINFO_LASTMODIFIEDBY).strip():
            Text = translate(
                "TitleBlock Workbench",
                "Last modified by value is mapped to:  " + str(DOCINFO_LASTMODIFIEDBY),
            )
        if str(DOCINFO_LASTMODIFIEDDATE).strip():
            Text = translate(
                "TitleBlock Workbench",
                "Last modified date is mapped to:  " + str(DOCINFO_LASTMODIFIEDDATE),
            )
        if str(DOCINFO_COMPANY).strip():
            Text = translate(
                "TitleBlock Workbench",
                "Company name is mapped to:  " + str(DOCINFO_COMPANY),
            )
        if str(DOCINFO_LICENSE).strip():
            Text = translate(
                "TitleBlock Workbench",
                "License name is mapped to:  " + str(DOCINFO_LICENSE),
            )
        if str(DOCINFO_LICENSEURL).strip():
            Text = translate(
                "TitleBlock Workbench",
                "License link is mapped to:  " + str(DOCINFO_LICENSEURL),
            )
        if str(DOCINFO_COMMENT).strip():
            Text = translate(
                "TitleBlock Workbench", "Comment is mapped to:  " + str(DOCINFO_COMMENT)
            )
        Standard_Functions.Print(Text, "Log")

    # Go through the column A in the spreadsheet and find the properties.
    for RowNum in range(1000):
        # Start with x+1 first, to make sure that x is at least 1.
        RowNum = RowNum + 2

        # Get the property name. If it starts with ' remove it.
        PropertyName = str(sheet.getContents("A" + str(RowNum)))
        if PropertyName.startswith("'"):
            PropertyName = PropertyName[1:]

        # DOCINFO_NAME
        if str(DOCINFO_NAME).strip():
            # If the cell in column A is equal to # DOCINFO_NAME, add the value in column B
            if PropertyName == DOCINFO_NAME:
                sheet.set("B" + str(RowNum), doc.Name)

        # DOCINFO_CREATEDBY
        if str(DOCINFO_CREATEDBY).strip():
            # If the cell in column A is equal to # DOCINFO_CREATEDBY, add the value in column B
            if PropertyName == DOCINFO_CREATEDBY:
                sheet.set("B" + str(RowNum), doc.CreatedBy)

        # DOCINFO_CREATEDDATE
        if str(DOCINFO_CREATEDDATE).strip():
            # If the cell in column A is equal to # DOCINFO_CREATEDDATE, add the value in column B
            if PropertyName == DOCINFO_CREATEDDATE:
                CreationDate = doc.CreationDate.split("T")[0]
                sheet.set("B" + str(RowNum), CreationDate)

        # DOCINFO_LASTMODIFIEDBY
        if str(DOCINFO_LASTMODIFIEDBY).strip():
            # If the cell in column A is equal to # DOCINFO_LASTMODIFIEDBY, add the value in column B
            if PropertyName == DOCINFO_LASTMODIFIEDBY:
                sheet.set("B" + str(RowNum), doc.LastModifiedBy)

        # DOCINFO_LASTMODIFIEDDATE
        if str(DOCINFO_LASTMODIFIEDDATE).strip():
            # If the cell in column A is equal to # DOCINFO_LASTMODIFIEDDATE, add the value in column B
            if PropertyName == DOCINFO_LASTMODIFIEDDATE:
                ModificationDate = doc.LastModifiedDate.split("T")[0]
                sheet.set("B" + str(RowNum), ModificationDate)

        # DOCINFO_COMPANY
        if str(DOCINFO_COMPANY).strip():
            # If the cell in column A is equal to # DOCINFO_COMPANY, add the value in column B
            if PropertyName == DOCINFO_COMPANY:
                sheet.set("B" + str(RowNum), doc.Company)

        # DOCINFO_LICENSE
        if str(DOCINFO_LICENSE).strip():
            # If the cell in column A is equal to # DOCINFO_LICENSE, add the value in column B
            if PropertyName == DOCINFO_LICENSE:
                sheet.set("B" + str(RowNum), doc.License)

        # DOCINFO_LICENSEURL
        if str(DOCINFO_LICENSEURL).strip():
            # If the cell in column A is equal to # DOCINFO_LICENSEURL, add the value in column B
            if PropertyName == DOCINFO_LICENSEURL:
                sheet.set("B" + str(RowNum), doc.LicenseURL)

        # DOCINFO_COMMENT
        if str(DOCINFO_COMMENT).strip():
            # If the cell in column A is equal to # DOCINFO_COMMENT, add the value in column B
            if PropertyName == DOCINFO_COMMENT:
                sheet.set("B" + str(RowNum), doc.Comment)
    return

# endregion


# Fill the spreadsheet with all the date from the titleblock
def FillSheet():
    try:
        # get the fist page. If there is no page, return
        pages = App.ActiveDocument.findObjects("TechDraw::DrawPage")
        if len(pages) > 0:
            page = pages[0]
        if len(pages) == 0:
            return

        # get the editable texts
        texts = page.Template.EditableTexts
        # get the spreadsheet "TitleBlock"
        sheet = App.ActiveDocument.getObject("TitleBlock")
        # Clear the sheet
        sheet.clearAll()

        # Debug mode is active, show all editable text in the page
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "the following editable text are present in your page:",
            )
            Standard_Functions.Print(Text, "Log")
            for EditableText in texts.items():
                Text = translate("TitleBlock Workbench", str(EditableText))
                Standard_Functions.Print(Text, "Log")

        # set the headers in the spreadsheet
        sheet.set("A1", "Property Name")
        sheet.set("B1", "Property Value")
        sheet.set("C1", "Increase value")
        sheet.set("D1", "Factor")
        sheet.set("E1", "Remarks")

        # set the start value for the start row.
        # (x=0, the spreadsheet whill be populated from the first row. the headers will be overwritten)
        StartRow = 1
        for key, value in texts.items():
            # Increase StartRow by one, to fill the next row
            StartRow = StartRow + 1
            # Fill the property name
            sheet.set("A" + str(StartRow), "{0}".format(key, value))
            # Fill the property value
            sheet.set("B" + str(StartRow), "{1}".format(key, value))
            # If there is no value yet, the increase function will be set empty by default.
            try:
                str(sheet.getContents("C" + str(StartRow)))
            except Exception:
                sheet.set("C" + str(StartRow), "")

        # Finally recompute the document
        App.ActiveDocument.recompute(None, True, True)

        # Run the def to map system data
        MapData(sheet=sheet)

        # Run the def to map document information
        MapDocInfo(sheet=sheet)

        # Run the def to add extra system data
        AddExtraData(sheet, StartRow)

        extraRows = 0
        if INCLUDE_LENGTH is True:
            extraRows = extraRows + 1
        if INCLUDE_ANGLE is True:
            extraRows = extraRows + 1
        if INCLUDE_MASS is True:
            extraRows = extraRows + 1
        if INCLUDE_NO_SHEETS is True:
            extraRows = extraRows + 1

        # region Format the data with the values as a Table
        #
        # Define the header range
        StartCell = "A1"
        RemarkCell = "E1"
        HeaderRange = str(f"{StartCell}:{RemarkCell}")

        # Get the first row below the header
        FirstTableRow = ""
        for i in range(len(StartCell)):
            if StartCell[i].isdigit():
                FirstTableRow = FirstTableRow + str(StartCell[i])
        FirstTableRow = int(FirstTableRow) + 1

        # Get the first column
        FirstColumn = Standard_Functions.RemoveNumbersFromString(StartCell)

        # Get the last column
        LastColumn = Standard_Functions.RemoveNumbersFromString(RemarkCell)

        # Define the table range
        TableRange = str(f"{FirstColumn}{FirstTableRow}:{LastColumn}{int(StartRow) + extraRows}")

        # Define the First column range
        FirstColumnRange = str(f"{FirstColumn}{FirstTableRow}:{FirstColumn}{int(StartRow) + extraRows}")

        # Format the table
        sheet = TableFormat_Functions.FormatTable(sheet=sheet, HeaderRange=HeaderRange,
                                                  TableRange=TableRange, FirstColumnRange=FirstColumnRange)

        # endregion

        # Finally recompute the document
        App.ActiveDocument.recompute(None, True, True)
    except Exception as e:
        Text = "TitleBlock Workbench: an error occurred!!\n"
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
            raise e
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0, IconType="Critical")
    return


# Import data from a (central) excel workbook
def ImportDataExcel():
    from openpyxl import load_workbook

    # if debug mode is enabled, show the external file including path.
    if ENABLE_DEBUG is True:
        Text = translate("TitleBlock Workbench", str(EXTERNAL_SOURCE_PATH))
        Standard_Functions.Print(Text, "Log")

    # Check if it is allowed to use an external source and if so, continue
    if USE_EXTERNAL_SOURCE is True:
        # try to open the source. if not show an messagebox and if debug mode is enabled, show the exeption as well
        try:
            wb = ""
            if EXTERNAL_SOURCE_PATH.lower().endswith(".fcstd"):
                Filter = [("Excel", "*.xlsx"), ]
                FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=False)
                if FileName != "":
                    wb = load_workbook(FileName, read_only=True, data_only=True)
                if FileName == "":
                    return
            else:
                wb = load_workbook(str(EXTERNAL_SOURCE_PATH), read_only=True, data_only=True)
            if EXTERNAL_SOURCE_SHEET_NAME == "":
                # Set the sheetname with a inputbox
                Worksheets_List = [i for i in wb.sheetnames if i != "Settings"]
                Text = translate(
                    "TitleBlock Workbench", "Please enter the name of the worksheet"
                )
                Input_SheetName = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=3,
                        default="TitleBlockData",
                        stringList=Worksheets_List,
                    )
                )
                # if the user canceled, exit this function.
                if not Input_SheetName.strip():
                    return
                ws = wb[str(Input_SheetName)]
            if EXTERNAL_SOURCE_SHEET_NAME != "":
                ws = wb[str(EXTERNAL_SOURCE_SHEET_NAME)]
        except IOError:
            Standard_Functions.Mbox("Permission error!!\nDo you have the file open?",
                                    "Titleblock Workbench", 0, IconType="Critical")
            return
        except Exception as e:
            if ENABLE_DEBUG is True:
                raise (e)
            Text = translate(
                "TitleBlock Workbench",
                "an problem occured while openening the excel file!\nDo you have it open in an another application",
            )
            Standard_Functions.Mbox(
                text=Text, title="TitleBlock Workbench", style=0
            )
            return

        try:
            # get the spreadsheet "TitleBlock"
            sheet = App.ActiveDocument.getObject("TitleBlock")
            # Clear the sheet
            sheet.clearAll()

            # Get the startcolumn and the other three columns from there
            StartCell = EXTERNAL_SOURCE_STARTCELL
            if EXTERNAL_SOURCE_SHEET_NAME == "":
                # Set EXTERNAL_SOURCE_SHEET_NAME to the chosen sheetname
                preferences.SetString("SheetName", Input_SheetName)
                ws = wb[Input_SheetName]

                # Set the startcell with an inputbox
                Text = translate(
                    "TitleBlock Workbench",
                    "Please enter the name of the cell.\n"
                    + "Enter a single cell like 'A1', 'B2', etc. Other notations will be ignored!",
                )
                StartCell = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=20,
                        default="A1",
                    )
                )
                if not StartCell.strip():
                    StartCell = "A1"

                # Set SHEETNAME_STARTCELL to the chosen sheetname
                preferences.SetString("StartCell", StartCell)

            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)
            StartColumn = StartCell[:1]
            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column is: " + str(StartColumn),
                        "Log",
                    )
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumn)),
                    ),
                    "Log",
                )
            Column2 = Standard_Functions.GetLetterFromNumber(
                int(Standard_Functions.GetNumberFromLetter(StartColumn) + 1), True
            )
            Column3 = Standard_Functions.GetLetterFromNumber(
                int(Standard_Functions.GetNumberFromLetter(StartColumn) + 2), True
            )
            Column4 = Standard_Functions.GetLetterFromNumber(
                int(Standard_Functions.GetNumberFromLetter(StartColumn) + 3), True
            )
            Column5 = Standard_Functions.GetLetterFromNumber(
                int(Standard_Functions.GetNumberFromLetter(StartColumn) + 4), True
            )

            # Get the start row
            StartRow = EXTERNAL_SOURCE_STARTCELL[1:2]
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row is: " + str(StartRow)
                )
                Standard_Functions.Print(Text, "Log")

            # import the headers from the excelsheet into the spreadsheet
            sheet.set("A1", str(ws[str(StartColumn) + str(StartRow)].value))
            sheet.set("B1", str(ws[str(Column2) + str(StartRow)].value))
            sheet.set("C1", str(ws[str(Column3) + str(StartRow)].value))
            sheet.set("D1", str(ws[str(Column4) + str(StartRow)].value))
            sheet.set("E1", str(ws[str(Column5) + str(StartRow)].value))

            # Go through the excel until the cell in the first column is empty.
            for i in range(1000):
                # Define the start row. This is the Header row +1 + i as counter
                RowNumber = int(StartRow) + i + 1

                # check if you reached the end of the data.
                if ws[str(StartColumn) + str(RowNumber)].value is None:
                    break

                # Get the number of row difference between the start row in the excelsheet
                # and the first row in the spreadsheet.
                # This to start at second row in the spreadsheet. (under the headers)
                Delta = int(StartRow) - 1

                # Fill the property name
                sheet.set(
                    str("A" + str(RowNumber - Delta)),
                    str(ws[str(StartColumn) + str(RowNumber)].value),
                )
                # Fill the property value
                if ws[Column2 + str(RowNumber)].value is not None:
                    sheet.set(
                        str("B" + str(RowNumber - Delta)),
                        str(ws[Column2 + str(RowNumber)].value),
                    )
                # Fill the value for auto increasement(yes or no)
                if ws[Column3 + str(RowNumber)].value is not None:
                    sheet.set(
                        str("C" + str(RowNumber - Delta)),
                        str(ws[Column3 + str(RowNumber)].value),
                    )
                # Fill the multipliers
                if ws[Column4 + str(RowNumber)].value is not None:
                    sheet.set(
                        str("D" + str(RowNumber - Delta)),
                        str(ws[Column4 + str(RowNumber)].value),
                    )
                # Fill the remarks
                if ws[Column5 + str(RowNumber)].value is not None:
                    sheet.set(
                        str("E" + str(RowNumber - Delta)),
                        str(ws[Column4 + str(RowNumber)].value),
                    )

            # Finally recompute the spreadsheet
            sheet.recompute()

            # Run the def to add extra system data.
            MapData(sheet=sheet)

            # Run the def to add document information
            MapDocInfo(sheet=sheet)

            # Run the def to add extra system data. This is the final value of "RowNumber" minus the "StartRow".
            AddExtraData(sheet, RowNumber - int(StartRow))

            # Format the spreadsheet
            extraRows = 0
            if INCLUDE_LENGTH is True:
                extraRows = extraRows + 1
            if INCLUDE_ANGLE is True:
                extraRows = extraRows + 1
            if INCLUDE_MASS is True:
                extraRows = extraRows + 1
            if INCLUDE_NO_SHEETS is True:
                extraRows = extraRows + 1

            # region Format the data with the values as a Table
            #
            # Define the header range
            StartCell = "A1"
            RemarkCell = "E1"
            HeaderRange = str(f"{StartCell}:{RemarkCell}")

            # Get the first row below the header
            FirstTableRow = ""
            for i in range(len(StartCell)):
                if StartCell[i].isdigit():
                    FirstTableRow = FirstTableRow + str(StartCell[i])
            FirstTableRow = int(FirstTableRow) + 1

            # Get the first column
            FirstColumn = Standard_Functions.RemoveNumbersFromString(StartCell)

            # Get the last column
            LastColumn = Standard_Functions.RemoveNumbersFromString(RemarkCell)

            # Define the table range
            TableRange = str(f"{FirstColumn}{FirstTableRow}:{LastColumn}{int(RowNumber) + extraRows}")

            # Define the First column range
            FirstColumnRange = str(f"{FirstColumn}{FirstTableRow}:{FirstColumn}{int(RowNumber) + extraRows}")

            # Format the table
            sheet = TableFormat_Functions.FormatTable(sheet=sheet, HeaderRange=HeaderRange,
                                                      TableRange=TableRange, FirstColumnRange=FirstColumnRange)

            # endregion

            # Finally recompute the spreadsheet
            sheet.recompute()
            App.ActiveDocument.recompute()

        except Exception as e:
            Text = translate(
                "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!\n"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: an error occurred!!\n"
                    + "See the report view for details",
                )
                raise e
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    else:
        Text = translate("TitleBlock Workbench", "External source is not enabled!")
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    return


# Import data from an central FreeCAD file
def ImportDataFreeCAD():
    # if debug mode is enabled, show the external file including path.
    if ENABLE_DEBUG is True:
        Text = translate("TitleBlock Workbench", str(EXTERNAL_SOURCE_PATH))
        Standard_Functions.Print(Text, "Log")

    try:
        # Check if it is allowed to use an external source and if so, continue
        if USE_EXTERNAL_SOURCE is True:
            # Get the active document
            doc = App.ActiveDocument
            # Get the name of the external source
            Input_SheetName = EXTERNAL_SOURCE_SHEET_NAME
            # get the spreadsheet "TitleBlock"
            sheet = doc.getObject("TitleBlock")
            # Clear the sheet
            sheet.clearAll()
            # Save the name of the active document to reactivate it at the end of this function.
            LastActiveDoc = doc.Name
            # Define the External sheet and document
            ExtSheet = None
            # ff = None

            # try to open the source. if not show an messagebox and if debug mode is enabled, show the exeption as well
            try:
                if EXTERNAL_SOURCE_PATH.lower().endswith(".xlsx"):
                    Filter = [("FreeCAD", "*.FCStd"), ]
                    FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=False)
                    if FileName != "":
                        ff = App.openDocument(FileName, True)
                    if FileName == "":
                        return
                else:
                    ff = App.openDocument(EXTERNAL_SOURCE_PATH, True)
                if EXTERNAL_SOURCE_SHEET_NAME == "":
                    # Set the sheetname with a inputbox
                    Spreadsheet_List = ff.findObjects('Spreadsheet::Sheet')
                    Text = translate(
                        "TitleBlock Workbench", "Please enter the name of the spreadsheet"
                    )
                    Input_SheetName = str(
                        Standard_Functions.Mbox(
                            text=Text,
                            title="TitleBlock Workbench",
                            style=21,
                            default="Settings",
                            stringList=Spreadsheet_List,
                        )
                    )
                    # if the user canceled, exit this function.
                    if not Input_SheetName.strip():
                        return

                    ExtSheet = ff.getObject(Input_SheetName)
                if EXTERNAL_SOURCE_SHEET_NAME != "":
                    ExtSheet = ff.getObject(Input_SheetName)
            except Exception as e:
                if ENABLE_DEBUG is True:
                    raise (e)
                Text = translate(
                    "TitleBlock Workbench",
                    "an problem occured while openening the FreeCAD file!\nDo you have it open in an another application",
                )
                Standard_Functions.Mbox(
                    text=Text, title="TitleBlock Workbench", style=0
                )
                return

            # Get the startcolumn and the other three columns from there
            StartCellExt = "A1"
            if EXTERNAL_SOURCE_SHEET_NAME == "":
                # Set EXTERNAL_SOURCE_SHEET_NAME to the chosen sheetname
                preferences.SetString("SheetName", Input_SheetName)
                ExtSheet = ff.getObject(Input_SheetName)
                # Set the startcell with an inputbox
                Text = translate(
                    "TitleBlock Workbench",
                    "Please enter the name of the cell.\n"
                    + "Enter a single cell like 'A1', 'B2', etc. Other notations will be ignored!",
                )
                StartCellExt = str(
                    Standard_Functions.Mbox(
                        text=Text,
                        title="TitleBlock Workbench",
                        style=20,
                        default="A1",
                    )
                )
                if not StartCellExt.strip():
                    StartCellExt = "A1"

                # Set SHEETNAME_STARTCELL to the chosen sheetname
                preferences.SetString("StartCell", StartCellExt)

            if (Standard_Functions.GetA1fromR1C1(StartCellExt)).strip():
                StartCellExt = Standard_Functions.GetA1fromR1C1(StartCellExt)
            StartColumnExt = Standard_Functions.RemoveNumbersFromString(StartCellExt)
            # If debug mode is on, show the start colum and its number
            if ENABLE_DEBUG is True:
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Start column is: " + str(StartColumnExt),
                        "Log",
                    )
                )
                Standard_Functions.Print(
                    translate(
                        "TitleBlock Workbench",
                        "Column number is: "
                        + str(Standard_Functions.GetNumberFromLetter(StartColumnExt)),
                    ),
                    "Log",
                )
            Column2 = Standard_Functions.GetLetterFromNumber(
                int(Standard_Functions.GetNumberFromLetter(StartColumnExt) + 1), True
            )
            Column3 = Standard_Functions.GetLetterFromNumber(
                int(Standard_Functions.GetNumberFromLetter(StartColumnExt) + 2), True
            )
            Column4 = Standard_Functions.GetLetterFromNumber(
                int(Standard_Functions.GetNumberFromLetter(StartColumnExt) + 3), True
            )
            Column5 = Standard_Functions.GetLetterFromNumber(
                int(Standard_Functions.GetNumberFromLetter(StartColumnExt) + 4), True
            )

            # Get the start row
            StartRow = Standard_Functions.RemoveLettersFromString(EXTERNAL_SOURCE_STARTCELL)
            # if debug mode is on, show your start row
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench", "the start row is: " + str(StartRow)
                )
                Standard_Functions.Print(Text, "Log")

            # import the headers from the excelsheet into the spreadsheet
            sheet.set("A1", str(ExtSheet.getContents(str(StartColumnExt) + str(StartRow))))
            sheet.set("B1", str(ExtSheet.getContents(str(Column2) + str(StartRow))))
            sheet.set("C1", str(ExtSheet.getContents(str(Column3) + str(StartRow))))
            sheet.set("D1", str(ExtSheet.getContents(str(Column4) + str(StartRow))))
            sheet.set("E1", str(ExtSheet.getContents(str(Column5) + str(StartRow))))

            # Go through the excel until the cell in the first column is empty.
            for i in range(1000):
                # Define the start row. This is the Header row +1 + i as counter
                RowNumber = int(StartRow) + i + 1

                # check if you reached the end of the data.
                if ExtSheet.getContents(str(StartColumnExt) + str(RowNumber)) is None:
                    break

                # Get the number of row difference between the start row in the excelsheet
                # and the first row in the spreadsheet.
                # This to start at second row in the spreadsheet. (under the headers)
                Delta = int(StartRow) - 1

                # Fill the property name
                sheet.set(
                    str("A" + str(RowNumber - Delta)),
                    str(ExtSheet.getContents(str(StartColumnExt) + str(RowNumber))),
                )
                # Fill the property value
                if ExtSheet.getContents(Column2 + str(RowNumber)) is not None:
                    sheet.set(
                        str("B" + str(RowNumber - Delta)),
                        str(ExtSheet.getContents(Column2 + str(RowNumber))),
                    )
                # Fill the value for auto increasement(yes or no)
                if ExtSheet.getContents(Column3 + str(RowNumber)) is not None:
                    sheet.set(
                        str("C" + str(RowNumber - Delta)),
                        str(ExtSheet.getContents(Column3 + str(RowNumber))),
                    )
                # Fill the multipliers
                if ExtSheet.getContents(Column4 + str(RowNumber)) is not None:
                    sheet.set(
                        str("D" + str(RowNumber - Delta)),
                        str(ExtSheet.getContents(Column4 + str(RowNumber))),
                    )
                # Fill the remarks
                if ExtSheet.getContents(Column5 + str(RowNumber)) is not None:
                    sheet.set(
                        str("E" + str(RowNumber - Delta)),
                        str(ExtSheet.getContents(Column4 + str(RowNumber))),
                    )

                # Check if the next row of the spreadsheet has data. If not this is the end of all the available values.
                try:
                    test = sheet.getContents("A" + str(RowNumber))
                    if test == "" or test is None:
                        break
                except Exception:
                    break

            # # Finally recompute the document
            # App.ActiveDocument.recompute(None, True, True)

            # Run the def to add extra system data.
            MapData(sheet=sheet, doc=doc)

            # Run the def to add document information
            MapDocInfo(sheet=sheet, doc=doc)

            # Run the def to add extra system data. This is the final value of "RowNumber" minus the "StartRow".
            AddExtraData(sheet, RowNumber - int(StartRow), doc)

            extraRows = 0
            if INCLUDE_LENGTH is True:
                extraRows = extraRows + 1
            if INCLUDE_ANGLE is True:
                extraRows = extraRows + 1
            if INCLUDE_MASS is True:
                extraRows = extraRows + 1
            if INCLUDE_NO_SHEETS is True:
                extraRows = extraRows + 1

            # region Format the data with the values as a Table
            #
            # Define the header range
            StartCell = "A1"
            RemarkCell = "E1"
            HeaderRange = str(f"{StartCell}:{RemarkCell}")

            # Get the first row below the header
            FirstTableRow = ""
            for i in range(len(StartCell)):
                if StartCell[i].isdigit():
                    FirstTableRow = FirstTableRow + str(StartCell[i])
            FirstTableRow = int(FirstTableRow) + 1

            # Get the first column
            FirstColumn = Standard_Functions.RemoveNumbersFromString(StartCell)

            # Get the last column
            LastColumn = Standard_Functions.RemoveNumbersFromString(RemarkCell)

            # Define the table range
            TableRange = str(f"{FirstColumn}{FirstTableRow}:{LastColumn}{int(RowNumber) + extraRows-1}")

            # Define the First column range
            FirstColumnRange = str(f"{FirstColumn}{FirstTableRow}:{FirstColumn}{int(RowNumber) + extraRows-1}")

            # Format the table
            sheet = TableFormat_Functions.FormatTable(sheet=sheet, HeaderRange=HeaderRange,
                                                      TableRange=TableRange, FirstColumnRange=FirstColumnRange)

            # endregion

            # recompute the document
            doc.recompute(None, True, True)
            # Save the workbook
            doc.save()
            # Close the FreeCAD file
            App.closeDocument(ff.Name)
            # Activate the document which was active when this command started.
            try:
                App.setActiveDocument(LastActiveDoc)
            except Exception:
                pass
        else:
            Text = translate("TitleBlock Workbench", "External source is not enabled!")
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!\n"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
            raise e
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
    return


# Use this function to run the function "Populate sheet"
def Start(command):
    try:
        sheet = App.ActiveDocument.getObject("TitleBlock")
        # check if the result is not empty
        if sheet is not None:
            # Proceed with the macro.
            if command == "FillSpreadsheet":
                FillSheet()
            if command == "ImportExcel":
                ImportDataExcel()
            if command == "ImportFreeCAD":
                ImportDataFreeCAD()

            # if the debug mode is on, report presense of titleblock spreadsheet
            if ENABLE_DEBUG is True:
                Text = translate("TitleBlock Workbench", "TitleBlock already present")
                Standard_Functions.Print(Text, "Log")

            return
        # if the result is empty, create a new titleblock spreadsheet
        if sheet is None:
            sheet = App.ActiveDocument.addObject("Spreadsheet::Sheet", "TitleBlock")

            # Proceed with the macro.
            if command == "FillSpreadsheet":
                FillSheet()
            if command == "ImportExcel":
                ImportDataExcel()
            if command == "ImportFreeCAD":
                ImportDataFreeCAD()

            # if the debug mode is on, report creation of titleblock spreadsheet
            if ENABLE_DEBUG is True:
                Text = translate("TitleBlock Workbench", "TitleBlock created")
                Standard_Functions.Print(Text, "Log")

            return
    except Exception as e:
        if ENABLE_DEBUG is True:
            raise e
    return
