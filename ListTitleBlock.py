# ***************************************************************************
# *   Copyright (c) 2015 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/


# This macro creates an Spreadsheet named "TitleBlock", reads editable texts from the first page
# and fills the spreadsheet.
# The data in spreadsheet can be changed by the user as long the data in column A remains unchanged.
# There will be three colums created in the spreadsheet:
#  - Property Name           -> the name of the editable text item
#  - Property Value          -> the value of the editable text item
#  - Increase, Yes or No?    -> do you want to increase the  value by 1 for the next sheet.
# there a multiple senarios for using these macro's:
#  - When changing the template for an different size, you can quickly refill the titleblock.
#    Of course, the editable text in the new template must be equal.
#  - Link the date to model properties like parameters. Basicly all what is achievable with the Spreadsheet workbench
#  - Import data from excel or libre office
#  - Etc.

import FreeCAD as App
import ctypes  # An included library with Python install.

# Get the settings
from Prefereneces import INCLUDE_LENGTH
from Prefereneces import INCLUDE_ANGLE
from Prefereneces import INCLUDE_MASS
from Prefereneces import INCLUDE_NO_SHEETS
from Prefereneces import USE_EXTERNAL_SOURCE
from Prefereneces import EXTERNAL_SOURCE_PATH
from Prefereneces import EXTERNAL_SOURCE_SHEET_NAME


#  Message Styles:
#  0 : OK
#  1 : OK | Cancel
#  2 : Abort | Retry | Ignore
#  3 : Yes | No | Cancel
#  4 : Yes | No
#  5 : Retry | Cancel
#  6 : Cancel | Try Again | Continuemport ctypes  # An included library with Python install.
def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


def AddExtraData(sheet, StartRow):
    # The following values can be added. You can use these for you specific templates.
    # Use the "bind function" of the spreadsheet workbench to create the correct entry for your template.:
    #   1. add the correct Property name
    #   2. bind the cell to the value of your specific template property.
    #   3. Set the increase value to "Yes" or "No".
    # Or use them directly if your templates has the exact fields

    print(str(INCLUDE_LENGTH))
    print(str(INCLUDE_ANGLE))
    print(str(INCLUDE_MASS))

    # get units scheme
    SchemeNumber = App.Units.getSchema()

    # Add the length units of your FreeCAD application
    if INCLUDE_LENGTH is True:
        sheet.set("A" + str(StartRow + 1), "Length_Units")
        sheet.set(
            "B" + str(StartRow + 1),
            str(
                App.Units.schemaTranslate(
                    App.Units.Quantity(50, App.Units.Length), SchemeNumber
                )
            ).split()[1],
        )
        sheet.set("C" + str(StartRow + 1), "No")
        StartRow = StartRow + 1

    # Add the angular units of your FreeCAD application
    if INCLUDE_ANGLE is True:
        sheet.set("A" + str(StartRow + 1), "Angle_Units")
        sheet.set(
            "B" + str(StartRow + 1),
            str(
                App.Units.schemaTranslate(
                    App.Units.Quantity(50, App.Units.Angle), SchemeNumber
                )
            ).split()[1],
        )
        sheet.set("C" + str(StartRow + 1), "No")
        StartRow = StartRow + 1

    # Add the mass units of your FreeCAD application
    if INCLUDE_MASS is True:
        sheet.set("A" + str(StartRow + 1), "Mass_Units")
        sheet.set(
            "B" + str(StartRow + 1),
            str(
                App.Units.schemaTranslate(
                    App.Units.Quantity(50, App.Units.Mass), SchemeNumber
                )
            ).split()[1],
        )
        sheet.set("C" + str(StartRow + 1), "No")
        StartRow = StartRow + 1

    # Add the total number of sheets. You can use this for your title block
    if INCLUDE_NO_SHEETS is True:
        pages = App.ActiveDocument.findObjects("TechDraw::DrawPage")
        sheet.set("A" + str(StartRow + 1), "Number of sheets")
        sheet.set("B" + str(StartRow + 1), str(len(pages)))
        sheet.set("C" + str(StartRow + 1), "No")

    return


# Fill the spreadsheet with all the date from the titleblock
def FillSheet():
    # get the fist page
    page = App.ActiveDocument.Page
    # get the editable texts
    texts = page.Template.EditableTexts
    # get the spreadsheet "TitleBlock"
    sheet = App.ActiveDocument.getObject("TitleBlock")

    # set the headers in the spreadsheet
    sheet.set("A1", "Property Name")
    sheet.set("B1", "Property Value")
    sheet.set("C1", "Increase, Yes or No?")

    # set the start value for the start row.
    # (x=0, the spreadsheet whill be populated from the first row. the headers will be overwritten)
    StartRow = 1
    for key, value in texts.items():
        # Increase StartRow by one, to fill the next row
        StartRow = StartRow + 1
        # Fill the property name
        sheet.set("A" + str(StartRow), "{0}".format(key, value))
        # Fill the property value
        sheet.set("B" + str(StartRow), "{1}".format(key, value))
        # If there is no value yet, the increase function will be set to "No" of by default.
        try:
            str(sheet.get("C" + str(StartRow)))
        except Exception:
            sheet.set("C" + str(StartRow), "No")

    # Run the def to add extra data
    AddExtraData(sheet, StartRow)


# Import data from a (central) excel workbook
def ImportDataExcel():
    from openpyxl import load_workbook

    print(str(EXTERNAL_SOURCE_PATH))

    # Check if it is allowed to use an external source and if so, continue
    if USE_EXTERNAL_SOURCE is True:
        wb = load_workbook(str(EXTERNAL_SOURCE_PATH))
        ws = wb[str(EXTERNAL_SOURCE_SHEET_NAME)]

        # get the spreadsheet "TitleBlock"
        sheet = App.ActiveDocument.getObject("TitleBlock")

        # import the headers from the excelsheet into the spreadsheet
        sheet.set("A1", str(ws["A1"].value))
        sheet.set("B1", str(ws["B1"].value))
        sheet.set("C1", str(ws["C1"].value))
        sheet.set("D1", str(ws["D1"].value))

        # set the start value for the start row.
        # (x=0, the spreadsheet whill be populated from the first row. the headers will be overwritten)
        StartRow = 1
        for i in range(1000):
            # check if you reached the end of the data.
            if ws["A" + str(StartRow)].value is None:
                break

            # Increase StartRow by one, to fill the next row
            StartRow = StartRow + 1
            # Fill the property name
            sheet.set("A" + str(StartRow), str(ws["A" + str(StartRow)].value))
            # Fill the property value
            if ws["B" + str(StartRow)].value is not None:
                sheet.set("B" + str(StartRow), str(ws["B" + str(StartRow)].value))
            # Fill the value for auto increasement(yes or no)
            if ws["C" + str(StartRow)].value is not None:
                sheet.set("C" + str(StartRow), str(ws["C" + str(StartRow)].value))
            # Fill the remarks
            if ws["D" + str(StartRow)].value is not None:
                sheet.set("D" + str(StartRow), str(ws["D" + str(StartRow)].value))

        # Run the def to add extra data
        AddExtraData(sheet, StartRow)

    else:
        Mbox("", "External source is not enabled!", 0)


def Start(command):
    # Find or Create spreadsheet
    try:
        # check if there is already an spreadsheet called "TitleBlock"
        sheet = App.ActiveDocument.getObject("TitleBlock")
        print("TitleBlock already present")
        # Proceed with the macro.
        if command == "ListTitleBlock":
            FillSheet()
        if command == "ImportExcel":
            ImportDataExcel()
    except Exception:
        # if there is not yet an spreadsheet called "TitleBlock", create one
        sheet = App.ActiveDocument.addObject("Spreadsheet::Sheet", "TitleBlock")
        print("TitleBlock created")
        # set the label to "TitleBlock"
        sheet.Label = "TitleBlock"
        # Proceed with the macro.
        if command == "ListTitleBlock":
            FillSheet()
        if command == "ImportExcel":
            ImportDataExcel()
