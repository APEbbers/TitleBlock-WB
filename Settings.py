# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/

import FreeCAD as App
import Standard_Functions_TitleBlock as Standard_Functions
import TableFormat_Functions
import os

# region defenitions
# Define the translation
translate = App.Qt.translate

preferences = App.ParamGet(
    "User parameter:BaseApp/Preferences/Mod/TitleBlock Workbench"
)
# endregion


# region -- functions to make sure that a None type result is ""
def GetStringSetting(settingName: str) -> str:
    result = preferences.GetString(settingName)

    if result.lower() == "none":
        result = ""
    return result


def GetIntSetting(settingName: str) -> int:
    result = preferences.GetInt(settingName)
    if result == "":
        result = None
    return result


def GetFloatSetting(settingName: str) -> int:
    result = preferences.GetFloat(settingName)
    if result == "":
        result = None
    return result


def GetBoolSetting(settingName: str) -> bool:
    result = preferences.GetBool(settingName)
    if str(result).lower() == "none":
        result = False
    return result


def GetColorSetting(settingName: str) -> object:
    from PySide2.QtGui import QColor

    # Create a tuple from the int value of the color
    result = QColor.fromRgba(preferences.GetUnsigned(settingName)).toTuple()

    # correct the order of the tuple and devide them by 255
    result = (result[3] / 255, result[0] / 255, result[1] / 255, result[2] / 255)

    return result


def SetStringSetting(settingName: str, value: str):
    Text = translate(
        "TitleBlock Workbench",
        f"string setting not applied!!\n Settings was: {settingName} and value was {value}",
    )
    if value.lower() == "none":
        if ENABLE_DEBUG is True:
            Standard_Functions.Print(Text, "Log")
            value = ""
    preferences.SetString(settingName, value)


def SetBoolSetting(settingName: str, value):
    if value.lower() == "true":
        Bool = True
    if str(value).lower() == "none" or value.lower() != "true":
        Text = translate(
            "TitleBlock Workbench",
            f"bool setting not applied!!\n Settings was: {settingName} and value was {value}",
        )
        if ENABLE_DEBUG is True:
            Standard_Functions.Print(Text, "Log")
        Bool = False
    preferences.SetBool(settingName, Bool)


# endregion


# region -- All settings from the UI

# TechDraw settings
ADD_TOOLBAR_TECHDRAW = GetBoolSetting("AddToolBarTechDraw")
IMPORT_EXAMPLE_TEMPLATES = GetBoolSetting("Import_templates")
DEFAULT_TEMPLATE = GetIntSetting("Default_Template")

# External source
USE_EXTERNAL_SOURCE = GetBoolSetting("UseExternalSource")
EXTERNAL_SOURCE_PATH = GetStringSetting("ExternalFile")
EXTERNAL_SOURCE_SHEET_NAME = GetStringSetting("SheetName")
EXTERNAL_SOURCE_STARTCELL = GetStringSetting("StartCell")
AUTOFILL_TITLEBLOCK = GetBoolSetting("AutoFillTitleBlock")
IMPORT_SETTINGS_XL = GetBoolSetting("ImportSettingsXL")
SHEETNAME_SETTINGS_XL = GetStringSetting("SheetName_Settings")
SHEETNAME_STARTCELL_XL = GetStringSetting("StartCell_Settings")

# Use filename as drawingnumber
USE_FILENAME_DRAW_NO = GetBoolSetting("UseFileName")
DRAW_NO_FIELD = GetStringSetting("DrwNrFieldName")
<<<<<<< HEAD
=======

# Use pagename as drawing number
USE_PAGENAME_DRAW_NO = GetBoolSetting("UsePageName")
DRAW_NO_FIELD_PAGE = GetStringSetting("DrwNrFieldName_Page")
>>>>>>> Develop

# The values that are mapped
MAP_LENGTH = GetStringSetting("MapLength")
MAP_ANGLE = GetStringSetting("MapAngle")
MAP_MASS = GetStringSetting("MapMass")
MAP_NOSHEETS = GetStringSetting("MapNoSheets")

# Document information that are mapped
DOCINFO_NAME = GetStringSetting("DocInfo_Name")
DOCINFO_CREATEDBY = GetStringSetting("DocInfo_CreatedBy")
DOCINFO_CREATEDDATE = GetStringSetting("DocInfo_CreatedDate")
DOCINFO_LASTMODIFIEDBY = GetStringSetting("DocInfo_LastModifiedBy")
DOCINFO_LASTMODIFIEDDATE = GetStringSetting("DocInfo_LastModifiedDate")
DOCINFO_COMPANY = GetStringSetting("DocInfo_Company")
DOCINFO_LICENSE = GetStringSetting("DocInfo_License")
DOCINFO_LICENSEURL = GetStringSetting("DocInfo_LicenseURL")
DOCINFO_COMMENT = GetStringSetting("DocInfo_Comment")

# Included values
INCLUDE_LENGTH = GetBoolSetting("IncludeLength")
INCLUDE_ANGLE = GetBoolSetting("IncludeAngle")
INCLUDE_MASS = GetBoolSetting("IncludeMass")
INCLUDE_NO_SHEETS = GetBoolSetting("IncludeNoOfSheets")

# UI Settings
SPREADSHEET_HEADERBACKGROUND = GetColorSetting("SpreadSheetHeaderBackGround")
SPREADSHEET_HEADERFOREGROUND = GetColorSetting("SpreadSheetHeaderForeGround")
SPREADSHEET_HEADERFONTSTYLE_BOLD = GetBoolSetting("SpreadsheetHeaderFontStyle_Bold")
SPREADSHEET_HEADERFONTSTYLE_ITALIC = GetBoolSetting("SpreadsheetHeaderFontStyle_Italic")
SPREADSHEET_HEADERFONTSTYLE_UNDERLINE = GetBoolSetting(
    "SpreadsheetHeaderFontStyle_Underline"
)
SPREADSHEET_TABLEBACKGROUND_1 = GetColorSetting("SpreadSheetTableBackGround_1")
SPREADSHEET_TABLEBACKGROUND_2 = GetColorSetting("SpreadSheetTableBackGround_2")
SPREADSHEET_TABLEFOREGROUND = GetColorSetting("SpreadSheetTableForeGround")
SPREADSHEET_TABLEFONTSTYLE_BOLD = GetBoolSetting("SpreadsheetTableFontStyle_Bold")
SPREADSHEET_TABLEFONTSTYLE_ITALIC = GetBoolSetting("SpreadsheetTableFontStyle_Italic")
SPREADSHEET_TABLEFONTSTYLE_UNDERLINE = GetBoolSetting(
    "SpreadsheetTableFontStyle_Underline"
)
SPREADSHEET_COLUMNFONTSTYLE_BOLD = GetBoolSetting("SpreadsheetColumnFontStyle_Bold")
SPREADSHEET_COLUMNFONTSTYLE_ITALIC = GetBoolSetting("SpreadsheetColumnFontStyle_Italic")
SPREADSHEET_COLUMNFONTSTYLE_UNDERLINE = GetBoolSetting(
    "SpreadsheetColumnFontStyle_Underline"
)
AUTOFIT_FACTOR = GetFloatSetting("AutoFitFactor")

# Enable debug mode. This will enable additional report messages
ENABLE_DEBUG = GetBoolSetting("EnableDebug")

# All the settings in a List
SettingsList = [
    USE_EXTERNAL_SOURCE,
    EXTERNAL_SOURCE_PATH,
    EXTERNAL_SOURCE_SHEET_NAME,
    EXTERNAL_SOURCE_STARTCELL,
    AUTOFILL_TITLEBLOCK,
    IMPORT_SETTINGS_XL,
    SHEETNAME_SETTINGS_XL,
    SHEETNAME_STARTCELL_XL,
    USE_FILENAME_DRAW_NO,
    DRAW_NO_FIELD,
    USE_PAGENAME_DRAW_NO,
    DRAW_NO_FIELD_PAGE,
    MAP_LENGTH,
    MAP_ANGLE,
    MAP_MASS,
    MAP_NOSHEETS,
    DOCINFO_NAME,
    DOCINFO_CREATEDBY,
    DOCINFO_CREATEDDATE,
    DOCINFO_LASTMODIFIEDBY,
    DOCINFO_LASTMODIFIEDDATE,
    DOCINFO_LICENSEURL,
    DOCINFO_COMMENT,
    DOCINFO_COMPANY,
    DOCINFO_LICENSE,
    INCLUDE_LENGTH,
    INCLUDE_ANGLE,
    INCLUDE_MASS,
    INCLUDE_NO_SHEETS,
    SPREADSHEET_HEADERBACKGROUND,
    SPREADSHEET_HEADERFOREGROUND,
    SPREADSHEET_HEADERFONTSTYLE_BOLD,
    SPREADSHEET_HEADERFONTSTYLE_ITALIC,
    SPREADSHEET_HEADERFONTSTYLE_UNDERLINE,
    SPREADSHEET_TABLEBACKGROUND_1,
    SPREADSHEET_TABLEBACKGROUND_2,
    SPREADSHEET_HEADERFOREGROUND,
    SPREADSHEET_TABLEFONTSTYLE_BOLD,
    SPREADSHEET_TABLEFONTSTYLE_ITALIC,
    SPREADSHEET_TABLEFONTSTYLE_UNDERLINE,
    SPREADSHEET_COLUMNFONTSTYLE_BOLD,
    SPREADSHEET_COLUMNFONTSTYLE_ITALIC,
    SPREADSHEET_COLUMNFONTSTYLE_UNDERLINE,
    ENABLE_DEBUG,
]


Extenstion = EXTERNAL_SOURCE_PATH.rsplit(".", 1)[0]

# endregion


def ExportSettings_FreeCAD(Silent=False):
    try:
        # Get the name of the current document and use it at the end of this function to reactivate this document
        doc = App.ActiveDocument
        LastActiveDoc = doc.Name

        # Define sheet and freecad file (ff)
        sheet = None
        ff = None

        # region -- Get the workbook, create a new sheet and set the startcell (top left cell of table).
        # If the user wants to export the settins, start an input dialog.
        if Silent is False:
            # load the excel file with the custom function
            if USE_EXTERNAL_SOURCE is False:
                Filter = [
                    ("FreeCAD", "*.FCStd"),
                ]
                FileName = Standard_Functions.GetFileDialog(Filter, False)
                if FileName != "":
                    ff = App.openDocument(FileName, True)
                    preferences.SetString("ExternalFile", FileName)
                if FileName == "":
                    return
            if USE_EXTERNAL_SOURCE is True:
                if (
                    Standard_Functions.CheckIfFreeCADfileExists(EXTERNAL_SOURCE_PATH, True)
                    is True
                ):
                    ff = App.openDocument(EXTERNAL_SOURCE_PATH, True)
                else:
                    Text = translate(
                        "TitleBlock Workbench",
                        f"TitleBlock Workbench: Something went wrong with loading {EXTERNAL_SOURCE_PATH}",
                    )
                    Standard_Functions.Print(Text, "Error")
                    return

            # Set the sheetname with a inputbox.
            # Create a list with the current spreadsheets for input in the inputdialog.
            # Make sure that "Settings" is on that list.
            Spreadsheet_List = []
            # Add "Settings" to the new list.
            Spreadsheet_List.append("Settings")
            # Get the sheet objects and go through them.
            sheets = ff.findObjects('Spreadsheet::Sheet')
            for i in range(len(sheets)):
                # If the sheet is not named "Settings", add it to the list.
                if sheets[i].Name != "Settings":
                    Spreadsheet_List.append(sheets[i].Name)
            # Show the inputdialog
            Text = translate("TitleBlock Workbench", "Please enter the name of the spreadsheet")
            Input_SheetName = str(
                Standard_Functions.Mbox(
                    text=Text,
                    title="TitleBlock Workbench",
                    style=21,
                    default="Settings",
                    stringList=Spreadsheet_List,
                )
            )
            # if the user canceled, exit this function.
            if not Input_SheetName.strip():
                return

            # Set SHEETNAME_SETTINGS_XL to the chosen sheetname
            preferences.SetString("SheetName_Settings", Input_SheetName)

            # Delete the current sheet if it exists
            sheets = ff.findObjects('Spreadsheet::Sheet')
            for i in range(len(sheets)):
                # If the sheet is not named "Settings", add it to the list.
                if sheets[i].Name == Input_SheetName:
                    ff.removeObject(Input_SheetName)

            # create a new sheet
            sheet = ff.addObject("Spreadsheet::Sheet", Input_SheetName)

            # Set the startcell
            StartCell = SHEETNAME_STARTCELL_XL
            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)
                if ENABLE_DEBUG is True:
                    Text = translate(
                        "TitleBlock Workbench",
                        f"TitleBlock Workbench: the startcell converted from {SHEETNAME_STARTCELL_XL} to {StartCell}",
                    )
                    Standard_Functions.Print(Text, "Log")

            # Set SHEETNAME_STARTCELL_XL to the chosen sheetname
            preferences.SetString("StartCell_Settings", StartCell)

        # If a new excel file is created and the sittings must be imported from that excel (IMPORT_SETTINGS_XL = True),
        # Load the workbook with the sheet from the preference menu.
        if Silent is True:
            try:
                if (
                    Standard_Functions.CheckIfFreeCADfileExists(
                        EXTERNAL_SOURCE_PATH, False
                    )
                    is True
                ):
                    ff = App.openDocument(EXTERNAL_SOURCE_PATH, True)
                else:
                    Text = translate(
                        "TitleBlock Workbench",
                        f"TitleBlock Workbench: FreeCAD file didn't exist!. ({EXTERNAL_SOURCE_PATH})",
                    )
                    Standard_Functions.Print(Text, "Error")
                    return
            except Exception:
                return
            StartCell = SHEETNAME_STARTCELL_XL
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: Spreadsheetname and startcell for the settings is: "
                    + f"{SHEETNAME_SETTINGS_XL}, {SHEETNAME_STARTCELL_XL}",
                )
                Standard_Functions.Print(Text, "Log")
        # endregion

        # region -- Create the headers
        # Define the sheet. If it not exists, exit this function
        sheet = ff.getObject(SHEETNAME_SETTINGS_XL)
        if sheet is None:
            if Silent is False:
                Standard_Functions.Print(f"This FreeCAD file doesn't have a spreadsheet named {SHEETNAME_SETTINGS_XL}")
                return
            if Silent is True:
                # Create a spreadsheet for the titleblock data.
                sheet = ff.addObject("Spreadsheet::Sheet", SHEETNAME_SETTINGS_XL)

        # Set the header for the first column
        sheet.set(StartCell, "Name")
        # Set the header for the second column
        TopRow = int(StartCell[1:])
        ValueCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
            )
        ) + str(TopRow)
        sheet.set(ValueCell, "Value")
        # endregion

        # region -- Export the external source settings
        #
        # USE_EXTERNAL_SOURCE
        RowNumber = 1
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "UseExternalSource")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(USE_EXTERNAL_SOURCE).upper())
        RowNumber = RowNumber + 1

        # EXTERNAL_SOURCE_PATH
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "ExternalFile")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(EXTERNAL_SOURCE_PATH))
        RowNumber = RowNumber + 1

        # EXTERNAL_SOURCE_SHEET_NAME
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "SheetName")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(EXTERNAL_SOURCE_SHEET_NAME))
        RowNumber = RowNumber + 1

        # EXTERNAL_SOURCE_STARTCELL
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "StartCell")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(EXTERNAL_SOURCE_STARTCELL))
        RowNumber = RowNumber + 1

        # AUTOFILL_TITLEBLOCK
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "AutoFillTitleBlock")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(AUTOFILL_TITLEBLOCK).upper())
        RowNumber = RowNumber + 1

        # IMPORT_SETTINGS_XL
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "ImportSettingsXL")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(IMPORT_SETTINGS_XL).upper())
        RowNumber = RowNumber + 1

        # SHEETNAME_SETTINGS_XL
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "SheetName_Settings")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(SHEETNAME_SETTINGS_XL))
        RowNumber = RowNumber + 1

        # SHEETNAME_STARTCELL_XL
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "StartCell_Settings")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(SHEETNAME_STARTCELL_XL))
        RowNumber = RowNumber + 1

        # endregion

        # region -- Export the filename settings
        #
        # USE_FILENAME_DRAW_NO
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "UseFileName")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(USE_FILENAME_DRAW_NO).upper())
        RowNumber = RowNumber + 1

        # DRAW_NO_FiELD
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DrwNrFieldName")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DRAW_NO_FIELD))
<<<<<<< HEAD
=======
        RowNumber = RowNumber + 1

        # endregion

        # region -- Export the pagename settings
        #
        # USE_PAGENAME_DRAW_NO
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "UsePageName")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(USE_PAGENAME_DRAW_NO).upper())
        RowNumber = RowNumber + 1

        # DRAW_NO_FiELD_PAGE
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DrwNrFieldName_Page")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DRAW_NO_FIELD_PAGE))
>>>>>>> Develop
        RowNumber = RowNumber + 1

        # endregion

        # region -- Export the Mapping settings
        #
        # MAP_LENGTH
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "MapLength")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(MAP_LENGTH))
        RowNumber = RowNumber + 1

        # MAP_ANGLE
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "MapAngle")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(MAP_ANGLE))
        RowNumber = RowNumber + 1

        # MAP_MASS
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "MapMass")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(MAP_MASS))
        RowNumber = RowNumber + 1

        # MAP_NOSHEETS
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "MapNoSheets")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(MAP_NOSHEETS))
        RowNumber = RowNumber + 1
        # endregion

        # region -- Map the document information settings
        #
        # DOCINFO_NAME
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_Name")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_NAME))
        RowNumber = RowNumber + 1

        # DOCINFO_CREATEDBY
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_CreatedBy")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_CREATEDBY))
        RowNumber = RowNumber + 1

        # DOCINFO_CREATEDDATE
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_CreatedDate")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_CREATEDDATE))
        RowNumber = RowNumber + 1

        # DOCINFO_LASTMODIFIEDBY
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_LastModifiedBy")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_LASTMODIFIEDBY))
        RowNumber = RowNumber + 1

        # DOCINFO_NAME
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_LastModifiedDate")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_LASTMODIFIEDDATE))
        RowNumber = RowNumber + 1

        # DOCINFO_COMPANY
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_Company")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_COMPANY))
        RowNumber = RowNumber + 1

        # DOCINFO_LICENSE
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_License")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_LICENSE))
        RowNumber = RowNumber + 1

        # DOCINFO_LICENSEURL
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_LicenseURL")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_LICENSEURL))
        RowNumber = RowNumber + 1

        # DOCINFO_COMMENT
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "DocInfo_Comment")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(DOCINFO_COMMENT))
        RowNumber = RowNumber + 1
        # endregion

        # region -- Export the included value settings
        #
        # INCLUDE_LENGTH
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "IncludeLength")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(INCLUDE_LENGTH).upper())
        RowNumber = RowNumber + 1

        # INCLUDE_ANGLE
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "IncludeAngle")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(INCLUDE_ANGLE).upper())
        RowNumber = RowNumber + 1

        # INCLUDE_MASS
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "IncludeMass")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(INCLUDE_MASS).upper())
        RowNumber = RowNumber + 1

        # INCLUDE_NO_SHEETS
        sheet.set(str(StartCell[:1] + str(TopRow + RowNumber)), "IncludeNoOfSheets")
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        sheet.set(SettingValue, str(INCLUDE_NO_SHEETS).upper())
        RowNumber = RowNumber + 1
        # endregion

        # Notes:
        # - ENABLE_DEBUG is excluded from export.
        #   This is a setting only needed for debuggin and thus a per user setting.
        # - Spreadsheet format settings are excluded from export.
        #   This is a per user setting

        # region Format the settings with the values as a Table
        #
        # Define the header range
        HeaderRange = str(f"{StartCell}:{ValueCell}")

        # Get the first row below the header
        FirstTableRow = ""
        for i in range(len(StartCell)):
            if StartCell[i].isdigit():
                FirstTableRow = FirstTableRow + str(StartCell[i])
        FirstTableRow = int(FirstTableRow) + 1

        # Get the first column
        FirstColumn = Standard_Functions.RemoveNumbersFromString(StartCell)

        # Get the last column
        LastColumn = Standard_Functions.RemoveNumbersFromString(ValueCell)

        # Define the table range
<<<<<<< HEAD
        TableRange = str(f"{FirstColumn}{FirstTableRow}:{LastColumn}{RowNumber + 1}")

        # Define the First column range
        FirstColumnRange = str(f"{FirstColumn}{FirstTableRow}:{FirstColumn}{RowNumber + 1}")
=======
        TableRange = str(f"{FirstColumn}{FirstTableRow}:{LastColumn}{RowNumber}")

        # Define the First column range
        FirstColumnRange = str(f"{FirstColumn}{FirstTableRow}:{FirstColumn}{RowNumber}")
>>>>>>> Develop

        # Format the table
        sheet = TableFormat_Functions.FormatTable(sheet=sheet, HeaderRange=HeaderRange,
                                                  TableRange=TableRange, FirstColumnRange=FirstColumnRange)
        # endregion

        # recompute the document
        ff.recompute(None, True, True)
        # Save the workbook
        ff.save()
        # Close the FreeCAD file
        App.closeDocument(ff.Name)
        # Activate the document which was active when this command started.
        try:
            App.setActiveDocument(LastActiveDoc)
        except Exception:
            pass
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            raise (e)


def ImportSettings_FreeCAD():
    import os.path
    import errno

    # Get the name of the current document and use it at the end of this function to reactivate this document
    doc = App.ActiveDocument
    LastActiveDoc = doc.Name

    # Get the workbook. If it doesn't exist. Let the user now.
    ff = None
    if USE_EXTERNAL_SOURCE is False:
        Filter = [
            ("FreeCAD", "*.FCStd"),
        ]
        FileName = Standard_Functions.GetFileDialog(files=Filter, SaveAs=False)
        if FileName != "":
            ff = App.openDocument(FileName, True)
        if FileName == "":
            return
    if USE_EXTERNAL_SOURCE is True:
        if os.path.exists(EXTERNAL_SOURCE_PATH) is True:
            ff = App.openDocument(EXTERNAL_SOURCE_PATH, True)
        if os.path.exists(EXTERNAL_SOURCE_PATH) is False:
            Text = translate(
                "TitleBlock Workbench",
                "There is no FreeCAD file available, while import from external source is enabled!\n"
                + "Please create an FreeCAD file to export your settings to or disable import from external source.",
                "TitleBlock Workbench",
            )
            Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
            return

    try:
        # Get the sheetname
        ExtSheet = ff.getObject(SHEETNAME_SETTINGS_XL)
        if ExtSheet is None:
            Text = translate(
                "TitleBlock Workbench", f"No spreadsheet named '{SHEETNAME_SETTINGS_XL}'!!!"
            )
            Standard_Functions.Print(Input=Text, Type="Warning")
            return

        # Get the startcell
        StartCell = SHEETNAME_STARTCELL_XL
        OriginalStartCell = StartCell
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench", f"The entered startcell is: {StartCell}"
            )
            Standard_Functions.Print(Text, "Log")
        if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
            StartCell = Standard_Functions.GetA1fromR1C1(StartCell)
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    f"TitleBlock Workbench: the startcell converted from {OriginalStartCell} to {StartCell}",
                )
                Standard_Functions.Print(Text, "Log")

        # Get the columns
        FirstColumn = Standard_Functions.RemoveNumbersFromString(StartCell)
        SecondColumn = Standard_Functions.GetLetterFromNumber(Standard_Functions.GetNumberFromLetter(FirstColumn) + 1)

        # Get the first table row
        FirstTableRow = int(Standard_Functions.RemoveLettersFromString(StartCell)) + 1

        # go through the excel until all settings are imported.
        counter = 0
        for i in range(FirstTableRow, 1000):
            Cell_Name = ExtSheet.getContents(str(FirstColumn) + str(i))
            if Cell_Name.startswith("'"):
                Cell_Name = Cell_Name[1:]
            Cell_Value = ExtSheet.getContents(str(SecondColumn) + str(i))
            if Cell_Value.startswith("'"):
                Cell_Value = Cell_Value[1:]

            # region -- Import the external source settings
            #
            # Import USE_EXTERNAL_SOURCE
            if Cell_Name == "UseExternalSource":
                SetBoolSetting("UseExternalSource", Cell_Value)
                counter = counter + 1

            # This is desabled because the preference must be leading at all time.
            # The are not allowed to be overridden by importing the settings.
            # -----------------------------------------------------------
            # Import EXTERNAL_SOURCE_PATH
            # if Cell_Name == "ExternalFile":
            #     SetStringSetting("ExternalFile", str(Cell_Value))
            #     counter = counter + 1

            # Import EXTERNAL_SOURCE_SHEET_NAME
            # if Cell_Name == "SheetName":
            #     SetStringSetting("SheetName", str(Cell_Value))
            #     counter = counter + 1

            # Import EXTERNAL_SOURCE_STARTCELL
            # if Cell_Name == "StartCell":
            #     SetStringSetting("StartCell", str(Cell_Value))
            #     counter = counter + 1
            # -----------------------------------------------------------

            # Import AUTOFILL_TITLEBLOCK
            if Cell_Name == "AutoFillTitleBlock":
                SetBoolSetting("AutoFillTitleBlock", Cell_Value)
                counter = counter + 1

            # Import IMPORT_SETTINGS_XL
            if Cell_Name == "ImportSettingsXL":
                SetBoolSetting("ImportSettingsXL", Cell_Value)
                counter = counter + 1

            # Import SHEETNAME_SETTINGS_XL
            if Cell_Name == "SheetName_Settings":
                SetStringSetting("SheetName_Settings", str(Cell_Value))
                counter = counter + 1

            # Import SHEETNAME_STARTCELL_XL
            if Cell_Name == "StartCell_Settings":
                SetStringSetting("StartCell_Settings", str(Cell_Value))
                counter = counter + 1

            # endregion

            # region -- Import the filename settings
            #
            # Import USE_FILENAME_DRAW_NO
            if Cell_Name == "UseFileName":
                SetBoolSetting("UseFileName", Cell_Value)
                counter = counter + 1

            # Import DRAW_NO_FiELD
            if Cell_Name == "DrwNrFieldName":
                SetStringSetting("DrwNrFieldName", str(Cell_Value))
                counter = counter + 1

            # endregion

            # region -- Import the pagename settings
            #
            # Import USE_PAGENAME_DRAW_NO
            if Cell_Name == "UsePageName":
                SetBoolSetting("UsePageName", Cell_Value)
                counter = counter + 1

            # Import DRAW_NO_FiELD_PAGE
            if Cell_Name == "DrwNrFieldName_Page":
                SetStringSetting("DrwNrFieldName_Page", str(Cell_Value))
                counter = counter + 1

            # endregion

            # region -- Import the mapping settings
            #
            # Import MAP_LENGTH
            if Cell_Name == "MapLength":
                SetStringSetting("MapLength", str(Cell_Value))
                counter = counter + 1

            # Import MAP_ANGLE
            if Cell_Name == "MapAngle":
                SetStringSetting("MapAngle", str(Cell_Value))
                counter = counter + 1

            # Import MAP_MASS
            if Cell_Name == "MapMass":
                SetStringSetting("MapMass", str(Cell_Value))
                counter = counter + 1

            # Import MAP_NOSHEETS
            if Cell_Name == "MapNoSheets":
                SetStringSetting("MapNoSheets", str(Cell_Value))
                counter = counter + 1

            # endregion

            # region -- Import the document information settings
            #
            # Import DOCINFO_NAME
            if Cell_Name == "DocInfo_Name":
                SetStringSetting("DocInfo_Name", str(Cell_Value))
                counter = counter + 1

            # DOCINFO_CREATEDBY
            if Cell_Name == "DocInfo_CreatedBy":
                SetStringSetting("DocInfo_CreatedBy", str(Cell_Value))
                counter = counter + 1

            # DOCINFO_CREATEDDATE
            if Cell_Name == "DocInfo_CreatedDate":
                SetStringSetting("DocInfo_CreatedDate", str(Cell_Value))
                counter = counter + 1

            # DOCINFO_LASTMODIFIEDBY
            if Cell_Name == "DocInfo_LastModifiedBy":
                SetStringSetting("DocInfo_LastModifiedBy", str(Cell_Value))
                counter = counter + 1

            # DOCINFO_LASTMODIFIEDDATE
            if Cell_Name == "DocInfo_LastModifiedDate":
                SetStringSetting("DocInfo_LastModifiedDate", str(Cell_Value))
                counter = counter + 1

            # DOCINFO_COMPANY
            if Cell_Name == "DocInfo_Company":
                SetStringSetting("DocInfo_Company", str(Cell_Value))
                counter = counter + 1

            # DOCINFO_LICENSE
            if Cell_Name == "DocInfo_License":
                SetStringSetting("DocInfo_License", str(Cell_Value))
                counter = counter + 1

            # DOCINFO_LICENSEURL
            if Cell_Name == "DocInfo_LicenseURL":
                SetStringSetting("DocInfo_LicenseURL", str(Cell_Value))
                counter = counter + 1

            # DOCINFO_COMMENT
            if Cell_Name == "DocInfo_Comment":
                SetStringSetting("DocInfo_Comment", str(Cell_Value))
                counter = counter + 1

            # endregion

            # region -- Import the Include settings
            #
            # Import INCLUDE_LENGTH
            if Cell_Name == "IncludeLength":
                SetBoolSetting("IncludeLength", Cell_Value)
                counter = counter + 1

            # Import INCLUDE_ANGLE
            if Cell_Name == "IncludeAngle":
                SetBoolSetting("IncludeAngle", Cell_Value)
                counter = counter + 1

            # Import INCLUDE_MASS
            if Cell_Name == "IncludeMass":
                SetBoolSetting("IncludeMass", Cell_Value)
                counter = counter + 1

            # Import INCLUDE_NO_SHEETS
            if Cell_Name == "IncludeNoOfSheets":
                SetBoolSetting("IncludeNoOfSheets", Cell_Value)
                counter = counter + 1

            # endregion

            # Notes:
            # - ENABLE_DEBUG is excluded from export.
            #   This is a setting only needed for debuggin and thus a per user setting.
            # - Spreadsheet format settings are excluded from export.
            #   This is a per user setting

            if Cell_Name == "":
                break

            if counter == len(SettingsList) + 1:
                break

        if counter > 0:
            Text = translate(
                "TitleBlock Workbench",
                "Titleblock workbench: Settings imported from "
                + f"{EXTERNAL_SOURCE_PATH} from worksheet: {EXTERNAL_SOURCE_SHEET_NAME} at {StartCell}",
            )
            Standard_Functions.Print(Text, "Log")
        if counter == 0:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: Settings are not imported",
            )
            Standard_Functions.Print(Text, "Log")

        # recompute the document
        ff.recompute(None, True, True)
        # Save the workbook
        ff.save()
        # Close the FreeCAD file
        App.closeDocument(ff.Name)
        # Activate the document which was active when this command started.
        App.setActiveDocument(LastActiveDoc)
    # If there is an IO Error continue:
    except IOError as e:
        # If there is an read/write error, print an error in the report view
        if e.errno == errno.EACCES:
            Text = translate(
                "TitleBlock Workbench", f"No permision to open {EXTERNAL_SOURCE_PATH}!"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    f"No permision to open {EXTERNAL_SOURCE_PATH}!\nSee the report view for details",
                )

            # Close the FreeCAD file
            App.closeDocument(ff.Name)
            return Standard_Functions.Mbox(
                text=Text, title="TitleBlock Workbench", style=0
            )
        # For all other IO errors, raise e.
        # Close the FreeCAD file
        App.closeDocument(ff.Name)
        raise (e)
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            # Close the FreeCAD file
            App.closeDocument(ff.Name)
            raise (e)
        # Close the FreeCAD file
        App.closeDocument(ff.Name)
        return


def ExportSettings_XL(Silent=False):
    import openpyxl.utils.exceptions
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo

    try:
        # region -- Get the workbook, create a new sheet and set the startcell (top left cell of table).
        # If the user wants to export the settins, start an input dialog.
        if Silent is False:
            # load the excel file with the custm function
            if USE_EXTERNAL_SOURCE is False:
                Filter = [
                    ("Excel", "*.xlsx"),
                ]
                FileName = Standard_Functions.GetFileDialog(Filter, False)
                if FileName != "":
                    wb = load_workbook(FileName)
                    preferences.SetString("ExternalFile", FileName)
                if FileName == "":
                    return
            if USE_EXTERNAL_SOURCE is True:
                if (
                    Standard_Functions.CheckIfWorkbookExists(EXTERNAL_SOURCE_PATH, True)
                    is True
                ):
                    wb = load_workbook(EXTERNAL_SOURCE_PATH)
                else:
                    Text = translate(
                        "TitleBlock Workbench",
                        f"TitleBlock Workbench: Something went wrong with loading {EXTERNAL_SOURCE_PATH}",
                    )
                    Standard_Functions.Print(Text, "Error")
                    return

            # Set the sheetname with a inputbox
            Worksheets_List = [i for i in wb.sheetnames if i != "TitleBlockData"]
            Text = translate(
                "TitleBlock Workbench", "Please enter the name of the worksheet"
            )
            Input_SheetName = str(
                Standard_Functions.Mbox(
                    text=Text,
                    title="TitleBlock Workbench",
                    style=21,
                    default="Settings",
                    stringList=Worksheets_List,
                )
            )
            # if the user canceled, exit this function.
            if not Input_SheetName.strip():
                return

            # Set SHEETNAME_SETTINGS_XL to the chosen sheetname
            preferences.SetString("SheetName_Settings", Input_SheetName)

            # Delete the current sheet if it exists
            for sheetname in wb.sheetnames:
                if sheetname == Input_SheetName:
                    del wb[str(Input_SheetName)]
                    break

            # create a new sheet
            ws = wb.create_sheet(title=Input_SheetName)

            # Set the startcell with an inputbox
            Text = translate(
                "TitleBlock Workbench",
                "Please enter the name of the cell.\n"
                + "Enter a single cell like 'A1', 'B2', etc. Other notations will be ignored!",
            )
            StartCell = str(
                Standard_Functions.Mbox(
                    text=Text, title="TitleBlock Workbench", style=20, default="A1"
                )
            )
            if StartCell == "" or StartCell is None:
                StartCell = "A1"

            # Set SHEETNAME_STARTCELL_XL to the chosen sheetname
            preferences.SetString("StartCell_Settings", StartCell)

        # If a new excel file is created and the sittings must be imported from that excel (IMPORT_SETTINGS_XL = True),
        # Load the workbook with the sheet from the preference menu.
        if Silent is True:
            try:
                if (
                    Standard_Functions.CheckIfWorkbookExists(
                        EXTERNAL_SOURCE_PATH, False
                    )
                    is True
                ):
                    wb = load_workbook(str(EXTERNAL_SOURCE_PATH))
                else:
                    Text = translate(
                        "TitleBlock Workbench",
                        f"TitleBlock Workbench: Workbook didn't exist!. ({EXTERNAL_SOURCE_PATH})",
                    )
                    Standard_Functions.Print(Text, "Error")
                    return
            except Exception:
                return
            ws = wb.create_sheet(SHEETNAME_SETTINGS_XL)
            StartCell = SHEETNAME_STARTCELL_XL
            if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
                StartCell = Standard_Functions.GetA1fromR1C1(StartCell)
                if ENABLE_DEBUG is True:
                    Text = translate(
                        "TitleBlock Workbench",
                        f"TitleBlock Workbench: the startcell converted from {SHEETNAME_STARTCELL_XL} to {StartCell}",
                    )
                    Standard_Functions.Print(Text, "Log")
            if StartCell is None or StartCell == "":
                StartCell = "A1"
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    "TitleBlock Workbench: Sheetname and startcell for the settings is: "
                    + f"{SHEETNAME_SETTINGS_XL}, {SHEETNAME_STARTCELL_XL}",
                )
                Standard_Functions.Print(Text, "Log")
        # endregion

        # region -- Create the headers
        ws[StartCell].value = "Name"
        TopRow = int(StartCell[1:])
        ValueCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
            )
        ) + str(TopRow)
        ws[ValueCell].value = "Value"
        # endregion

        # region -- Export the external source settings
        #
        # USE_EXTERNAL_SOURCE
        RowNumber = 1
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "UseExternalSource"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(USE_EXTERNAL_SOURCE).upper()
        RowNumber = RowNumber + 1

        # EXTERNAL_SOURCE_PATH
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "ExternalFile"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = EXTERNAL_SOURCE_PATH
        RowNumber = RowNumber + 1

        # EXTERNAL_SOURCE_SHEET_NAME
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "SheetName"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = EXTERNAL_SOURCE_SHEET_NAME
        RowNumber = RowNumber + 1

        # EXTERNAL_SOURCE_STARTCELL
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "StartCell"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = EXTERNAL_SOURCE_STARTCELL
        RowNumber = RowNumber + 1

        # AUTOFILL_TITLEBLOCK
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "AutoFillTitleBlock"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(AUTOFILL_TITLEBLOCK).upper()
        RowNumber = RowNumber + 1

        # IMPORT_SETTINGS_XL
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "ImportSettingsXL"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(IMPORT_SETTINGS_XL).upper()
        RowNumber = RowNumber + 1

        # SHEETNAME_SETTINGS_XL
        ws[str(StartCell[:1] + str(TopRow + RowNumber))] = "SheetName_Settings"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = SHEETNAME_SETTINGS_XL
        RowNumber = RowNumber + 1

        # SHEETNAME_STARTCELL_XL
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "StartCell_Settings"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = SHEETNAME_STARTCELL_XL
        RowNumber = RowNumber + 1

        # endregion

        # region -- Export the filename settings
        #
        # USE_FILENAME_DRAW_NO
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "UseFileName"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(USE_FILENAME_DRAW_NO).upper()
        RowNumber = RowNumber + 1

        # DRAW_NO_FiELD
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DrwNrFieldName"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DRAW_NO_FIELD
<<<<<<< HEAD
=======
        RowNumber = RowNumber + 1

        # endregion

        # region -- Export the pagename settings
        #
        # USE_FILENAME_DRAW_NO
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "UsePageName"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(USE_PAGENAME_DRAW_NO).upper()
        RowNumber = RowNumber + 1

        # DRAW_NO_FiELD
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DrwNrFieldName_Page"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DRAW_NO_FIELD_PAGE
>>>>>>> Develop
        RowNumber = RowNumber + 1

        # endregion

        # region -- Export the Mapping settings
        #
        # MAP_LENGTH
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "MapLength"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = MAP_LENGTH
        RowNumber = RowNumber + 1

        # MAP_ANGLE
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "MapAngle"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = MAP_ANGLE
        RowNumber = RowNumber + 1

        # MAP_MASS
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "MapMass"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = MAP_MASS
        RowNumber = RowNumber + 1

        # MAP_NOSHEETS
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "MapNoSheets"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = MAP_NOSHEETS
        RowNumber = RowNumber + 1
        # endregion

        # region -- Map the document information settings
        #
        # DOCINFO_NAME
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DocInfo_Name"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_NAME
        RowNumber = RowNumber + 1

        # DOCINFO_CREATEDBY
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DocInfo_CreatedBy"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_CREATEDBY
        RowNumber = RowNumber + 1

        # DOCINFO_CREATEDDATE
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DocInfo_CreatedDate"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_CREATEDDATE
        RowNumber = RowNumber + 1

        # DOCINFO_LASTMODIFIEDBY
        ws[
            str(StartCell[:1] + str(TopRow + RowNumber))
        ].value = "DocInfo_LastModifiedBy"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_LASTMODIFIEDBY
        RowNumber = RowNumber + 1

        # DOCINFO_NAME
        ws[
            str(StartCell[:1] + str(TopRow + RowNumber))
        ].value = "DocInfo_LastModifiedDate"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_LASTMODIFIEDDATE
        RowNumber = RowNumber + 1

        # DOCINFO_COMPANY
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DocInfo_Company"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_COMPANY
        RowNumber = RowNumber + 1

        # DOCINFO_LICENSE
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DocInfo_License"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_LICENSE
        RowNumber = RowNumber + 1

        # DOCINFO_LICENSEURL
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DocInfo_LicenseURL"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_LICENSEURL
        RowNumber = RowNumber + 1

        # DOCINFO_COMMENT
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "DocInfo_Comment"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        ws[SettingValue].value = DOCINFO_COMMENT
        RowNumber = RowNumber + 1
        # endregion

        # region -- Export the included value settings
        #
        # INCLUDE_LENGTH
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "IncludeLength"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(INCLUDE_LENGTH).upper()
        RowNumber = RowNumber + 1

        # INCLUDE_ANGLE
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "IncludeAngle"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(INCLUDE_ANGLE).upper()
        RowNumber = RowNumber + 1

        # INCLUDE_MASS
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "IncludeMass"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(INCLUDE_MASS).upper()
        RowNumber = RowNumber + 1

        # INCLUDE_NO_SHEETS
        ws[str(StartCell[:1] + str(TopRow + RowNumber))].value = "IncludeNoOfSheets"
        # Write the value
        SettingValue = str(ValueCell[:1] + str(TopRow + RowNumber))
        # Create a dropdown for the boolan
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[SettingValue])
        ws[SettingValue].value = str(INCLUDE_NO_SHEETS).upper()
        RowNumber = RowNumber + 1
        # endregion

        # Note: ENABLE_DEBUG is excluded from export.
        # This is a setting only needed for debuggin and thus a per user setting.

        # region Format the settings with the values as a Table
        #
        # Define the the last cell
        EndCell = str(
            Standard_Functions.GetLetterFromNumber(
                Standard_Functions.GetNumberFromLetter(StartCell[:1]) + 1
            )
        ) + str(RowNumber + TopRow - 1)

        # Define the table
        NumberOfSheets = str(len(wb.sheetnames))
        tab = Table(
            displayName="SettingsTable_" + NumberOfSheets,
            ref=f"{StartCell}:{EndCell}",
        )

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium11",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )

        # add the style to the table
        tab.tableStyleInfo = style

        # add the table to the worksheet
        ws.add_table(tab)

        # Align the columns
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                f"TitleBlock Workbench: Table range is: {StartCell}:{EndCell}",
            )
            Standard_Functions.Print(Text, "Log")
        for row in ws[1: ws.max_row]:
            Column_1 = row[Standard_Functions.GetNumberFromLetter(StartCell[:1]) - 1]
            Column_2 = row[Standard_Functions.GetNumberFromLetter(EndCell[:1]) - 1]
            Column_1.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            Column_2.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
        # endregion

        # Make the columns to autofit the date
        for col in ws.columns:
            SetLen = 0
            column = col[0].column_letter  # Get the column name

            for cell in col:
                if len(str(cell.value)) > SetLen:
                    SetLen = len(str(cell.value))

            set_col_width = SetLen + 5
            # Setting the column width
            ws.column_dimensions[column].width = set_col_width

        # Save the workbook
        wb.save(EXTERNAL_SOURCE_PATH)

        # Close the workbook
        wb.close()
    except openpyxl.utils.exceptions.ReadOnlyWorkbookException as e:
        Text = translate("TitleBlock Workbench", "The excel file is read only!")
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            raise (e)
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            raise (e)


def ImportSettings_XL():
    from openpyxl import load_workbook
    import os.path
    import errno

    try:
        # Get the workbook. If it doesn't exist. Let the user now.
        wb = None
        if USE_EXTERNAL_SOURCE is False:
            Filter = [
                ("Excel", "*.xlsx"),
            ]
            FileName = Standard_Functions.GetFileDialog(Filter, False)
            if FileName != "":
                wb = load_workbook(str(FileName), read_only=True)
            if FileName == "":
                return
        if USE_EXTERNAL_SOURCE is True:
            if os.path.exists(EXTERNAL_SOURCE_PATH) is True:
                wb = load_workbook(str(EXTERNAL_SOURCE_PATH), read_only=True)
            if os.path.exists(EXTERNAL_SOURCE_PATH) is False:
                Text = translate(
                    "TitleBlock Workbench",
                    "There is no excel workbook available, while import from external source is enabled!\n" +
                    "Please create an excel workbook to export your settings to or disable import from external source.",
                    "TitleBlock Workbench",)
                Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
                return

        # Get the sheetname
        counter = 0
        for sheetname in wb.sheetnames:
            if sheetname == SHEETNAME_SETTINGS_XL:
                ws = wb[str(SHEETNAME_SETTINGS_XL)]
                counter = 1
        if counter == 0:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: The sheet didn't exists when trying to import the settings!\n"
                + f"The sheetname should be {SHEETNAME_SETTINGS_XL}",
            )
            Standard_Functions.Print(Input=Text, Type="Warning")
            return

        # Get the startcell
        StartCell = SHEETNAME_STARTCELL_XL
        if StartCell == "" or StartCell is None:
            StartCell = "A1"
        OriginalStartCell = StartCell
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench", f"The entered startcell is: {StartCell}"
            )
            Standard_Functions.Print(Text, "Log")
        if (Standard_Functions.GetA1fromR1C1(StartCell)).strip():
            StartCell = Standard_Functions.GetA1fromR1C1(StartCell)
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    f"TitleBlock Workbench: the startcell converted from {OriginalStartCell} to {StartCell}",
                )
                Standard_Functions.Print(Text, "Log")

        # Get the columns
        FirstColumn = int(Standard_Functions.GetNumberFromLetter(StartCell[:1]))
        SecondColumn = FirstColumn + 1

        # go through the excel until all settings are imported.
        counter = 0

        for i in range(1, 1000):
            Cell_Name = ws.cell(i, FirstColumn)
            Cell_Value = ws.cell(i, SecondColumn)

            # region -- Import the external source settings
            #
            # Import USE_EXTERNAL_SOURCE
            if Cell_Name.value == "UseExternalSource":
                SetBoolSetting("UseExternalSource", Cell_Value.value)
                counter = counter + 1

            # This is desabled because the preference must be leading at all time.
            # The are not allowed to be overridden by importing the settings.
            # -----------------------------------------------------------
            # Import EXTERNAL_SOURCE_PATH
            # if Cell_Name.value == "ExternalFile":
            #     SetStringSetting("ExternalFile", str(Cell_Value.value))
            #     counter = counter + 1

            # Import EXTERNAL_SOURCE_SHEET_NAME
            # if Cell_Name.value == "SheetName":
            #     SetStringSetting("SheetName", str(Cell_Value.value))
            #     counter = counter + 1

            # Import EXTERNAL_SOURCE_STARTCELL
            # if Cell_Name.value == "StartCell":
            #     SetStringSetting("StartCell", str(Cell_Value.value))
            #     counter = counter + 1
            # -----------------------------------------------------------

            # Import AUTOFILL_TITLEBLOCK
            if Cell_Name.value == "AutoFillTitleBlock":
                SetBoolSetting("AutoFillTitleBlock", Cell_Value.value)
                counter = counter + 1

            # Import IMPORT_SETTINGS_XL
            if Cell_Name.value == "ImportSettingsXL":
                SetBoolSetting("ImportSettingsXL", Cell_Value.value)
                counter = counter + 1

            # Import SHEETNAME_SETTINGS_XL
            if Cell_Name.value == "SheetName_Settings":
                SetStringSetting("SheetName_Settings", str(Cell_Value.value))
                counter = counter + 1

            # Import SHEETNAME_STARTCELL_XL
            if Cell_Name.value == "StartCell_Settings":
                SetStringSetting("StartCell_Settings", str(Cell_Value.value))
                counter = counter + 1

            # endregion

            # region -- Import the filename settings
            #
            # Import USE_FILENAME_DRAW_NO
            if Cell_Name.value == "UseFileName":
                SetBoolSetting("UseFileName", Cell_Value.value)
                counter = counter + 1

            # Import DRAW_NO_FiELD
            if Cell_Name.value == "DrwNrFieldName":
                SetStringSetting("DrwNrFieldName", str(Cell_Value.value))
                counter = counter + 1

            # endregion

            # region -- Import the pagename settings
            #
            # Import USE_FILENAME_DRAW_NO
            if Cell_Name.value == "UsePageName":
                SetBoolSetting("UsePageName", Cell_Value.value)
                counter = counter + 1

            # Import DRAW_NO_FiELD
            if Cell_Name.value == "DrwNrFieldName_Page":
                SetStringSetting("DrwNrFieldName_Page", str(Cell_Value.value))
                counter = counter + 1

            # endregion

            # region -- Import the mapping settings
            #
            # Import MAP_LENGTH
            if Cell_Name.value == "MapLength":
                SetStringSetting("MapLength", str(Cell_Value.value))
                counter = counter + 1

            # Import MAP_ANGLE
            if Cell_Name.value == "MapAngle":
                SetStringSetting("MapAngle", str(Cell_Value.value))
                counter = counter + 1

            # Import MAP_MASS
            if Cell_Name.value == "MapMass":
                SetStringSetting("MapMass", str(Cell_Value.value))
                counter = counter + 1

            # Import MAP_NOSHEETS
            if Cell_Name.value == "MapNoSheets":
                SetStringSetting("MapNoSheets", str(Cell_Value.value))
                counter = counter + 1

            # endregion

            # region -- Import the document information settings
            #
            # Import DOCINFO_NAME
            if Cell_Name.value == "DocInfo_Name":
                SetStringSetting("DocInfo_Name", str(Cell_Value.value))
                counter = counter + 1

            # DOCINFO_CREATEDBY
            if Cell_Name.value == "DocInfo_CreatedBy":
                SetStringSetting("DocInfo_CreatedBy", str(Cell_Value.value))
                counter = counter + 1

            # DOCINFO_CREATEDDATE
            if Cell_Name.value == "DocInfo_CreatedDate":
                SetStringSetting("DocInfo_CreatedDate", str(Cell_Value.value))
                counter = counter + 1

            # DOCINFO_LASTMODIFIEDBY
            if Cell_Name.value == "DocInfo_LastModifiedBy":
                SetStringSetting("DocInfo_LastModifiedBy", str(Cell_Value.value))
                counter = counter + 1

            # DOCINFO_LASTMODIFIEDDATE
            if Cell_Name.value == "DocInfo_LastModifiedDate":
                SetStringSetting("DocInfo_LastModifiedDate", str(Cell_Value.value))
                counter = counter + 1

            # DOCINFO_COMPANY
            if Cell_Name.value == "DocInfo_Company":
                SetStringSetting("DocInfo_Company", str(Cell_Value.value))
                counter = counter + 1

            # DOCINFO_LICENSE
            if Cell_Name.value == "DocInfo_License":
                SetStringSetting("DocInfo_License", str(Cell_Value.value))
                counter = counter + 1

            # DOCINFO_LICENSEURL
            if Cell_Name.value == "DocInfo_LicenseURL":
                SetStringSetting("DocInfo_LicenseURL", str(Cell_Value.value))
                counter = counter + 1

            # DOCINFO_COMMENT
            if Cell_Name.value == "DocInfo_Comment":
                SetStringSetting("DocInfo_Comment", str(Cell_Value.value))
                counter = counter + 1

            # endregion

            # region -- Import the Include settings
            #
            # Import INCLUDE_LENGTH
            if Cell_Name.value == "IncludeLength":
                SetBoolSetting("IncludeLength", Cell_Value.value)
                counter = counter + 1

            # Import INCLUDE_ANGLE
            if Cell_Name.value == "IncludeAngle":
                SetBoolSetting("IncludeAngle", Cell_Value.value)
                counter = counter + 1

            # Import INCLUDE_MASS
            if Cell_Name.value == "IncludeMass":
                SetBoolSetting("IncludeMass", Cell_Value.value)
                counter = counter + 1

            # Import INCLUDE_NO_SHEETS
            if Cell_Name.value == "IncludeNoOfSheets":
                SetBoolSetting("IncludeNoOfSheets", Cell_Value.value)
                counter = counter + 1

            # endregion

            # Note: ENABLE_DEBUG is excluded from import.
            # This is a setting only needed for debuggin and thus a per user setting.

            if Cell_Name == "":
                break

            if counter == len(SettingsList) + 1:
                break

        if counter > 0:
            Text = translate(
                "TitleBlock Workbench",
                "Titleblock workbench: Settings imported from "
                + f"{EXTERNAL_SOURCE_PATH} from worksheet: {sheetname} at {StartCell}",
            )
            Standard_Functions.Print(Text, "Log")
        if counter == 0:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: Settings are not imported",
            )
            Standard_Functions.Print(Text, "Log")
    # If there is an IO Error continue:
    except IOError as e:
        # If there is an read/write error, print an error in the report view
        if e.errno == errno.EACCES:
            Text = translate(
                "TitleBlock Workbench", f"No permision to open {EXTERNAL_SOURCE_PATH}!"
            )
            if ENABLE_DEBUG is True:
                Text = translate(
                    "TitleBlock Workbench",
                    f"No permision to open {EXTERNAL_SOURCE_PATH}!\nSee the report view for details",
                )
            return Standard_Functions.Mbox(
                text=Text, title="TitleBlock Workbench", style=0
            )
        # For all other IO errors, raise e.
        raise (e)
    except Exception as e:
        Text = translate(
            "TitleBlock Workbench", "TitleBlock Workbench: an error occurred!!"
        )
        if ENABLE_DEBUG is True:
            Text = translate(
                "TitleBlock Workbench",
                "TitleBlock Workbench: an error occurred!!\n"
                + "See the report view for details",
            )
        Standard_Functions.Mbox(text=Text, title="TitleBlock Workbench", style=0)
        if ENABLE_DEBUG is True:
            raise (e)
        return

    # Close the workbook
    wb.close()
